"""Combined risk scoring + attendance-query routing + Attendance Watch
action — covering items 6–12 of the Phase H test list.

Items 1–5 (attendance ingestion, matching, metrics, risk threshold) live in
tests/test_attendance.py.
Item 13 (existing tests stay green) is covered by the full pytest run.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.combined_risk import attach_combined_risk
from core.confirmed_actions import execute_academic_watch_action
from core.risk_settings import RiskSettings
from core.query_engine import run_query
from nlp.planner_router import plan_user_request
from nlp.request_intents import is_attendance_watch_request


# ---- shared synthetic roster ----------------------------------------------


@pytest.fixture
def enriched_roster() -> pd.DataFrame:
    """A 10-student fixture with roster + attendance + risk columns merged."""
    return pd.DataFrame({
        "Student ID": [f"S{i:03d}" for i in range(10)],
        "Name": list("ABCDEFGHIJ"),
        "Department": ["Biology"] * 5 + ["Math"] * 5,
        "Teacher": ["Smith", "Smith", "Jones", "Jones", "Lee",
                    "Smith", "Lee", "Lee", "Jones", "Lee"],
        "GPA": [3.5, 1.8, 2.2, 1.5, 3.9, 2.7, 1.7, 2.0, 1.9, 3.1],
        "Academic Standing": ["Good Standing"] * 8 + ["Probation", "Warning"],
        "Attendance Rate": [98.0, 85.0, 92.0, 75.0, 96.0,
                            91.0, 78.0, 88.0, 95.0, 70.0],
        "Days Absent": [2, 6, 4, 10, 1, 3, 9, 5, 2, 12],
        "Attendance Risk": [False, True, False, True, False,
                            False, True, True, False, True],
    })


@pytest.fixture
def sheets(enriched_roster):
    return {"Students": enriched_roster}


@pytest.fixture
def sheet_columns(enriched_roster):
    return {"Students": list(enriched_roster.columns)}


def _route(message: str, sheets, sheet_columns):
    return plan_user_request(
        user_message=message, sheets=sheets, sheet_columns=sheet_columns,
        selected_sheet="Students", conversation_state=None,
        settings={"llm_enabled": False},
    )


# ---- Test 12: combined risk logic ------------------------------------------


def test_combined_risk_attaches_signals_and_level():
    """attach_combined_risk fold: 2+ signals → High Risk, 1 → Moderate, 0 → Low.

    Attendance Rate as a 0-1 fraction, matching every real roster's actual
    convention (and attach_combined_risk's own comparison, which converts the
    percentage-scale threshold to match) -- not a 0-100 percentage.
    """
    df = pd.DataFrame({
        "Student ID": ["A", "B", "C", "D"],
        "GPA": [3.9, 1.5, 1.5, 3.0],
        "Attendance Rate": [0.98, 0.95, 0.70, 0.88],
        "Academic Standing": ["Good Standing", "Good Standing",
                              "Probation", "Good Standing"],
    })
    out = attach_combined_risk(df).set_index("Student ID")
    # A: no signals → Low.
    assert out.loc["A", "Risk Signals"] == 0
    assert out.loc["A", "Risk Level"] == "Low Risk"
    # B: GPA only → Moderate.
    assert out.loc["B", "Risk Signals"] == 1
    assert out.loc["B", "Risk Level"] == "Moderate Risk"
    # C: GPA + Attendance + Standing → High.
    assert out.loc["C", "Risk Signals"] == 3
    assert out.loc["C", "Risk Level"] == "High Risk"
    # D: Attendance only (88 < 90) → Moderate.
    assert out.loc["D", "Risk Signals"] == 1
    assert out.loc["D", "Risk Level"] == "Moderate Risk"


def test_attendance_risk_does_not_flag_every_row():
    # Regression: Attendance Rate is a 0-1 fraction; the threshold is a 0-100
    # percentage. Comparing them directly without scaling meant every row
    # satisfied "< 90" regardless of actual attendance, silently flagging
    # 100% of any roster as attendance-risk (confirmed on real rosters this
    # session: 300/300 and 250/250 before this fix).
    df = pd.DataFrame({
        "Student ID": ["A", "B", "C"],
        "Attendance Rate": [0.99, 0.95, 0.85],
    })
    out = attach_combined_risk(df).set_index("Student ID")
    assert out.loc["A", "Attendance Risk"] == False
    assert out.loc["B", "Attendance Risk"] == False
    assert out.loc["C", "Attendance Risk"] == True


def test_risk_settings_defaults_are_correct():
    settings = RiskSettings()
    assert settings.gpa_risk_threshold == 2.0
    assert settings.attendance_risk_threshold == 90.0
    assert settings.severe_attendance_risk_threshold == 80.0
    assert settings.unexcused_absence_concern == 3
    assert settings.tardy_concern == 5
    assert settings.high_risk_signal_count == 2
    assert settings.moderate_risk_signal_count == 1


def test_risk_settings_mention_thresholds():
    settings = RiskSettings()
    assert "GPA below 2" in settings.mention_gpa_risk()
    assert "Attendance Rate below 90%" in settings.mention_attendance_risk()


def test_combined_risk_skips_missing_columns():
    """A roster without Attendance Rate or Standing should still get a Risk
    Level derived from whatever signals do exist."""
    df = pd.DataFrame({
        "Student ID": ["A", "B"],
        "GPA": [3.5, 1.5],
    })
    out = attach_combined_risk(df).set_index("Student ID")
    assert "GPA Risk" in out.columns
    assert "Attendance Risk" not in out.columns
    assert out.loc["A", "Risk Level"] == "Low Risk"
    assert out.loc["B", "Risk Level"] == "Moderate Risk"


# ---- Test 6: "students at attendance risk" routes ---------------------------


def test_who_is_at_attendance_risk_routes_to_filter(sheets, sheet_columns):
    routing = _route("who is at attendance risk", sheets, sheet_columns)
    plan = routing["plan"]
    assert routing["intent"] == "query"
    assert any(
        f.get("column") == "Attendance Risk" and f.get("value") is True
        for f in plan.get("filters") or []
    )
    result = run_query(plan, sheets)
    # 5 students flagged True in the fixture.
    assert result.row_count == 5


def test_attendance_below_threshold_filter(sheets, sheet_columns):
    routing = _route("show students with attendance below 90%",
                     sheets, sheet_columns)
    plan = routing["plan"]
    assert any(
        f.get("column") == "Attendance Rate" and f.get("operator") == "less_than"
        for f in plan.get("filters") or []
    )


# ---- Test 7: combined GPA + attendance filter ------------------------------


def test_gpa_and_attendance_combined_filter(sheets, sheet_columns):
    """The two numeric clauses should both land — exercises the
    AND-split fix in _detect_filters."""
    routing = _route(
        "students with GPA below 2.0 and attendance below 90%",
        sheets, sheet_columns,
    )
    plan = routing["plan"]
    cols = {f.get("column") for f in plan.get("filters") or []}
    assert "GPA" in cols
    assert "Attendance Rate" in cols
    assert plan.get("filter_mode", "all") == "all"
    result = run_query(plan, sheets)
    # Fixture students with GPA<2 AND attendance<90: S001 (1.8, 85), S003
    # (1.5, 75), S006 (1.7, 78), S008 (1.9, 95) excluded — 3 rows.
    assert result.row_count == 3


# ---- Test 8: teachers with most attendance-risk students -------------------


def test_teachers_with_most_attendance_risk_students_groups_by_teacher(
    sheets, sheet_columns,
):
    routing = _route(
        "which teachers have the most attendance-risk students",
        sheets, sheet_columns,
    )
    plan = routing["plan"]
    assert plan["operation"] == "groupby_count"
    assert plan["group_by"] == "Teacher"
    # Each group's count comes back filtered to attendance-risk students.
    result = run_query(plan, sheets)
    # Lee + Jones each have 2 of the 5 risk students; Smith has 1.
    counts = {row["Teacher"]: row["Count"] for row in result.table}
    # Lee owns 3 risk students (indices 6, 7, 9), Jones owns 1 (index 3),
    # Smith owns 1 (index 1) — total 5.
    assert counts == {"Lee": 3, "Jones": 1, "Smith": 1}


# ---- Test 9: Mark Attendance Watch is gated on confirmation ----------------


def test_attendance_watch_request_is_detected_separately_from_academic_watch():
    assert is_attendance_watch_request("mark these students attendance watch")
    assert is_attendance_watch_request("put them on attendance watch")
    # Academic-watch phrasings should NOT trip the attendance detector.
    assert not is_attendance_watch_request("mark these students academic watch")
    assert not is_attendance_watch_request("flag these students")


def test_attendance_watch_routing_requires_confirmation(sheets, sheet_columns):
    routing = _route("mark these students attendance watch", sheets, sheet_columns)
    assert routing["intent"] == "attendance_watch"
    assert routing["requires_confirmation"] is True
    assert routing["plan"]["column_name"] == "Attendance Watch"


# ---- Test 10 + 11: confirm writes new workbook; original unchanged ---------


def test_attendance_watch_writes_new_workbook_and_preserves_original(
    enriched_roster, tmp_path,
):
    sheets = {"Students": enriched_roster}
    original_snapshot = enriched_roster.copy()
    filters = [{"column": "Attendance Risk", "operator": "equals", "value": True}]
    result = execute_academic_watch_action(
        filters=filters, sheets=sheets, sheet="Students",
        column_name="Attendance Watch", value="Yes",
        request_summary="mark these students attendance watch",
        output_dir=tmp_path / "outputs",
        audit_path=tmp_path / "audit.jsonl",
    )
    assert result.success
    assert result.action_type == "academic_watch"
    # 5 attendance-risk students in the fixture.
    assert result.rows_affected == 5
    # Original DataFrame is unchanged.
    pd.testing.assert_frame_equal(sheets["Students"], original_snapshot)
    # New workbook exists and the column was added there.
    output_path = Path(result.output_file)
    assert output_path.exists()
    wb = load_workbook(output_path)
    ws = wb["Students"]
    headers = [cell.value for cell in ws[1]]
    assert "Attendance Watch" in headers
    assert "Attendance Watch Reason" in headers
    # And the audit log was appended.
    audit = tmp_path / "audit.jsonl"
    assert audit.exists()
    assert audit.read_text().strip()


def test_attendance_watch_does_not_clobber_academic_watch_column(
    enriched_roster, tmp_path,
):
    """Marking Attendance Watch must NOT touch any Academic Watch column."""
    df = enriched_roster.copy()
    df["Academic Watch"] = ""
    sheets = {"Students": df}
    filters = [{"column": "Attendance Risk", "operator": "equals", "value": True}]
    result = execute_academic_watch_action(
        filters=filters, sheets=sheets, sheet="Students",
        column_name="Attendance Watch", value="Yes",
        output_dir=tmp_path / "outputs",
        audit_path=tmp_path / "audit.jsonl",
    )
    output = pd.read_excel(result.output_file, sheet_name="Students")
    # Academic Watch column is untouched (all blank), Attendance Watch is
    # populated for the 5 risk students.
    assert (output["Academic Watch"].fillna("") == "").all()
    assert (output["Attendance Watch"] == "Yes").sum() == 5
