from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.assessment import (
    canonicalise_assessment_frame,
    compute_assessment_metrics,
    detect_workbook_assessments,
    latest_assessment_by_student,
    match_assessments_to_roster,
)
from core.combined_risk import attach_combined_risk
from core.confirmed_actions import execute_academic_watch_action
from core.data_sources import DataSourceRegistry
from core.excel_loader import LoadedWorkbook
from core.field_policy import is_protected
from core.query_engine import run_query
from core.schema import canonical_for
from core.workbook_capabilities import detect_capabilities, missing_field_messages
from nlp.planner_router import plan_user_request


def _roster() -> pd.DataFrame:
    return pd.DataFrame({
        "Student ID": ["S001", "S002", "S003", "S004"],
        "Student Name": ["Alice", "Bob", "Cara", "Dan"],
        "Teacher": ["Smith", "Smith", "Lee", "Lee"],
        "Department": ["Math", "Math", "English", "English"],
        "GPA": [3.6, 1.8, 2.4, 1.7],
        "Academic Standing": ["Good Standing", "Probation", "Good Standing", "Warning"],
        "Attendance Rate": [97, 88, 91, 76],
    })


def _inline_assessment() -> pd.DataFrame:
    df = _roster()
    df["SAT Math"] = [610, 470, 540, 430]
    df["SAT EBRW"] = [620, 490, 530, 450]
    df["SAT Total"] = [1230, 960, 1070, 880]
    df["Benchmark Status"] = ["Met", "Below Benchmark", "Met", "Did Not Meet"]
    df["Math Benchmark Met"] = [True, False, True, False]
    df["Reading Benchmark Met"] = [True, False, True, False]
    return df


def _assessment_sheet() -> pd.DataFrame:
    return pd.DataFrame({
        "Student ID": ["S001", "S002", "S002", "S003", "S004", "S999"],
        "Test Type": ["SAT", "PSAT", "SAT", "PSAT", "SAT", "SAT"],
        "Test Date": pd.to_datetime([
            "2026-04-01", "2026-03-01", "2026-05-01",
            "2026-02-15", "2026-05-03", "2026-05-03",
        ]),
        "Math Score": [610, 430, 470, 510, 430, 500],
        "Reading/Writing Score": [620, 440, 490, 520, 450, 500],
        "Total Score": [1230, 870, 960, 1030, 880, 1000],
        "Benchmark Status": ["Met", "Below Benchmark", "Below Benchmark", "Met", "Did Not Meet", "Met"],
        "Math Benchmark Met": [True, False, False, True, False, True],
        "Reading Benchmark Met": [True, False, False, True, False, True],
    })


def _loaded(sheets: dict[str, pd.DataFrame]) -> LoadedWorkbook:
    return LoadedWorkbook(file_name="assessment.xlsx", workbook=None, sheets=sheets, warnings=[])


def _route(message: str, df: pd.DataFrame) -> dict:
    return plan_user_request(
        user_message=message,
        sheets={"Students": df},
        sheet_columns={"Students": list(df.columns)},
        selected_sheet="Students",
        conversation_state=None,
        settings={"llm_enabled": False},
    )


def test_assessment_canonical_mappings():
    assert canonical_for("SAT Math") == "sat_math"
    assert canonical_for("PSAT Reading/Writing") == "psat_reading_writing"
    assert canonical_for("SAT EBRW") == "sat_ebrw"
    assert canonical_for("Math Benchmark Met") == "math_benchmark_met"


def test_inline_sat_psat_columns_detected():
    detection = detect_workbook_assessments({"Students": _inline_assessment()}, "Students")
    assert detection.mode == "inline"
    assert "SAT Math" in detection.inline_columns
    assert "Benchmark Status" in detection.inline_columns


def test_sibling_assessment_sheet_detected_and_matched_by_student_id():
    sheets = {"Students": _roster(), "Assessments": _assessment_sheet()}
    detection = detect_workbook_assessments(sheets, "Students")
    assert detection.mode == "sheet"
    canonical, warnings = canonicalise_assessment_frame(sheets["Assessments"])
    assert not warnings
    matched, unmatched, ids = match_assessments_to_roster(sheets["Students"], canonical)
    assert matched == 5
    assert unmatched == 1
    assert ids == ["S999"]


def test_missing_student_id_prevents_assessment_matching_with_warning():
    canonical, warnings = canonicalise_assessment_frame(pd.DataFrame({"SAT Math": [500]}))
    assert canonical.empty
    assert any("Student ID missing" in warning for warning in warnings)


def test_latest_assessment_by_test_date_selected():
    canonical, _ = canonicalise_assessment_frame(_assessment_sheet())
    latest = latest_assessment_by_student(canonical).set_index("Student ID")
    assert latest.loc["S002", "Test Date"] == pd.Timestamp("2026-05-01")
    assert latest.loc["S002", "Math Score"] == 470


def test_benchmark_fields_map_to_assessment_risk():
    canonical, _ = canonicalise_assessment_frame(_assessment_sheet())
    metrics = compute_assessment_metrics(canonical).set_index("Student ID")
    assert bool(metrics.loc["S002", "Assessment Risk"])
    assert "Benchmark Status" in metrics.loc["S002", "Assessment Reason"]
    assert bool(metrics.loc["S004", "Assessment Risk"])


def test_assessment_score_fields_are_protected():
    for column in ("SAT Math", "SAT EBRW", "SAT Total", "PSAT Math",
                   "PSAT Reading/Writing", "PSAT Total", "Benchmark Status",
                   "Assessment Risk", "College Readiness"):
        assert is_protected(column), column


def test_assistant_refuses_score_edits_as_protected():
    routing = plan_user_request(
        user_message="Change SAT score to 1200",
        sheets={"Students": _inline_assessment()},
        sheet_columns={"Students": list(_inline_assessment().columns)},
        selected_sheet="Students",
        conversation_state=None,
        settings={"llm_enabled": False},
    )
    assert routing["intent"] == "field_update"
    assert routing["validation"]["status"] == "failed"
    assert "protected field" in routing["confirmation_reason"]


def test_below_sat_math_benchmark_routes_when_benchmark_field_exists():
    routing = _route("Show students below SAT math benchmark", _inline_assessment())
    plan = routing["plan"]
    assert any(f["column"] == "Math Benchmark Met" and f["value"] is False for f in plan["filters"])
    assert run_query(plan, {"Students": _inline_assessment()}).row_count == 2


def test_below_benchmark_routes_when_benchmark_status_exists():
    routing = _route("Show students below benchmark", _inline_assessment())
    plan = routing["plan"]
    assert any(f["column"] in {"Assessment Risk", "Benchmark Status"} for f in plan["filters"])


def test_benchmark_unavailable_asks_for_threshold_or_configuration():
    df = _roster()
    df["SAT Math"] = [610, 470, 540, 430]
    routing = _route("Show students below SAT math benchmark", df)
    assert routing["intent"] == "clarify"
    assert "SAT Math below 500" in routing["confirmation_reason"]


def test_sat_math_numeric_threshold_and_average_by_teacher_work():
    df = _inline_assessment()
    threshold = _route("Show students with SAT Math below 500", df)["plan"]
    assert run_query(threshold, {"Students": df}).row_count == 2
    avg = _route("Show average SAT Math by teacher", df)["plan"]
    assert avg["operation"] == "groupby_average"
    assert avg["value_column"] == "SAT Math"
    assert avg["group_by"] == "Teacher"


def test_combined_queries_with_gpa_and_attendance_work():
    df = _inline_assessment()
    gpa = _route("Show students with GPA below 2.0 and below benchmark", df)["plan"]
    assert run_query(gpa, {"Students": df}).row_count == 2
    attendance = _route("Show students with poor attendance and below benchmark", df)["plan"]
    assert run_query(attendance, {"Students": df}).row_count == 2


def test_data_source_enrichment_merges_sibling_assessment_sheet():
    registry = DataSourceRegistry()
    registry.set_roster(_loaded({"Students": _roster(), "Assessments": _assessment_sheet()}))
    enriched = registry.enriched_sheets()["Students"]
    assert "SAT Math" in enriched.columns
    assert "Assessment Risk" in enriched.columns
    assert bool(enriched.loc[enriched["Student ID"] == "S002", "Assessment Risk"].iloc[0])


def test_combined_risk_includes_assessment_risk_and_reason():
    out = attach_combined_risk(_inline_assessment()).set_index("Student ID")
    assert out.loc["S002", "Risk Signals"] >= 2
    assert out.loc["S002", "Risk Level"] == "High Risk"
    assert "Benchmark Status" in out.loc["S002", "Risk Reason"]


def test_assessment_capabilities_and_missing_messages():
    caps = detect_capabilities(list(_inline_assessment().columns))
    by_key = {cap.key: cap for cap in caps}
    assert by_key["assessment_review"].available
    assert by_key["benchmark_risk"].available
    messages = missing_field_messages(["Student ID", "Student Name", "GPA"])
    assert any("Assessment scores not detected" in message for message in messages)


def test_mark_below_benchmark_students_academic_watch_export_audit_and_original(tmp_path):
    df = _inline_assessment()
    snapshot = df.copy()
    result = execute_academic_watch_action(
        filters=[{"column": "Math Benchmark Met", "operator": "equals", "value": False}],
        sheets={"Students": df},
        sheet="Students",
        column_name="Academic Watch",
        output_dir=tmp_path / "outputs",
        audit_path=tmp_path / "audit.jsonl",
        request_summary="mark below benchmark students academic watch and export",
    )
    assert result.success
    pd.testing.assert_frame_equal(df, snapshot)
    assert Path(result.output_file).exists()
    headers = [cell.value for cell in load_workbook(result.output_file)["Students"][1]]
    assert "Academic Watch Reason" in headers
    assert (tmp_path / "audit.jsonl").exists()
