"""PR-1 unit tests for core.session_workbook.

Covers: deterministic naming + conclusion templates, write-after-each-turn,
sensitive-column redaction at write, reset-for-new-source, suppress-last-turn,
cover sheet + session log presence, audit-log entries, and that the original
sheets dict is never mutated.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.session_workbook import (
    SessionWorkbook,
    auto_sheet_name,
    build_conclusion,
)


# ---- fixtures --------------------------------------------------------------


@dataclass
class FakeResult:
    operation: str
    description: str = ""
    value: Any = None
    row_count: int | None = None
    table: list[dict[str, Any]] = field(default_factory=list)
    columns_used: list[str] = field(default_factory=list)


@pytest.fixture()
def roster():
    df = pd.DataFrame({
        "Student ID": ["S1", "S2", "S3", "S4"],
        "Email": ["a@x.edu", "b@x.edu", "c@x.edu", "d@x.edu"],
        "Department": ["Accounting", "Biology", "Accounting", "Biology"],
        "GPA": [2.1, 3.5, 1.9, 3.2],
        "Notes": ["old", "", "", ""],
    })
    return {"Students": df}


@pytest.fixture()
def audit_path(tmp_path):
    return tmp_path / "logs" / "audit_log.jsonl"


@pytest.fixture()
def sw(tmp_path):
    return SessionWorkbook(
        source_file_name="synthetic_students.xlsx",
        schema_hash="abcd1234",
        output_dir=tmp_path / "outputs",
    )


def _record(sw, audit_path, **kwargs):
    """Helper: record_turn with a per-test tmp audit_path."""
    return sw.record_turn(audit_path=audit_path, **kwargs)


ACC_LT_25 = [
    {"column": "Department", "operator": "equals", "value": "Accounting"},
    {"column": "GPA", "operator": "less_than", "value": 2.5},
]


def _filtered_preview_plan(filters=None, sort=None):
    return {
        "operation": "filtered_preview",
        "sheet": "Students",
        "filters": filters or ACC_LT_25,
        "sort": sort or {},
        "group_by": "",
        "limit": 10,
    }


# ---- deterministic naming + conclusion ------------------------------------


def test_auto_sheet_name_filtered_preview_includes_filters():
    name = auto_sheet_name(_filtered_preview_plan())
    assert "Department" in name and "GPA" in name


def test_auto_sheet_name_groupby_uses_group_column():
    plan = {"operation": "groupby_count", "group_by": "Department",
            "filters": [{"column": "GPA", "operator": "less_than", "value": 2.0}]}
    name = auto_sheet_name(plan)
    assert name.startswith("Count by Department")
    assert "GPA" in name  # filter phrase included in parens


def test_auto_sheet_name_falls_back_for_unknown_operation():
    assert auto_sheet_name({"operation": ""}) == "Result"


def test_build_conclusion_uses_description_when_present():
    result = FakeResult(operation="filtered_preview", description="3 students below 2.0 GPA.")
    assert build_conclusion(plan=_filtered_preview_plan(), result=result, row_count=3) \
        == "3 students below 2.0 GPA."


def test_build_conclusion_deterministic_filtered_preview_without_description():
    result = FakeResult(operation="filtered_preview")
    conclusion = build_conclusion(plan=_filtered_preview_plan(), result=result, row_count=2)
    assert conclusion.startswith("2 rows match Department = Accounting")
    assert "GPA < 2.5" in conclusion


def test_build_conclusion_aggregate_includes_value():
    plan = {"operation": "average_column", "value_column": "GPA",
            "filters": [{"column": "Department", "operator": "equals", "value": "Accounting"}]}
    result = FakeResult(operation="average_column", value=2.0)
    conclusion = build_conclusion(plan=plan, result=result, row_count=None)
    assert "Average of GPA" in conclusion and "2.0" in conclusion


# ---- record_turn writes a sheet -------------------------------------------


def test_record_turn_creates_sheet_with_header_rows_and_data(sw, roster, audit_path):
    result = FakeResult(operation="filtered_preview", row_count=2)
    outcome = _record(
        sw, audit_path,
        user_message="show me accounting students below 2.5 gpa",
        plan=_filtered_preview_plan(),
        result=result,
        loaded_sheets=roster,
    )
    assert outcome.action == "created"
    assert Path(sw.path).exists()

    wb = load_workbook(sw.path)
    assert "Cover" in wb.sheetnames
    assert "Session log" in wb.sheetnames
    data_sheets = [n for n in wb.sheetnames if n not in {"Cover", "Session log"}]
    assert len(data_sheets) == 1

    ws = wb[data_sheets[0]]
    assert ws["A1"].value.startswith("Investigation:")
    assert ws["A2"].value == "Question: show me accounting students below 2.5 gpa"
    assert ws["A3"].value.startswith("Filters:")
    assert "Rows: 2" in ws["A4"].value
    assert ws["A5"].value.startswith("Conclusion:")
    # Row 6 is the blank separator; data header row is 7.
    header = [ws.cell(row=7, column=i).value for i in range(1, 5)]
    assert "Student ID" in header  # full rows materialized from the source


def test_record_turn_redacts_sensitive_columns_in_written_data(sw, roster, audit_path):
    result = FakeResult(operation="filtered_preview", row_count=2)
    _record(sw, audit_path, user_message="show accounting", plan=_filtered_preview_plan(),
            result=result, loaded_sheets=roster)
    wb = load_workbook(sw.path)
    data_sheet = next(n for n in wb.sheetnames if n not in {"Cover", "Session log"})
    ws = wb[data_sheet]
    header = [ws.cell(row=7, column=i).value for i in range(1, 10) if ws.cell(row=7, column=i).value]
    # Email and Notes are sensitive-by-default and must not appear in the data.
    assert "Email" not in header
    assert "Notes" not in header
    # The header row also records what was hidden.
    assert "Email" in ws["A4"].value and "Notes" in ws["A4"].value


def test_record_turn_materializes_full_filtered_population_not_preview(sw, roster, audit_path):
    # The preview the user saw might be 1 row, but the sheet should contain the
    # full 2-row filtered set (Accounting AND GPA<2.5 → S1, S3).
    preview_of_one = [{"Student ID": "S1", "Department": "Accounting", "GPA": 2.1}]
    result = FakeResult(operation="filtered_preview", row_count=2, table=preview_of_one)
    _record(sw, audit_path, user_message="show accounting", plan=_filtered_preview_plan(),
            result=result, loaded_sheets=roster)
    wb = load_workbook(sw.path)
    data_sheet = next(n for n in wb.sheetnames if n not in {"Cover", "Session log"})
    ws = wb[data_sheet]
    # 2 data rows means rows 8 and 9 have student IDs, row 10 is empty.
    assert ws.cell(row=8, column=1).value in {"S1", "S3"}
    assert ws.cell(row=9, column=1).value in {"S1", "S3"}
    assert ws.cell(row=10, column=1).value in (None, "")


# ---- write-after-each-turn (atomic rename) --------------------------------


def test_write_after_each_turn(sw, roster, audit_path):
    assert not Path(sw.path).exists() or sw.path.stat().st_size == 0
    result = FakeResult(operation="filtered_preview", row_count=2)

    _record(sw, audit_path, user_message="first", plan=_filtered_preview_plan(),
            result=result, loaded_sheets=roster)
    after_first = sw.path.stat().st_mtime
    assert Path(sw.path).exists()

    _record(sw, audit_path, user_message="second", plan=_filtered_preview_plan(
        filters=[{"column": "Department", "operator": "equals", "value": "Biology"}]),
        result=result, loaded_sheets=roster)
    wb = load_workbook(sw.path)
    data_sheets = [n for n in wb.sheetnames if n not in {"Cover", "Session log"}]
    assert len(data_sheets) == 2
    # Same path is reused for the whole session (write-after-each-turn rewrites the file).
    assert sw.path.stat().st_mtime >= after_first
    # No leftover .tmp file on success.
    assert not sw.path.with_suffix(sw.path.suffix + ".tmp").exists()


# ---- suppress_last_turn ---------------------------------------------------


def test_suppress_last_turn_removes_latest_sheet(sw, roster, audit_path):
    result = FakeResult(operation="filtered_preview", row_count=2)
    _record(sw, audit_path, user_message="first", plan=_filtered_preview_plan(),
            result=result, loaded_sheets=roster)
    _record(sw, audit_path, user_message="second", plan=_filtered_preview_plan(
        filters=[{"column": "Department", "operator": "equals", "value": "Biology"}]),
        result=result, loaded_sheets=roster)
    assert len(sw.sheets) == 2

    assert sw.suppress_last_turn() is True
    assert len(sw.sheets) == 1

    wb = load_workbook(sw.path)
    data_sheets = [n for n in wb.sheetnames if n not in {"Cover", "Session log"}]
    assert len(data_sheets) == 1


def test_suppress_last_turn_when_empty_returns_false(sw):
    assert sw.suppress_last_turn() is False


# ---- reset_for_new_source --------------------------------------------------


def test_reset_for_new_source_opens_fresh_path_and_drops_sheets(sw, roster, audit_path):
    result = FakeResult(operation="filtered_preview", row_count=2)
    _record(sw, audit_path, user_message="x", plan=_filtered_preview_plan(),
            result=result, loaded_sheets=roster)
    original_path = sw.path
    original_session_id = sw.session_id

    sw.reset_for_new_source(source_file_name="other.xlsx", schema_hash="zzzz9999")
    assert sw.source_file_name == "other.xlsx"
    assert sw.schema_hash == "zzzz9999"
    assert sw.session_id != original_session_id
    assert sw.sheets == []
    # New path is resolved lazily on next save.
    assert sw.path != original_path


def test_is_bound_to_matches_filename_and_schema(sw):
    assert sw.is_bound_to(source_file_name="synthetic_students.xlsx", schema_hash="abcd1234") is True
    assert sw.is_bound_to(source_file_name="other.xlsx", schema_hash="abcd1234") is False
    assert sw.is_bound_to(source_file_name="synthetic_students.xlsx", schema_hash="0000") is False


# ---- the source dict is never mutated -------------------------------------


def test_record_turn_does_not_mutate_source_sheets(sw, roster, audit_path):
    before = roster["Students"].copy(deep=True)
    result = FakeResult(operation="filtered_preview", row_count=2)
    _record(sw, audit_path, user_message="x", plan=_filtered_preview_plan(),
            result=result, loaded_sheets=roster)
    pd.testing.assert_frame_equal(roster["Students"], before)


# ---- audit log -------------------------------------------------------------


def test_record_turn_writes_sheet_added_audit_entry(sw, roster, audit_path):
    result = FakeResult(operation="filtered_preview", row_count=2)
    _record(sw, audit_path, user_message="show accounting",
            plan=_filtered_preview_plan(), result=result, loaded_sheets=roster)

    assert audit_path.exists()
    lines = [json.loads(line) for line in audit_path.read_text().splitlines() if line.strip()]
    assert len(lines) == 1
    entry = lines[0]
    assert entry["action_type"] == "sheet_added"
    assert entry["session_id"] == sw.session_id
    assert entry["source_workbook"] == "synthetic_students.xlsx"
    assert entry["source_schema_hash"] == "abcd1234"
    assert entry["operation"] == "filtered_preview"
    assert entry["rows_affected"] == 2
    # Metadata-only — no raw row values, just column names + filter dicts.
    assert "Email" in entry["sensitive_fields_redacted"]
    assert entry["filters_applied"] == ACC_LT_25
    assert entry["confirmation_status"] == "auto_saved"
