"""Headless UI test that actually RUNS edits through the app and checks the
exported workbook copy. Confirms the Phase 3 execute_plan path is wired into the
Run button, that multi-step plans execute, and that the original file on disk is
never overwritten (a new copy is always produced).
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
os.chdir(REPO_ROOT)

import streamlit as st  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from streamlit.testing.v1 import AppTest  # noqa: E402

WORKBOOK = REPO_ROOT / "outputs" / "mock_dean_student_roster_edited_20260519_180232_331696.xlsx"


class FakeUpload:
    def __init__(self, path: Path):
        self._bytes = path.read_bytes()
        self.name = path.name

    def getvalue(self) -> bytes:
        return self._bytes


st.file_uploader = lambda *a, **k: FakeUpload(WORKBOOK)


def seed(at):
    at.session_state["current_user"] = {"username": "tester", "role": "Editor"}
    at.session_state["workbook_upload"] = FakeUpload(WORKBOOK)


def fresh():
    at = AppTest.from_file(str(REPO_ROOT / "app.py"), default_timeout=180)
    seed(at)
    at.run()
    return at


def click(at, label):
    for b in at.button:
        if b.label == label:
            b.click()
            return True
    return False


def send(at, text):
    at.text_input[0].set_value(text)
    for b in at.button:
        if b.label == "Send":
            b.click()
            break
    seed(at)
    at.run()


def check(name, cond, detail=""):
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))
    return cond


def main() -> int:
    ok = True
    original_bytes = WORKBOOK.read_bytes()

    print("1) Plan + RUN a single chart edit through the UI")
    at = fresh()
    send(at, "Create a chart of students by Major")
    ok &= check("planned, executable", at.session_state["current_can_execute"] is True)
    ok &= check("run button present", click(at, "Run Action"), "Run Action")
    seed(at)
    at.run()
    ok &= check("no exception on run", not at.exception, str(at.exception))
    out = at.session_state["latest_output_file"] if "latest_output_file" in at.session_state else None
    ok &= check("exported a new file", bool(out) and Path(out).exists(), str(out))
    if out and Path(out).exists():
        names = load_workbook(out).sheetnames
        ok &= check("export contains a chart sheet", any("Chart" in n for n in names), str(names))

    print("\n2) Inject a MULTI-STEP plan and RUN it")
    at = fresh()
    multi = {
        "plan_type": "multi_step_plan",
        "commands": [
            {"action": "filter_rows", "sheet": "Student Roster",
             "conditions": [{"column": "Standing", "operator": "is_not_missing"}],
             "output_sheet": "All Students"},
            {"action": "count_by_group", "sheet": "All Students", "group_by": "Major",
             "output_sheet": "Major Counts"},
            {"action": "format_report", "sheets": ["Major Counts"]},
        ],
    }
    at.session_state["assistant_mode"] = "edit_workbook"
    at.session_state["current_plan"] = multi
    at.session_state["current_command"] = multi["commands"][0]
    at.session_state["current_request"] = "break down students by major"
    at.session_state["current_confidence"] = 0.9
    at.session_state["current_confidence_level"] = "high"
    at.session_state["current_validation_error"] = None
    at.session_state["current_can_execute"] = True
    seed(at)
    at.run()
    ok &= check("multi-step run button present", click(at, "Run Action"))
    seed(at)
    at.run()
    ok &= check("no exception on multi-step run", not at.exception, str(at.exception))
    out = at.session_state["latest_output_file"] if "latest_output_file" in at.session_state else None
    ok &= check("exported a new file", bool(out) and Path(out).exists(), str(out))
    if out and Path(out).exists():
        names = load_workbook(out).sheetnames
        ok &= check("export has both intermediate sheets",
                    "All Students" in names and "Major Counts" in names, str(names))

    print("\n3) Original workbook on disk is untouched")
    ok &= check("source bytes unchanged", WORKBOOK.read_bytes() == original_bytes)

    print("\n" + ("ALL PASS" if ok else "SOME FAILURES"))
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
