"""Headless test of the Phase 3 plan executor.

Builds an in-memory workbook with proper numeric dtypes (so numeric ops aren't
blocked by the str-typed-loader limitation), then runs a multi-step plan and
each newly-wired single action through execute_plan, asserting output sheets are
created and the source sheet is preserved.
"""

from __future__ import annotations

import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

import pandas as pd  # noqa: E402
from openpyxl import Workbook  # noqa: E402

from core.action_engine import execute_plan  # noqa: E402


def build():
    df = pd.DataFrame({
        "Student ID": [f"S{i:03d}" for i in range(1, 13)],
        "Program": (["Nursing", "Biology", "Nursing", "Biology", "History", "Nursing"] * 2),
        "Balance Due": [0, 250, 500, 0, 125, 800, 0, 300, 50, 0, 990, 0],
        "Standing": (["Good", "Probation", "Good", "Good", "Probation", "Good"] * 2),
    })
    wb = Workbook()
    ws = wb.active
    ws.title = "Enrollment"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    sheets = {"Enrollment": df}
    return wb, sheets


def check(name, cond, detail=""):
    print(f"  [{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))
    return cond


def main() -> int:
    ok = True

    print("1) MULTI-STEP plan: filter -> sum_by_group -> chart -> format")
    wb, sheets = build()
    plan = {
        "plan_type": "multi_step_plan",
        "commands": [
            {"action": "filter_rows", "sheet": "Enrollment",
             "conditions": [{"column": "Balance Due", "operator": "greater_than", "value": 0}],
             "output_sheet": "Students With Balance"},
            {"action": "sum_by_group", "sheet": "Students With Balance", "group_by": "Program",
             "sum_column": "Balance Due", "output_sheet": "Balance by Program"},
            {"action": "create_chart", "sheet": "Balance by Program", "chart_type": "bar",
             "category_column": "Program", "value_column": "Balance Due",
             "title": "Outstanding Balance by Program", "output_sheet": "Balance Chart"},
            {"action": "format_report", "sheets": ["Students With Balance", "Balance by Program"]},
        ],
    }
    res = execute_plan(plan, wb, sheets, "roster.xlsx")
    ok &= check("4 step messages", len(res.step_messages) == 4, str(res.step_messages))
    ok &= check("filtered sheet created", "Students With Balance" in wb.sheetnames)
    ok &= check("grouped sheet created", "Balance by Program" in wb.sheetnames)
    ok &= check("chart sheet created", "Balance Chart" in wb.sheetnames)
    ok &= check("original Enrollment preserved", "Enrollment" in wb.sheetnames and wb["Enrollment"].max_row == 13)
    # 7 students have a balance > 0; grouped totals should be in the dict
    grouped = sheets.get("Balance by Program")
    ok &= check("grouped totals computed", grouped is not None and "Balance Due" in grouped.columns,
                str(None if grouped is None else list(grouped.columns)))

    print("\n2) Single new actions")
    for label, command, expect_sheet in [
        ("sort_rows", {"action": "sort_rows", "sheet": "Enrollment",
                       "sort_by": [{"column": "Balance Due", "direction": "desc"}],
                       "output_sheet": "Sorted"}, "Sorted"),
        ("average_column", {"action": "average_column", "sheet": "Enrollment", "column": "Balance Due",
                            "group_by": "Program", "output_sheet": "Avg by Program"}, "Avg by Program"),
        ("move_rows_to_sheet", {"action": "move_rows_to_sheet", "sheet": "Enrollment",
                                "conditions": [{"column": "Standing", "operator": "equals", "value": "Probation"}],
                                "output_sheet": "Probation"}, "Probation"),
        ("create_summary_sheet", {"action": "create_summary_sheet", "sheet": "Enrollment",
                                  "metrics": [{"kind": "count", "label": "Students"},
                                              {"kind": "sum", "column": "Balance Due", "label": "Total Owed"},
                                              {"kind": "count_missing", "column": "Standing", "label": "Missing Standing"}],
                                  "output_sheet": "Summary"}, "Summary"),
        ("freeze_header", {"action": "freeze_header", "sheet": "Enrollment"}, None),
        ("autofit_columns", {"action": "autofit_columns", "sheet": "Enrollment"}, None),
        ("apply_conditional_formatting", {"action": "apply_conditional_formatting", "sheet": "Enrollment",
                                          "conditions": [{"column": "Balance Due", "operator": "greater_than", "value": 0}],
                                          "format": {"fill_color": "red"}}, None),
    ]:
        wb, sheets = build()
        try:
            res = execute_plan({"plan_type": "single_action", "commands": [command]}, wb, sheets, "roster.xlsx")
            crashed = False
        except Exception as exc:  # noqa: BLE001
            crashed = True
            res = None
            detail = str(exc)
        if crashed:
            ok &= check(label, False, detail)
            continue
        if expect_sheet:
            ok &= check(label, expect_sheet in wb.sheetnames, res.message)
        else:
            ok &= check(label, "Enrollment" in wb.sheetnames, res.message)

    print("\n3) freeze_header sets pane")
    wb, sheets = build()
    execute_plan({"plan_type": "single_action", "commands": [{"action": "freeze_header", "sheet": "Enrollment"}]},
                 wb, sheets, "roster.xlsx")
    ok &= check("freeze pane is A2", wb["Enrollment"].freeze_panes == "A2", str(wb["Enrollment"].freeze_panes))

    print("\n" + ("ALL PASS" if ok else "SOME FAILURES"))
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
