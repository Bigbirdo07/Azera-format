"""E2E smoke test for PSAT/SAT assessment risk workflow.

Runs without cloud services and without sending workbook rows to any model.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from core.confirmed_actions import execute_academic_watch_action
from core.excel_loader import load_excel_workbook
from core.query_engine import run_query
from core.workbook_capabilities import detect_capabilities
from nlp.planner_router import plan_user_request


FIXTURE = ROOT / "tests" / "fixtures" / "academic_assessment_roster.xlsx"


class FakeUpload:
    def __init__(self, path: Path):
        self.name = path.name
        self._payload = path.read_bytes()

    def getvalue(self) -> bytes:
        return self._payload


def _route(message: str, sheets: dict, columns_by_sheet: dict) -> dict:
    return plan_user_request(
        user_message=message,
        sheets=sheets,
        sheet_columns=columns_by_sheet,
        selected_sheet="Students",
        conversation_state=None,
        settings={"llm_enabled": False},
    )


def main() -> int:
    loaded = load_excel_workbook(FakeUpload(FIXTURE))
    sheets = loaded.sheets
    original = sheets["Students"].copy()
    columns_by_sheet = {"Students": list(original.columns)}
    chat_history: list[dict[str, str]] = []

    caps = detect_capabilities(list(original.columns))
    assert any(c.key == "assessment_review" and c.available for c in caps)

    first = _route("Show students below SAT math benchmark", sheets, columns_by_sheet)
    chat_history.append({"role": "user", "content": "Show students below SAT math benchmark"})
    result = run_query(first["plan"], sheets)
    assert result.row_count == 3

    second = _route("Which teachers have the most students below benchmark?", sheets, columns_by_sheet)
    chat_history.append({"role": "user", "content": "Which teachers have the most students below benchmark?"})
    teacher_result = run_query(second["plan"], sheets)
    assert teacher_result.table

    third = _route("Now only students with GPA below 2.0", sheets, columns_by_sheet)
    chat_history.append({"role": "user", "content": "Now only students with GPA below 2.0"})
    gpa_result = run_query(third["plan"], sheets)
    assert gpa_result.row_count >= 1

    output_dir = ROOT / "outputs"
    audit_path = ROOT / "logs" / "assessment_e2e_audit.jsonl"
    if audit_path.exists():
        audit_path.unlink()
    action = execute_academic_watch_action(
        filters=[
            {"column": "Math Benchmark Met", "operator": "equals", "value": False},
            {"column": "GPA", "operator": "less_than", "value": 2.0},
        ],
        sheets=sheets,
        sheet="Students",
        column_name="Academic Watch",
        output_dir=output_dir,
        audit_path=audit_path,
        request_summary="Mark below benchmark low GPA students Academic Watch and export",
    )
    chat_history.append({"role": "assistant", "content": action.message})

    assert action.success
    assert action.output_file and Path(action.output_file).exists()
    pd.testing.assert_frame_equal(sheets["Students"], original)

    wb = load_workbook(action.output_file)
    ws = wb["Students"]
    headers = [cell.value for cell in ws[1]]
    assert "Academic Watch" in headers
    assert "Academic Watch Reason" in headers
    assert "Date Flagged" in headers

    assert audit_path.exists()
    audit = json.loads(audit_path.read_text(encoding="utf-8").splitlines()[-1])
    assert audit["action_type"] == "academic_watch"
    assert audit["rows_affected"] >= 1
    assert len(chat_history) >= 4

    print("ASSESSMENT E2E: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
