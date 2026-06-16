from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class FormulaResult:
    message: str
    preview: pd.DataFrame | None = None
    result_sheet: str | None = None


def create_formula(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> FormulaResult:
    formula_type = command["formula_type"].upper()

    if command.get("new_column"):
        return add_formula_column(command, workbook)

    return add_summary_formula_sheet(command, workbook)


def add_formula_column(command: dict[str, Any], workbook: Workbook) -> FormulaResult:
    sheet_name = command["sheet"]
    worksheet = workbook[sheet_name]
    formula_type = command["formula_type"].upper()
    new_column = command["new_column"]
    new_column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=new_column_index, value=new_column)

    for row_index in range(2, worksheet.max_row + 1):
        formula = _row_formula(command, workbook, worksheet, row_index, formula_type)
        worksheet.cell(row=row_index, column=new_column_index, value=formula)

    _format_header_cell(worksheet.cell(row=1, column=new_column_index))
    _autosize_column(worksheet, new_column_index)

    preview = pd.DataFrame(
        [
            {
                "sheet": sheet_name,
                "new_column": new_column,
                "formula_type": formula_type,
                "first_formula": worksheet.cell(row=2, column=new_column_index).value if worksheet.max_row >= 2 else "",
            }
        ]
    )
    return FormulaResult(
        message=f"Added {formula_type} formulas to new column {new_column} in {sheet_name}.",
        preview=preview,
    )


def add_summary_formula_sheet(command: dict[str, Any], workbook: Workbook) -> FormulaResult:
    sheet_name = command["sheet"]
    source = workbook[sheet_name]
    formula_type = command["formula_type"].upper()
    summary_name = _unique_sheet_name(workbook, command.get("summary_sheet", f"{formula_type}_Summary"))
    summary = workbook.create_sheet(summary_name)
    _write_summary_header(summary)

    formulas = _summary_formulas(command, workbook, source, formula_type)
    for row_index, item in enumerate(formulas, start=2):
        summary.cell(row=row_index, column=1, value=item["label"])
        summary.cell(row=row_index, column=2, value=item["formula"])

    _autosize_column(summary, 1)
    _autosize_column(summary, 2)

    preview = pd.DataFrame(formulas)
    return FormulaResult(
        message=f"Added {formula_type} summary formulas to {summary_name}.",
        preview=preview,
        result_sheet=summary_name,
    )


def _row_formula(
    command: dict[str, Any],
    workbook: Workbook,
    worksheet: Worksheet,
    row_index: int,
    formula_type: str,
) -> str:
    if formula_type == "IF":
        logic = command["logic"]
        condition_ref = _cell_ref(worksheet, logic["condition_column"], row_index)
        condition = _condition_expression(condition_ref, logic["operator"], logic.get("value"))
        return f'=IF({condition},{_formula_value(logic.get("true_value", "Yes"))},{_formula_value(logic.get("false_value", "No"))})'

    if formula_type == "IFS":
        parts = []
        for rule in command["logic"]["rules"]:
            condition_ref = _cell_ref(worksheet, rule["condition_column"], row_index)
            parts.append(_condition_expression(condition_ref, rule["operator"], rule.get("value")))
            parts.append(_formula_value(rule.get("result", "")))
        if command["logic"].get("default") is not None:
            parts.extend(["TRUE", _formula_value(command["logic"]["default"])])
        return f'=IFS({",".join(parts)})'

    if formula_type == "XLOOKUP":
        lookup_sheet = workbook[command["lookup"]["lookup_sheet"]]
        lookup_value = _cell_ref(worksheet, command["lookup"]["lookup_value_column"], row_index)
        lookup_range = _column_range(lookup_sheet, command["lookup"]["lookup_key_column"])
        return_range = _column_range(lookup_sheet, command["lookup"]["return_column"])
        return f'=XLOOKUP({lookup_value},{lookup_range},{return_range},"")'

    if formula_type == "VLOOKUP":
        lookup_sheet = workbook[command["lookup"]["lookup_sheet"]]
        lookup_value = _cell_ref(worksheet, command["lookup"]["lookup_value_column"], row_index)
        table_range = _table_range(lookup_sheet)
        return_index = _column_index(lookup_sheet, command["lookup"]["return_column"])
        return f'=VLOOKUP({lookup_value},{table_range},{return_index},FALSE)'

    if formula_type == "CONCAT":
        refs = [_cell_ref(worksheet, column, row_index) for column in command["columns"]]
        return f'=CONCAT({",".join(refs)})'

    if formula_type == "TEXT":
        value_ref = _cell_ref(worksheet, command["column"], row_index)
        return f'=TEXT({value_ref},{_formula_value(command.get("format_text", "0"))})'

    if formula_type == "TODAY":
        return "=TODAY()"

    raise ValueError(f"{formula_type} formulas are not supported as row formulas.")


def _summary_formulas(
    command: dict[str, Any],
    workbook: Workbook,
    source: Worksheet,
    formula_type: str,
) -> list[dict[str, str]]:
    if formula_type in {"SUM", "AVERAGE", "COUNT", "COUNTA"}:
        column = command["column"]
        return [{"label": f"{formula_type} of {column}", "formula": f"={formula_type}({_column_range(source, column)})"}]

    if formula_type == "COUNTIF":
        criteria = command["criteria"][0]
        return [
            {
                "label": f"COUNTIF {criteria['column']}",
                "formula": f'=COUNTIF({_column_range(source, criteria["column"])},{_formula_value(_criteria_value(criteria))})',
            }
        ]

    if formula_type == "COUNTIFS":
        formula_parts = []
        for criteria in command["criteria"]:
            formula_parts.extend([_column_range(source, criteria["column"]), _formula_value(_criteria_value(criteria))])
        return [{"label": "COUNTIFS summary", "formula": f'=COUNTIFS({",".join(formula_parts)})'}]

    if formula_type == "SUMIF":
        criteria = command["criteria"][0]
        return [
            {
                "label": f"SUMIF {command['sum_column']}",
                "formula": f'=SUMIF({_column_range(source, criteria["column"])},{_formula_value(_criteria_value(criteria))},{_column_range(source, command["sum_column"])})',
            }
        ]

    if formula_type == "SUMIFS":
        formula_parts = [_column_range(source, command["sum_column"])]
        for criteria in command["criteria"]:
            formula_parts.extend([_column_range(source, criteria["column"]), _formula_value(_criteria_value(criteria))])
        return [{"label": f"SUMIFS {command['sum_column']}", "formula": f'=SUMIFS({",".join(formula_parts)})'}]

    if formula_type == "UNIQUE":
        return [{"label": f"UNIQUE {command['column']}", "formula": f"=UNIQUE({_column_range(source, command['column'])})"}]

    if formula_type == "SORT":
        return [{"label": f"SORT {command['column']}", "formula": f"=SORT({_column_range(source, command['column'])})"}]

    if formula_type == "FILTER":
        criteria = command["criteria"][0]
        condition = _condition_expression(_column_range(source, criteria["column"]), criteria["operator"], criteria.get("value"))
        return [{"label": "FILTER summary", "formula": f"=FILTER({_table_range(source)},{condition})"}]

    if formula_type == "TODAY":
        return [{"label": "Today", "formula": "=TODAY()"}]

    raise ValueError(f"{formula_type} formulas are not supported as summary formulas.")


def _condition_expression(reference: str, operator: str, value: Any) -> str:
    operator_map = {
        "equals": "=",
        "not_equals": "<>",
        "greater_than": ">",
        "greater_or_equal": ">=",
        "less_than": "<",
        "less_or_equal": "<=",
    }
    if operator == "is_missing":
        return f'{reference}=""'
    if operator == "is_not_missing":
        return f'{reference}<>""'
    if operator == "contains":
        return f'ISNUMBER(SEARCH({_formula_value(value)},{reference}))'
    if operator == "not_contains":
        return f'NOT(ISNUMBER(SEARCH({_formula_value(value)},{reference})))'
    if operator not in operator_map:
        raise ValueError(f"Unsupported formula operator: {operator}")
    return f"{reference}{operator_map[operator]}{_formula_value(value)}"


def _criteria_value(criteria: dict[str, Any]) -> Any:
    operator = criteria["operator"]
    value = criteria.get("value")
    prefix_map = {
        "equals": "",
        "not_equals": "<>",
        "greater_than": ">",
        "greater_or_equal": ">=",
        "less_than": "<",
        "less_or_equal": "<=",
    }
    if operator == "is_missing":
        return ""
    if operator == "is_not_missing":
        return "<>"
    if operator == "contains":
        return f"*{value}*"
    if operator == "not_contains":
        return f"<>*{value}*"
    return f"{prefix_map[operator]}{value}"


def _formula_value(value: Any) -> str:
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    return f'"{str(value).replace(chr(34), chr(34) + chr(34))}"'


def _cell_ref(worksheet: Worksheet, column_name: str, row_index: int) -> str:
    return f"{get_column_letter(_column_index(worksheet, column_name))}{row_index}"


def _column_range(worksheet: Worksheet, column_name: str) -> str:
    column_letter = get_column_letter(_column_index(worksheet, column_name))
    return f"'{worksheet.title}'!${column_letter}$2:${column_letter}${worksheet.max_row}"


def _table_range(worksheet: Worksheet) -> str:
    last_column = get_column_letter(worksheet.max_column)
    return f"'{worksheet.title}'!$A$2:${last_column}${worksheet.max_row}"


def _column_index(worksheet: Worksheet, column_name: str) -> int:
    for cell in worksheet[1]:
        if str(cell.value) == str(column_name):
            return cell.column
    raise ValueError(f"Column does not exist: {column_name}")


def _write_summary_header(worksheet: Worksheet) -> None:
    worksheet.append(["Metric", "Formula"])
    for cell in worksheet[1]:
        _format_header_cell(cell)


def _format_header_cell(cell) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    cell.font = Font(color="FFFFFF", bold=True)


def _autosize_column(worksheet: Worksheet, column_index: int) -> None:
    column_letter = get_column_letter(column_index)
    max_length = max(
        len(str(worksheet.cell(row=row, column=column_index).value or ""))
        for row in range(1, worksheet.max_row + 1)
    )
    worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)


def _unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    safe_base_name = base_name[:31]
    if safe_base_name not in workbook.sheetnames:
        return safe_base_name

    counter = 1
    while True:
        suffix = f"_{counter}"
        candidate = f"{safe_base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1
