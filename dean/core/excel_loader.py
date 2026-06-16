from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.workbook_diagnostics import diagnose_workbook


@dataclass(frozen=True)
class LoadedWorkbook:
    file_name: str
    workbook: Workbook
    sheets: dict[str, pd.DataFrame]
    warnings: list[str] = field(default_factory=list)


def load_excel_workbook(uploaded_file) -> LoadedWorkbook:
    """Load an uploaded .xlsx file with openpyxl and pandas.

    The original workbook is never modified; only the in-memory working copy is
    cleaned (trimmed/deduped headers, dropped blank rows/placeholder columns).
    Messy-input observations are collected as user-facing warnings.
    """
    file_name = Path(uploaded_file.name).name
    file_bytes = uploaded_file.getvalue()

    workbook = load_workbook(BytesIO(file_bytes))
    diagnostics = diagnose_workbook(workbook)
    sheets: dict[str, pd.DataFrame] = {}
    warnings: list[str] = []
    for sheet in diagnostics.sheets:
        if sheet.is_hidden:
            continue
        worksheet = workbook[sheet.name]
        frame, sheet_warnings = _worksheet_to_dataframe(
            worksheet,
            header_row=sheet.likely_header_row or 1,
        )
        warnings.extend(f"{sheet.name}: {w}" for w in sheet_warnings)
        if frame.empty or not list(frame.columns):
            warnings.append(f"Sheet '{sheet.name}' is empty and was ignored.")
            continue
        sheets[sheet.name] = frame

    # Analysis-level warnings (e.g. numeric values stored as text).
    try:
        from core.schema import schema_warnings

        warnings.extend(w for w in schema_warnings(sheets) if "is empty" not in w)
    except Exception:
        pass

    return LoadedWorkbook(
        file_name=file_name,
        workbook=workbook,
        sheets=sheets,
        warnings=warnings,
    )


def _worksheet_to_dataframe(
    worksheet: Worksheet,
    *,
    header_row: int,
) -> tuple[pd.DataFrame, list[str]]:
    warnings: list[str] = []
    max_row, max_column = _used_bounds(worksheet)
    if max_row == 0 or max_column == 0:
        return pd.DataFrame(), warnings

    header_row = min(max(header_row, 1), max_row)
    raw_headers = [worksheet.cell(row=header_row, column=column).value for column in range(1, max_column + 1)]
    if any(isinstance(v, str) and v != v.strip() for v in raw_headers):
        warnings.append("Some column names had leading/trailing spaces and were trimmed.")
    headers = [_header_name(value, column) for column, value in enumerate(raw_headers, start=1)]
    deduped = _deduplicate_headers(headers)
    if deduped != headers:
        warnings.append("The workbook contains duplicate column names. I renamed them internally.")
    headers = deduped

    rows: list[list[Any]] = []
    blank_rows = 0
    for row_index in range(header_row + 1, max_row + 1):
        row_values = [
            _excel_safe_value(worksheet.cell(row=row_index, column=column).value)
            for column in range(1, max_column + 1)
        ]
        if all(value in (None, "") for value in row_values):
            blank_rows += 1
            continue
        rows.append(row_values)
    if blank_rows:
        warnings.append(f"{blank_rows} completely blank row(s) were dropped from the working copy.")

    frame = pd.DataFrame(rows, columns=headers)
    frame, dropped_columns = _drop_placeholder_blank_columns(frame)
    if dropped_columns:
        warnings.append(f"{dropped_columns} unnamed blank column(s) were ignored.")
    return frame, warnings


def _drop_placeholder_blank_columns(frame: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Drop only auto-named ('Column N') columns that are entirely blank.
    Named-but-empty columns (e.g. Notes) are preserved."""
    to_drop = []
    for column in frame.columns:
        if str(column).startswith("Column ") and _is_blank_column(frame[column]):
            to_drop.append(column)
    if to_drop:
        return frame.drop(columns=to_drop), len(to_drop)
    return frame, 0


def _is_blank_column(series: pd.Series) -> bool:
    normalized = series.astype(str).str.strip()
    return bool((series.isna() | (normalized == "")).all())


def _used_bounds(worksheet: Worksheet) -> tuple[int, int]:
    max_row = 0
    max_column = 0
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_column = max(max_column, cell.column)
    return max_row, max_column


def _header_name(value: Any, column_number: int) -> str:
    if value in (None, ""):
        return f"Column {column_number}"
    return str(value).strip() or f"Column {column_number}"


def _deduplicate_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    deduplicated = []
    for header in headers:
        count = seen.get(header, 0)
        if count == 0:
            deduplicated.append(header)
        else:
            deduplicated.append(f"{header}_{count + 1}")
        seen[header] = count + 1
    return deduplicated


def _excel_safe_value(value: Any) -> Any:
    if hasattr(value, "item"):
        return value.item()
    return value
