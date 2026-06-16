from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ReportResult:
    message: str
    preview: pd.DataFrame | None = None
    result_sheet: str | None = None


def create_report(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ReportResult:
    source_sheet = command["sheet"]
    dataframe = sheets[source_sheet]
    filtered = dataframe[_condition_mask(dataframe, command.get("conditions", []))].copy()

    output_sheet = _unique_sheet_name(
        workbook,
        command.get("output_sheet") or command.get("report_type", "Report"),
    )
    worksheet = workbook.create_sheet(output_sheet)

    title = command.get("title") or output_sheet
    _write_title_block(worksheet, title)
    next_row = _write_summary(worksheet, dataframe, filtered, command, start_row=4)
    table_start = next_row + 2
    _write_table(worksheet, filtered, table_start)

    if command.get("include_chart"):
        _write_optional_chart(worksheet, filtered, command, table_start)

    worksheet.freeze_panes = f"A{table_start + 1}"
    _autosize_columns(worksheet)

    return ReportResult(
        message=f"Created report sheet {output_sheet}.",
        preview=filtered.head(100),
        result_sheet=output_sheet,
    )


def _write_title_block(worksheet: Worksheet, title: str) -> None:
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=16, color="1F4E78")
    worksheet["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    worksheet["A2"].font = Font(italic=True, color="666666")


def _write_summary(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    filtered: pd.DataFrame,
    command: dict[str, Any],
    start_row: int,
) -> int:
    worksheet.cell(row=start_row, column=1, value="Summary Metrics")
    worksheet.cell(row=start_row, column=1).font = Font(bold=True)

    metrics = [
        ("Total source rows", len(dataframe.index)),
        ("Rows in report", len(filtered.index)),
    ]

    for condition in command.get("conditions", []):
        metrics.append((f"Condition: {condition['column']}", _condition_label(condition)))

    balance_column = _find_column_like(filtered, ["balance", "amount due"])
    if balance_column and pd.api.types.is_numeric_dtype(filtered[balance_column]):
        metrics.append((f"Total {balance_column}", float(filtered[balance_column].sum())))

    for index, (label, value) in enumerate(metrics, start=start_row + 1):
        worksheet.cell(row=index, column=1, value=label)
        worksheet.cell(row=index, column=2, value=value)

    return start_row + len(metrics)


def _write_table(worksheet: Worksheet, dataframe: pd.DataFrame, start_row: int) -> None:
    if dataframe.empty:
        worksheet.cell(row=start_row, column=1, value="No matching rows")
        return

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col_index, column in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=start_row, column=col_index, value=str(column))
        cell.fill = header_fill
        cell.font = header_font

    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=start_row + 1):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=None if pd.isna(value) else value)


def _write_optional_chart(
    worksheet: Worksheet,
    filtered: pd.DataFrame,
    command: dict[str, Any],
    table_start: int,
) -> None:
    group_by = command.get("chart_group_by")
    if not group_by or group_by not in filtered.columns or filtered.empty:
        return

    summary = filtered.groupby(group_by, dropna=False).size().reset_index(name="Count")
    chart_start = table_start
    chart_col = max(len(filtered.columns) + 3, 5)

    worksheet.cell(row=chart_start, column=chart_col, value=group_by)
    worksheet.cell(row=chart_start, column=chart_col + 1, value="Count")
    for row_offset, row in enumerate(summary.itertuples(index=False, name=None), start=1):
        worksheet.cell(row=chart_start + row_offset, column=chart_col, value=row[0])
        worksheet.cell(row=chart_start + row_offset, column=chart_col + 1, value=row[1])

    chart = BarChart()
    chart.title = command.get("chart_title") or f"Rows by {group_by}"
    chart.type = "col"
    data = Reference(worksheet, min_col=chart_col + 1, min_row=chart_start, max_row=chart_start + len(summary.index))
    cats = Reference(worksheet, min_col=chart_col, min_row=chart_start + 1, max_row=chart_start + len(summary.index))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    worksheet.add_chart(chart, f"{get_column_letter(chart_col)}{chart_start + len(summary.index) + 3}")


def _condition_mask(dataframe: pd.DataFrame, conditions: list[dict[str, Any]]) -> pd.Series:
    mask = pd.Series(True, index=dataframe.index)
    for condition in conditions:
        column = condition["column"]
        operator = condition["operator"]
        value = condition.get("value")
        series = dataframe[column]

        if operator == "equals":
            condition_mask = series.astype(str).str.casefold() == str(value).casefold()
        elif operator == "not_equals":
            condition_mask = series.astype(str).str.casefold() != str(value).casefold()
        elif operator == "greater_than":
            condition_mask = pd.to_numeric(series, errors="coerce") > float(value)
        elif operator == "greater_or_equal":
            condition_mask = pd.to_numeric(series, errors="coerce") >= float(value)
        elif operator == "less_than":
            condition_mask = pd.to_numeric(series, errors="coerce") < float(value)
        elif operator == "less_or_equal":
            condition_mask = pd.to_numeric(series, errors="coerce") <= float(value)
        elif operator == "contains":
            condition_mask = series.astype(str).str.contains(str(value), case=False, na=False)
        elif operator == "contains_any":
            condition_mask = series.astype(str).apply(
                lambda item: any(str(term).casefold() in item.casefold() for term in value)
            )
        elif operator == "not_contains":
            condition_mask = ~series.astype(str).str.contains(str(value), case=False, na=False)
        elif operator == "is_missing":
            condition_mask = series.isna() | (series.astype(str).str.strip() == "")
        elif operator == "is_not_missing":
            condition_mask = series.notna() & (series.astype(str).str.strip() != "")
        else:
            raise ValueError(f"Unsupported operator: {operator}")

        mask &= condition_mask.fillna(False)
    return mask


def _condition_label(condition: dict[str, Any]) -> str:
    operator = condition["operator"].replace("_", " ")
    if "value" in condition:
        return f"{operator} {condition['value']}"
    return operator


def _find_column_like(dataframe: pd.DataFrame, terms: list[str]) -> str | None:
    for column in dataframe.columns:
        normalized = str(column).lower()
        if any(term in normalized for term in terms):
            return column
    return None


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    safe_base_name = base_name[:31]
    if safe_base_name not in workbook.sheetnames:
        return safe_base_name

    counter = 1
    while True:
        suffix = f"_{counter}"
        candidate = f"{safe_base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1
