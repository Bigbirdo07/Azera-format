from __future__ import annotations

from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ChartResult:
    message: str
    preview: pd.DataFrame | None = None
    result_sheet: str | None = None


def create_chart(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ChartResult:
    sheet_name = command["sheet"]
    dataframe = sheets[sheet_name]
    chart_type = command["chart_type"]
    group_by = command["group_by"]
    metric = command["metric"]
    output_sheet = _unique_sheet_name(
        workbook,
        command.get("output_sheet") or f"{group_by} Chart",
    )

    summary = _build_summary_table(dataframe, command)
    worksheet = workbook.create_sheet(output_sheet)
    _write_summary_table(worksheet, summary)
    chart = _build_chart(chart_type, worksheet, len(summary.index), _chart_title(command))
    worksheet.add_chart(chart, "D2")
    _autosize_columns(worksheet)

    return ChartResult(
        message=f"Created {chart_type.replace('_', ' ')} chart on {output_sheet}.",
        preview=summary,
        result_sheet=output_sheet,
    )


def _build_summary_table(dataframe: pd.DataFrame, command: dict[str, Any]) -> pd.DataFrame:
    group_by = command["group_by"]
    metric = command["metric"]

    if metric == "count_rows":
        summary = dataframe.groupby(group_by, dropna=False).size().reset_index(name="Count")
    elif metric == "sum":
        value_column = command["value_column"]
        summary = (
            dataframe.groupby(group_by, dropna=False)[value_column]
            .sum()
            .reset_index(name=f"{value_column} Sum")
        )
    elif metric == "count_missing":
        value_column = command["value_column"]
        summary = (
            dataframe.assign(_missing_value=dataframe[value_column].isna() | (dataframe[value_column].astype(str).str.strip() == ""))
            .groupby(group_by, dropna=False)["_missing_value"]
            .sum()
            .reset_index(name=f"Missing {value_column}")
        )
    else:
        raise ValueError(f"Unsupported chart metric: {metric}")

    return summary.sort_values(summary.columns[0]).reset_index(drop=True)


def _write_summary_table(worksheet: Worksheet, summary: pd.DataFrame) -> None:
    worksheet.append([str(column) for column in summary.columns])
    for row in summary.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def _build_chart(chart_type: str, worksheet: Worksheet, row_count: int, title: str):
    data = Reference(worksheet, min_col=2, min_row=1, max_row=row_count + 1)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=row_count + 1)

    if chart_type == "pie":
        chart = PieChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
    elif chart_type == "line":
        chart = LineChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.y_axis.title = worksheet.cell(row=1, column=2).value
        chart.x_axis.title = worksheet.cell(row=1, column=1).value
    else:
        chart = BarChart()
        chart.type = "bar" if chart_type in {"bar", "stacked_bar"} else "col"
        if chart_type == "stacked_bar":
            chart.grouping = "stacked"
            chart.overlap = 100
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.y_axis.title = worksheet.cell(row=1, column=2).value
        chart.x_axis.title = worksheet.cell(row=1, column=1).value

    chart.title = title
    chart.height = 8
    chart.width = 14
    return chart


def _chart_title(command: dict[str, Any]) -> str:
    if command.get("title"):
        return command["title"]

    group_by = command["group_by"]
    metric = command["metric"]
    if metric == "sum":
        return f"{command['value_column']} by {group_by}"
    if metric == "count_missing":
        return f"Missing {command['value_column']} by {group_by}"
    return f"Rows by {group_by}"


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    safe_base_name = base_name[:31]
    if safe_base_name not in workbook.sheetnames:
        return safe_base_name

    counter = 1
    while True:
        suffix = f"_{counter}"
        candidate = f"{safe_base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1
