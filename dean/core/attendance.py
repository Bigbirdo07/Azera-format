"""Attendance ingestion and per-student metric computation.

Input: a long-format attendance file — one row per (Student ID, Date) with
columns Student ID, Date, Attendance Status, Excused/Unexcused, Tardy, plus
optional Period / Teacher / Course.

Output:
- ``load_attendance_file`` → ``LoadedAttendance`` (canonicalised frame + a
  list of warnings + the original openpyxl workbook for audit).
- ``match_attendance_to_roster`` → matched/unmatched counts so the UI can
  warn about Student IDs in attendance that aren't in the roster.
- ``compute_attendance_metrics`` → per-student DataFrame keyed by Student ID
  with Days Present / Absent / Tardy / Unexcused Absences / Attendance Rate
  / Recent Absences / Attendance Risk / Severe Attendance Risk.

Risk thresholds (per the spec):
- Attendance Risk: Attendance Rate < 90%
- Severe Attendance Risk: Attendance Rate < 80%
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


ATTENDANCE_RISK_THRESHOLD = 90.0  # < 90% → at risk
SEVERE_ATTENDANCE_RISK_THRESHOLD = 80.0  # < 80% → severe risk
RECENT_ABSENCE_WINDOW_DAYS = 14  # "Recent Absences" looks at the last N days

# Status normalisation — schools spell statuses many ways. We canonicalise to
# four buckets: Present / Absent / Tardy / Excused. Anything else lands in
# Other and is counted as Absent for safety (better to over-report risk than
# under-report).
_PRESENT_TOKENS = {"present", "p", "in", "in attendance", "attended", "here"}
_ABSENT_TOKENS = {"absent", "a", "out", "missing", "no show", "no-show", "ns"}
_TARDY_TOKENS = {"tardy", "late", "t", "tdy", "tardies"}
_EXCUSED_TOKENS = {"excused", "excused absence", "ex", "e",
                   "excused absent", "ea"}

# Column-name resolution: case- and space-insensitive matching for each
# canonical field. The roster's "Student ID" column also lives in
# `_STUDENT_ID_VARIANTS` so the same matcher can be re-used.
_STUDENT_ID_VARIANTS = ("student id", "studentid", "student number",
                        "id", "sid", "banner id")
_DATE_VARIANTS = ("date", "attendance date", "day", "school day")
_STATUS_VARIANTS = ("attendance status", "status", "attendance",
                    "attendance type", "type")
_EXCUSED_VARIANTS = ("excused", "excused/unexcused", "excuse",
                     "excused absence", "excused status")
_TARDY_VARIANTS = ("tardy", "late", "tardiness", "tardies")
_PERIOD_VARIANTS = ("period", "class period", "block")
_TEACHER_VARIANTS = ("teacher", "instructor", "teacher name", "professor")
_COURSE_VARIANTS = ("course", "class", "subject", "course name")


@dataclass(frozen=True)
class LoadedAttendance:
    file_name: str
    workbook: Workbook  # kept for audit; never mutated
    frame: pd.DataFrame  # canonicalised long-format
    warnings: list[str] = field(default_factory=list)


# Canonical names of the attendance fields we look for ON the roster sheet
# itself. The detector treats the presence of any of these as "the workbook
# already has attendance" — even one column (e.g. just `Days Absent`) is
# enough to enable threshold queries.
_INLINE_ATTENDANCE_CANONICALS = {
    "attendance_rate",
    "days_absent",
    "days_tardy",
    "unexcused_absences",
    "excused_absences",
    "attendance_status",
    "attendance_watch",
}

# Sheet names that look like a sibling "Attendance" sheet inside the same
# workbook. Matched case-insensitively against the normalised sheet name.
_ATTENDANCE_SHEET_NAME_HINTS = (
    "attendance", "daily attendance", "attendance log",
    "attendance roster", "absences",
)


@dataclass(frozen=True)
class AttendanceDetection:
    """Where attendance data lives inside the uploaded workbook.

    ``mode`` is one of:
      - "inline"   — the roster sheet already has attendance columns;
                     no recomputation is needed.
      - "sheet"    — a sibling sheet inside the same workbook holds the
                     long-format attendance data we should compute metrics
                     from and merge onto the roster.
      - "none"     — no attendance signal in the workbook.
    """
    mode: str
    inline_columns: tuple[str, ...] = ()
    attendance_sheet: str | None = None
    inline_fields: tuple[str, ...] = ()  # canonical field names, for UI


def detect_workbook_attendance(
    sheets: dict[str, pd.DataFrame],
    roster_sheet: str,
) -> AttendanceDetection:
    """Inspect an uploaded workbook for attendance data.

    Priority: inline columns on the roster sheet win first (the school has
    already calculated attendance); else a sibling sheet that looks like a
    long-format attendance log; else 'none'.
    """
    from core.schema import canonical_for

    inline_cols: list[str] = []
    inline_fields: list[str] = []
    roster_frame = sheets.get(roster_sheet)
    if roster_frame is not None and not roster_frame.empty:
        for column in roster_frame.columns:
            canonical = canonical_for(column)
            if canonical in _INLINE_ATTENDANCE_CANONICALS:
                inline_cols.append(column)
                inline_fields.append(canonical)
    if inline_cols:
        return AttendanceDetection(
            mode="inline",
            inline_columns=tuple(inline_cols),
            inline_fields=tuple(dict.fromkeys(inline_fields)),
        )

    # Look for a sibling sheet that smells like an attendance log.
    for sheet_name, frame in sheets.items():
        if sheet_name == roster_sheet or frame is None or frame.empty:
            continue
        if not _looks_like_attendance_sheet_name(sheet_name):
            continue
        if not _has_attendance_long_columns(frame):
            continue
        return AttendanceDetection(mode="sheet", attendance_sheet=sheet_name)

    return AttendanceDetection(mode="none")


def _looks_like_attendance_sheet_name(name: str) -> bool:
    norm = str(name).strip().lower()
    if not norm:
        return False
    return any(hint in norm for hint in _ATTENDANCE_SHEET_NAME_HINTS)


def _has_attendance_long_columns(frame: pd.DataFrame) -> bool:
    """A sibling sheet only counts as an attendance source when it has at
    least Student ID and one of (Attendance Status | Date) — protects against
    a "Notes" sheet getting mistaken for attendance because the title contains
    'attendance'."""
    sid = _find_column(frame.columns, _STUDENT_ID_VARIANTS)
    if sid is None:
        return False
    status = _find_column(frame.columns, _STATUS_VARIANTS)
    date = _find_column(frame.columns, _DATE_VARIANTS)
    return status is not None or date is not None


# ---- loading ---------------------------------------------------------------


def load_attendance_file(uploaded_file) -> LoadedAttendance:
    """Parse an uploaded .xlsx attendance file into a canonical long frame.

    Reads only the first non-empty sheet; multi-sheet attendance files are
    rare and ambiguous (which sheet is the source of truth?), so we keep
    that out of scope for now and surface a warning.
    """
    file_name = Path(getattr(uploaded_file, "name", "attendance.xlsx")).name
    file_bytes = uploaded_file.getvalue()
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)

    warnings: list[str] = []
    frame: pd.DataFrame | None = None
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if worksheet.sheet_state != "visible":
            continue
        candidate = _worksheet_to_dataframe(worksheet)
        if candidate is None or candidate.empty:
            continue
        if frame is None:
            frame = candidate
            if len(workbook.sheetnames) > 1:
                warnings.append(
                    f"Attendance file has {len(workbook.sheetnames)} sheets; "
                    f"only the first non-empty sheet ('{sheet_name}') is used."
                )
            break

    if frame is None:
        return LoadedAttendance(
            file_name=file_name, workbook=workbook, frame=pd.DataFrame(),
            warnings=warnings + ["Attendance file has no readable rows."],
        )

    canonical, more_warnings = canonicalise_attendance_frame(frame)
    warnings.extend(more_warnings)
    return LoadedAttendance(
        file_name=file_name, workbook=workbook,
        frame=canonical, warnings=warnings,
    )


def _worksheet_to_dataframe(worksheet) -> pd.DataFrame | None:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return None
    header = [_normalize_header(cell) for cell in rows[0]]
    body = [list(row) for row in rows[1:]]
    if not header or not body:
        return None
    frame = pd.DataFrame(body, columns=header)
    # Drop fully-blank columns and rows so a stray trailing column doesn't
    # poison the canonicaliser.
    frame = frame.dropna(axis=1, how="all").dropna(axis=0, how="all")
    return frame


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


# ---- canonicalisation ------------------------------------------------------


def canonicalise_attendance_frame(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Rename source columns to canonical names and coerce types.

    Returns (frame, warnings). Missing required columns become warnings, not
    exceptions — we still want to surface what we DID load.
    """
    warnings: list[str] = []
    rename = {}
    used: set[str] = set()

    def take(variants: tuple[str, ...]) -> str | None:
        col = _find_column(frame.columns, variants, excluded=used)
        if col is not None:
            used.add(col)
        return col

    # Order matters: the most-required fields claim columns first so an
    # ambiguous header like "Status" lands on Attendance Status, not Excused.
    sid = take(_STUDENT_ID_VARIANTS)
    date = take(_DATE_VARIANTS)
    status = take(_STATUS_VARIANTS)
    excused = take(_EXCUSED_VARIANTS)
    tardy = take(_TARDY_VARIANTS)
    period = take(_PERIOD_VARIANTS)
    teacher = take(_TEACHER_VARIANTS)
    course = take(_COURSE_VARIANTS)

    if sid is None:
        warnings.append("Attendance file is missing a Student ID column.")
    else:
        rename[sid] = "Student ID"
    if date is None:
        warnings.append("Attendance file is missing a Date column — "
                        "date-bounded asks ('absent today') will be unavailable.")
    else:
        rename[date] = "Date"
    if status is None:
        warnings.append("Attendance file is missing a Status column — "
                        "metrics will assume every row is an absence.")
    else:
        rename[status] = "Attendance Status"
    if excused is not None:
        rename[excused] = "Excused"
    if tardy is not None:
        rename[tardy] = "Tardy"
    if period is not None:
        rename[period] = "Period"
    if teacher is not None:
        rename[teacher] = "Teacher"
    if course is not None:
        rename[course] = "Course"

    out = frame.rename(columns=rename).copy()
    keep = [c for c in ("Student ID", "Date", "Attendance Status", "Excused",
                        "Tardy", "Period", "Teacher", "Course") if c in out.columns]
    out = out[keep]

    if "Student ID" in out.columns:
        out["Student ID"] = out["Student ID"].apply(_coerce_id)
        out = out[out["Student ID"].astype(str).str.strip() != ""]
    if "Date" in out.columns:
        out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
    if "Attendance Status" in out.columns:
        out["Attendance Status"] = out["Attendance Status"].apply(_canonicalise_status)
    if "Excused" in out.columns:
        out["Excused"] = out["Excused"].apply(_canonicalise_yes_no)
    if "Tardy" in out.columns:
        out["Tardy"] = out["Tardy"].apply(_canonicalise_yes_no)
    return out, warnings


def _find_column(
    columns, variants: tuple[str, ...], *, excluded: set[str] | None = None,
) -> str | None:
    excluded = excluded or set()
    norm_map = {str(c).strip().lower(): c for c in columns}
    # Exact normalised match (skipping any already claimed by a higher-
    # priority canonical).
    for variant in variants:
        if variant in norm_map and norm_map[variant] not in excluded:
            return norm_map[variant]
    # Loose containment fallback so "Student ID Number" still resolves.
    for variant in variants:
        for norm, original in norm_map.items():
            if original in excluded:
                continue
            if variant in norm or norm in variant:
                return original
    return None


def _coerce_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _canonicalise_status(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "Other"
    token = str(value).strip().lower()
    if token in _PRESENT_TOKENS:
        return "Present"
    if token in _TARDY_TOKENS:
        return "Tardy"
    if token in _EXCUSED_TOKENS:
        return "Excused"
    if token in _ABSENT_TOKENS:
        return "Absent"
    return "Other"


def _canonicalise_yes_no(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    token = str(value).strip().lower()
    if token in {"yes", "y", "true", "t", "1", "excused"}:
        return "Yes"
    if token in {"no", "n", "false", "f", "0", "unexcused"}:
        return "No"
    return None


# ---- matching --------------------------------------------------------------


def match_attendance_to_roster(
    attendance: pd.DataFrame, roster: pd.DataFrame
) -> tuple[int, int, list[str]]:
    """Return (matched_count, unmatched_count, unmatched_ids) measured at
    the unique-Student-ID level (NOT per row), since a counsellor wants to
    know "how many students don't appear in the roster," not "how many
    attendance rows failed to join."
    """
    if attendance.empty or "Student ID" not in attendance.columns:
        return 0, 0, []
    roster_id_col = _find_column(roster.columns, _STUDENT_ID_VARIANTS)
    if roster_id_col is None:
        return 0, 0, []
    attendance_ids = set(
        str(v).strip() for v in attendance["Student ID"].dropna().unique().tolist()
        if str(v).strip()
    )
    roster_ids = set(
        str(v).strip() for v in roster[roster_id_col].dropna().unique().tolist()
        if str(v).strip()
    )
    matched = attendance_ids & roster_ids
    unmatched = attendance_ids - roster_ids
    return len(matched), len(unmatched), sorted(unmatched)


# ---- metrics ---------------------------------------------------------------


def compute_attendance_metrics(
    attendance: pd.DataFrame,
    *,
    today: datetime | None = None,
    risk_threshold: float | None = None,
    severe_threshold: float | None = None,
) -> pd.DataFrame:
    """Per-student metrics frame keyed by Student ID.

    Counting rules (deliberately conservative — we want at-risk students to
    surface, not get buried by ambiguous statuses):
    - Present / Tardy / Excused / Absent are counted as labelled.
    - "Other" (anything we couldn't categorise) is counted as Absent.
    - Days Tardy and Days Present are mutually exclusive — a row tagged
      Tardy is NOT also counted as Present (a tardy IS an attendance issue).
    - Attendance Rate = (Present + Tardy + Excused) / total recorded days.
    - Unexcused Absences = Absent rows where Excused != "Yes".

    ``risk_threshold`` / ``severe_threshold`` override the module defaults
    when the caller passes configured ``RiskSettings``. When omitted, the
    module constants apply — so existing callers keep the historical
    behaviour without any code change.
    """
    risk = risk_threshold if risk_threshold is not None else ATTENDANCE_RISK_THRESHOLD
    severe = severe_threshold if severe_threshold is not None else SEVERE_ATTENDANCE_RISK_THRESHOLD
    if attendance.empty or "Student ID" not in attendance.columns:
        return pd.DataFrame(columns=[
            "Student ID", "Days Present", "Days Absent", "Days Tardy",
            "Unexcused Absences", "Attendance Rate",
            "Attendance Category", "Recent Absences", "Attendance Risk", "Severe Attendance Risk",
        ])

    status_col = "Attendance Status" if "Attendance Status" in attendance else None
    excused_col = "Excused" if "Excused" in attendance else None
    tardy_col = "Tardy" if "Tardy" in attendance else None
    date_col = "Date" if "Date" in attendance else None

    out_rows: list[dict[str, Any]] = []
    cutoff = None
    if today is None:
        today = datetime.now()
    cutoff = today - timedelta(days=RECENT_ABSENCE_WINDOW_DAYS)

    for sid, group in attendance.groupby("Student ID"):
        total = int(len(group))
        present = absent = tardy = excused = 0
        unexcused_absences = 0
        recent_absences = 0
        for _, row in group.iterrows():
            status = row[status_col] if status_col else "Absent"
            is_tardy_flag = (tardy_col and row.get(tardy_col) == "Yes")
            is_excused_flag = (excused_col and row.get(excused_col) == "Yes")

            if is_tardy_flag or status == "Tardy":
                tardy += 1
                continue
            if status == "Present":
                present += 1
                continue
            if status == "Excused" or is_excused_flag and status != "Absent":
                excused += 1
                continue
            # Everything left (Absent + Other) counts as absent.
            absent += 1
            if not is_excused_flag:
                unexcused_absences += 1
            if date_col:
                row_date = row.get(date_col)
                if isinstance(row_date, pd.Timestamp) and not pd.isna(row_date):
                    if row_date.to_pydatetime(warn=False).replace(microsecond=0) >= cutoff:
                        recent_absences += 1

        # Attendance Rate counts Present + Tardy + Excused as "attended"
        # (tardies don't zero the day out for rate-keeping purposes, just
        # for the Days Tardy column).
        attended_days = present + tardy + excused
        rate = round((attended_days / total) * 100, 2) if total > 0 else 0.0
        if rate < severe:
            category = "Severe Attendance Support"
        elif rate < risk:
            category = "Needs Attendance Support"
        else:
            category = "On Track"
        out_rows.append({
            "Student ID": sid,
            "Days Present": present,
            "Days Absent": absent,
            "Days Tardy": tardy,
            "Unexcused Absences": unexcused_absences,
            "Attendance Rate": rate,
            "Attendance Category": category,
            "Recent Absences": recent_absences,
            "Attendance Risk": rate < risk,
            "Severe Attendance Risk": rate < severe,
        })

    return pd.DataFrame(out_rows)
