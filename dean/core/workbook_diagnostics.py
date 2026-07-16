from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class SheetDiagnostics:
    name: str
    state: str
    is_hidden: bool
    is_protected: bool
    merged_cell_ranges: list[str]
    formula_count: int
    blank_row_count: int
    blank_column_count: int
    duplicate_column_names: list[str]
    likely_header_row: int | None
    table_count: int
    has_existing_filter: bool
    chart_count: int
    pivot_table_count: int
    data_quality_warnings: list[str] = field(default_factory=list)
    complex_formatting_warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class WorkbookDiagnostics:
    sheet_count: int
    hidden_sheets: list[str]
    sheets: list[SheetDiagnostics]


def diagnose_workbook(workbook: Workbook) -> WorkbookDiagnostics:
    sheet_diagnostics = [_diagnose_sheet(worksheet) for worksheet in workbook.worksheets]
    return WorkbookDiagnostics(
        sheet_count=len(workbook.worksheets),
        hidden_sheets=[sheet.name for sheet in sheet_diagnostics if sheet.is_hidden],
        sheets=sheet_diagnostics,
    )


def _diagnose_sheet(worksheet: Worksheet) -> SheetDiagnostics:
    likely_header_row = _detect_likely_header_row(worksheet)
    duplicate_columns = _duplicate_column_names(worksheet, likely_header_row)
    blank_rows = _blank_row_count(worksheet)
    blank_columns = _blank_column_count(worksheet)
    formula_count = _formula_count(worksheet)
    merged_ranges = [str(item) for item in worksheet.merged_cells.ranges]
    table_count = len(worksheet.tables)
    has_filter = bool(worksheet.auto_filter and worksheet.auto_filter.ref)
    chart_count = len(getattr(worksheet, "_charts", []))
    pivot_count = len(getattr(worksheet, "_pivots", []))

    data_warnings: list[str] = []
    complex_warnings: list[str] = []

    if likely_header_row is None:
        data_warnings.append("Could not confidently detect a header row.")
    elif likely_header_row != 1:
        data_warnings.append(f"Likely header row is row {likely_header_row}, not row 1.")

    if duplicate_columns:
        data_warnings.append(f"Duplicate column names detected: {', '.join(duplicate_columns)}.")
    if blank_rows:
        data_warnings.append(f"{blank_rows} blank rows detected.")
    if blank_columns:
        data_warnings.append(f"{blank_columns} blank columns detected.")

    if merged_ranges:
        complex_warnings.append("Merged cells detected.")
    if worksheet.protection.sheet:
        complex_warnings.append("Sheet protection is enabled.")
    if table_count:
        complex_warnings.append("Excel tables detected.")
    if has_filter:
        complex_warnings.append("Existing filters detected.")
    if chart_count:
        complex_warnings.append("Existing charts detected.")
    if pivot_count:
        complex_warnings.append("Pivot tables detected.")
    if formula_count:
        complex_warnings.append("Existing formulas detected.")

    return SheetDiagnostics(
        name=worksheet.title,
        state=worksheet.sheet_state,
        is_hidden=worksheet.sheet_state != "visible",
        is_protected=bool(worksheet.protection.sheet),
        merged_cell_ranges=merged_ranges,
        formula_count=formula_count,
        blank_row_count=blank_rows,
        blank_column_count=blank_columns,
        duplicate_column_names=duplicate_columns,
        likely_header_row=likely_header_row,
        table_count=table_count,
        has_existing_filter=has_filter,
        chart_count=chart_count,
        pivot_table_count=pivot_count,
        data_quality_warnings=data_warnings,
        complex_formatting_warnings=complex_warnings,
    )


def _detect_likely_header_row(worksheet: Worksheet, scan_rows: int = 15) -> int | None:
    first_row_values = [
        _cell_value(worksheet.cell(row=1, column=column))
        for column in range(1, worksheet.max_column + 1)
    ]
    first_row_non_blank = [
        value
        for value in first_row_values
        if value not in (None, "")
    ]

    # Most university exports put real headers on row 1. Keep that as the
    # default unless row 1 is blank or looks like a single title/metadata cell.
    if len(first_row_non_blank) >= 2:
        return 1

    best_row: int | None = None
    best_score = 0
    max_row = min(worksheet.max_row, scan_rows)
    for row_index in range(1, max_row + 1):
        values = [_cell_value(worksheet.cell(row=row_index, column=column)) for column in range(1, worksheet.max_column + 1)]
        non_blank = [value for value in values if value not in (None, "")]
        if not non_blank:
            continue
        text_count = sum(isinstance(value, str) for value in non_blank)
        unique_count = len(set(str(value).strip() for value in non_blank))
        duplicate_penalty = len(non_blank) - unique_count
        next_row_values = []
        if row_index < worksheet.max_row:
            next_row_values = [
                _cell_value(worksheet.cell(row=row_index + 1, column=column))
                for column in range(1, worksheet.max_column + 1)
            ]
        next_row_non_blank = sum(value not in (None, "") for value in next_row_values)
        score = (text_count * 2) + unique_count + next_row_non_blank - (duplicate_penalty * 3)
        if score > best_score:
            best_row = row_index
            best_score = score
    return best_row


def _duplicate_column_names(worksheet: Worksheet, header_row: int | None) -> list[str]:
    if header_row is None:
        return []
    names = [
        str(_cell_value(worksheet.cell(row=header_row, column=column))).strip()
        for column in range(1, worksheet.max_column + 1)
        if _cell_value(worksheet.cell(row=header_row, column=column)) not in (None, "")
    ]
    duplicates = sorted({name for name in names if names.count(name) > 1})
    return duplicates


def _blank_row_count(worksheet: Worksheet) -> int:
    count = 0
    for row in worksheet.iter_rows():
        if all(_cell_value(cell) in (None, "") for cell in row):
            count += 1
    return count


def _blank_column_count(worksheet: Worksheet) -> int:
    count = 0
    for column in worksheet.iter_cols():
        if all(_cell_value(cell) in (None, "") for cell in column):
            count += 1
    return count


def _formula_count(worksheet: Worksheet) -> int:
    return sum(
        1
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


def _cell_value(cell: Any) -> Any:
    return cell.value
