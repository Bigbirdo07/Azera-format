from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import re
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
MAX_EXAMPLE_ROWS = 10

COLUMN_CONCEPTS = {
    "student_id": ["student id", "studentid", "student number", "id number", "banner id", "emplid"],
    "student_name": ["student name", "name", "full name", "student"],
    "email": ["email", "email address", "student email", "e-mail"],
    "advisor": ["advisor", "academic advisor", "assigned advisor"],
    "program": ["program", "major", "degree program", "academic program", "plan"],
    "fafsa_status": ["fafsa status", "fafsa", "financial aid status"],
    "enrollment_status": ["enrollment status", "status", "student status"],
    "registration_status": ["registration status", "registered status", "registration"],
    "balance_due": ["balance due", "balance", "amount due", "student balance", "outstanding balance"],
    "registered_credits": ["registered credits", "credits", "credit hours", "enrolled credits"],
}


@dataclass(frozen=True)
class DataQualityResult:
    message: str
    preview: pd.DataFrame
    result_sheet: str


def create_data_quality_report(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> DataQualityResult:
    sheet_name = command["sheet"]
    dataframe = sheets[sheet_name].copy()
    column_map = _match_columns(dataframe.columns)
    issues = _run_checks(dataframe, column_map)

    report = pd.DataFrame(
        issues,
        columns=[
            "issue_type",
            "affected_column",
            "affected_row_count",
            "example_row_numbers",
            "severity",
            "recommended_fix",
        ],
    )
    if report.empty:
        report = pd.DataFrame(
            [
                {
                    "issue_type": "No issues detected",
                    "affected_column": "",
                    "affected_row_count": 0,
                    "example_row_numbers": "",
                    "severity": "low",
                    "recommended_fix": "No data quality fixes are recommended based on the current checks.",
                }
            ]
        )

    output_sheet = _unique_sheet_name(
        workbook,
        str(command.get("output_sheet") or "Data Quality Report"),
    )
    worksheet = workbook.create_sheet(output_sheet)
    _write_report_sheet(worksheet, report, source_sheet=sheet_name)

    return DataQualityResult(
        message=f"Created Data Quality Report with {len(report)} issue summary rows.",
        preview=report,
        result_sheet=output_sheet,
    )


def _run_checks(dataframe: pd.DataFrame, column_map: dict[str, str]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []

    _check_missing_column(issues, column_map, "student_id", "Student ID", "high")
    _check_missing_column(issues, column_map, "student_name", "student name", "high")
    _check_missing_column(issues, column_map, "email", "email", "medium")
    _check_missing_column(issues, column_map, "advisor", "advisor", "medium")
    _check_missing_column(issues, column_map, "program", "program/major", "medium")
    _check_missing_column(issues, column_map, "fafsa_status", "FAFSA status", "medium")
    _check_missing_column(issues, column_map, "enrollment_status", "enrollment status", "high")

    _check_missing_values(issues, dataframe, column_map, "student_id", "Missing Student ID", "high", "Enter the correct Student ID before importing or reporting.")
    _check_missing_values(issues, dataframe, column_map, "student_name", "Missing student name", "high", "Enter the official student name or reconcile this row with the student system.")
    _check_missing_values(issues, dataframe, column_map, "email", "Missing email", "medium", "Add a valid institutional or preferred email address.")
    _check_missing_values(issues, dataframe, column_map, "advisor", "Missing advisor", "medium", "Assign an advisor or confirm that no advisor is required.")
    _check_missing_values(issues, dataframe, column_map, "program", "Missing program/major", "medium", "Add the student's current program or major.")
    _check_missing_values(issues, dataframe, column_map, "fafsa_status", "Missing FAFSA status", "medium", "Update FAFSA status from the financial aid source of record.")
    _check_missing_values(issues, dataframe, column_map, "enrollment_status", "Missing enrollment status", "high", "Update enrollment status from the registrar source of record.")

    _check_duplicate_student_ids(issues, dataframe, column_map)
    _check_invalid_email(issues, dataframe, column_map)
    _check_balance_values(issues, dataframe, column_map)
    _check_credit_status_conflicts(issues, dataframe, column_map)
    _check_duplicate_rows(issues, dataframe)
    _check_blank_required_fields(issues, dataframe, column_map)

    severity_order = {"high": 0, "medium": 1, "low": 2}
    return sorted(
        issues,
        key=lambda item: (severity_order.get(str(item["severity"]), 9), str(item["issue_type"])),
    )


def _check_missing_column(
    issues: list[dict[str, Any]],
    column_map: dict[str, str],
    concept: str,
    display_name: str,
    severity: str,
) -> None:
    if concept in column_map:
        return
    issues.append(
        _issue(
            issue_type=f"Missing expected column: {display_name}",
            affected_column=display_name,
            mask=None,
            severity=severity,
            recommended_fix=f"Map or add a column for {display_name}.",
        )
    )


def _check_missing_values(
    issues: list[dict[str, Any]],
    dataframe: pd.DataFrame,
    column_map: dict[str, str],
    concept: str,
    issue_type: str,
    severity: str,
    recommended_fix: str,
) -> None:
    column = column_map.get(concept)
    if not column:
        return
    mask = _blank_mask(dataframe[column])
    _append_if_any(issues, issue_type, column, mask, severity, recommended_fix)


def _check_duplicate_student_ids(
    issues: list[dict[str, Any]],
    dataframe: pd.DataFrame,
    column_map: dict[str, str],
) -> None:
    column = column_map.get("student_id")
    if not column:
        return
    normalized = dataframe[column].astype(str).str.strip().str.casefold()
    nonblank = ~_blank_mask(dataframe[column])
    mask = normalized.duplicated(keep=False) & nonblank
    _append_if_any(
        issues,
        "Duplicate Student ID values",
        column,
        mask,
        "high",
        "Review duplicate IDs and merge, correct, or remove duplicate student records.",
    )


def _check_invalid_email(
    issues: list[dict[str, Any]],
    dataframe: pd.DataFrame,
    column_map: dict[str, str],
) -> None:
    column = column_map.get("email")
    if not column:
        return
    nonblank = ~_blank_mask(dataframe[column])
    mask = nonblank & ~dataframe[column].astype(str).str.strip().str.match(EMAIL_PATTERN)
    _append_if_any(
        issues,
        "Invalid email format",
        column,
        mask,
        "medium",
        "Correct email addresses so they follow a standard name@example.edu format.",
    )


def _check_balance_values(
    issues: list[dict[str, Any]],
    dataframe: pd.DataFrame,
    column_map: dict[str, str],
) -> None:
    column = column_map.get("balance_due")
    if not column:
        return
    values = dataframe[column]
    numeric = pd.to_numeric(values, errors="coerce")
    nonblank = ~_blank_mask(values)

    _append_if_any(
        issues,
        "Negative balance values",
        column,
        numeric < 0,
        "medium",
        "Confirm whether negative balances are credits/refunds or correct the balance value.",
    )
    _append_if_any(
        issues,
        "Nonnumeric balance values",
        column,
        nonblank & numeric.isna(),
        "high",
        "Replace text or symbols with numeric balance values.",
    )


def _check_credit_status_conflicts(
    issues: list[dict[str, Any]],
    dataframe: pd.DataFrame,
    column_map: dict[str, str],
) -> None:
    status_column = column_map.get("enrollment_status")
    credits_column = column_map.get("registered_credits")
    if not status_column or not credits_column:
        return

    status = dataframe[status_column].astype(str).str.strip().str.casefold()
    credits = pd.to_numeric(dataframe[credits_column], errors="coerce").fillna(0)
    active_raw = status.str.contains(r"\bactive\b", regex=True, na=False)
    active = active_raw & ~status.str.contains("inactive|withdraw", regex=True, na=False)
    inactive = status.str.contains("inactive", na=False)
    withdrawn = status.str.contains("withdraw", na=False)

    _append_if_any(
        issues,
        "Active students with 0 registered credits",
        f"{status_column}; {credits_column}",
        active & (credits == 0),
        "medium",
        "Confirm active students are registered or update enrollment status.",
    )
    _append_if_any(
        issues,
        "Inactive students with registered credits",
        f"{status_column}; {credits_column}",
        inactive & (credits > 0),
        "high",
        "Confirm whether inactive students should remain registered for credits.",
    )
    _append_if_any(
        issues,
        "Withdrawn students with active status conflicts",
        status_column,
        withdrawn & active_raw,
        "high",
        "Resolve conflicting withdrawn and active status indicators.",
    )

    registration_column = column_map.get("registration_status")
    if registration_column and registration_column != status_column:
        registration = dataframe[registration_column].astype(str).str.strip().str.casefold()
        registration_active = registration.str.contains(r"\bactive\b|registered", regex=True, na=False)
        _append_if_any(
            issues,
            "Withdrawn students with active status conflicts",
            f"{status_column}; {registration_column}",
            withdrawn & registration_active,
            "high",
            "Resolve conflicts between enrollment and registration status.",
        )


def _check_duplicate_rows(issues: list[dict[str, Any]], dataframe: pd.DataFrame) -> None:
    if dataframe.empty:
        return
    mask = dataframe.duplicated(keep=False)
    _append_if_any(
        issues,
        "Duplicate rows",
        "All columns",
        mask,
        "medium",
        "Review duplicate rows and remove only confirmed duplicates.",
    )


def _check_blank_required_fields(
    issues: list[dict[str, Any]],
    dataframe: pd.DataFrame,
    column_map: dict[str, str],
) -> None:
    required_concepts = [
        "student_id",
        "student_name",
        "email",
        "advisor",
        "program",
        "fafsa_status",
        "enrollment_status",
    ]
    required_columns = [column_map[concept] for concept in required_concepts if concept in column_map]
    if not required_columns:
        return

    mask = pd.Series(False, index=dataframe.index)
    for column in required_columns:
        mask |= _blank_mask(dataframe[column])
    _append_if_any(
        issues,
        "Blank required fields",
        "; ".join(required_columns),
        mask,
        "high",
        "Fill required fields or confirm why the values are unavailable.",
    )


def _append_if_any(
    issues: list[dict[str, Any]],
    issue_type: str,
    affected_column: str,
    mask: pd.Series,
    severity: str,
    recommended_fix: str,
) -> None:
    if int(mask.sum()) == 0:
        return
    issues.append(_issue(issue_type, affected_column, mask, severity, recommended_fix))


def _issue(
    issue_type: str,
    affected_column: str,
    mask: pd.Series | None,
    severity: str,
    recommended_fix: str,
) -> dict[str, Any]:
    if mask is None:
        row_count = 0
        examples = ""
    else:
        row_count = int(mask.sum())
        examples = ", ".join(str(row_number) for row_number in _example_excel_rows(mask))
    return {
        "issue_type": issue_type,
        "affected_column": affected_column,
        "affected_row_count": row_count,
        "example_row_numbers": examples,
        "severity": severity,
        "recommended_fix": recommended_fix,
    }


def _example_excel_rows(mask: pd.Series) -> list[int]:
    rows = []
    for index in mask[mask].index[:MAX_EXAMPLE_ROWS]:
        try:
            rows.append(int(index) + 2)
        except (TypeError, ValueError):
            rows.append(len(rows) + 2)
    return rows


def _blank_mask(series: pd.Series) -> pd.Series:
    return series.isna() | (series.astype(str).str.strip() == "")


def _match_columns(columns: pd.Index) -> dict[str, str]:
    normalized_columns = {_normalize(column): str(column) for column in columns}
    matched: dict[str, str] = {}
    for concept, aliases in COLUMN_CONCEPTS.items():
        for alias in aliases:
            normalized_alias = _normalize(alias)
            if normalized_alias in normalized_columns:
                matched[concept] = normalized_columns[normalized_alias]
                break
        if concept in matched:
            continue
        for normalized_column, original_column in normalized_columns.items():
            if any(_normalize(alias) in normalized_column for alias in aliases):
                matched[concept] = original_column
                break
    return matched


def _normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).casefold()).strip()


def _write_report_sheet(
    worksheet: Worksheet,
    dataframe: pd.DataFrame,
    *,
    source_sheet: str,
) -> None:
    worksheet["A1"] = "Data Quality Report"
    worksheet["A1"].font = Font(bold=True, size=16, color="1F4E78")
    worksheet["A2"] = f"Source sheet: {source_sheet}"
    worksheet["A3"] = f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    worksheet["A4"] = "This report stores issue summaries and row numbers only, not student row contents."

    start_row = 6
    headers = [str(column) for column in dataframe.columns]
    worksheet.append([])
    worksheet.append(headers)
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    worksheet.freeze_panes = f"A{start_row + 1}"
    worksheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{worksheet.max_row}"
    for row in worksheet.iter_rows(min_row=start_row + 1):
        severity_cell = row[4]
        if severity_cell.value == "high":
            severity_cell.fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        elif severity_cell.value == "medium":
            severity_cell.fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
        elif severity_cell.value == "low":
            severity_cell.fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def _unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    safe_base_name = base_name[:31] or "Data Quality Report"
    if safe_base_name not in workbook.sheetnames:
        return safe_base_name

    counter = 1
    while True:
        suffix = f"_{counter}"
        candidate = f"{safe_base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1
