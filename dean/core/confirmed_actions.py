"""Confirmed, local-only workbook actions.

These run ONLY after the user confirms a pending action. They never modify the
original upload or the in-memory workbook — they operate on copies of the sheet
DataFrames and write a brand-new timestamped file under the outputs directory.
Every confirmed action appends a privacy-safe record to the audit log (filters,
counts, columns, output path — never raw student rows).
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.field_policy import field_status
from core.privacy import classify_sensitivity
from core.query_engine import QueryResult, _build_mask, run_query
from core.schema import canonical_for

# Module-level defaults; tests monkeypatch these to temporary directories.
DEFAULT_OUTPUT_DIR = Path("outputs")
DEFAULT_AUDIT_PATH = Path("logs/audit_log.jsonl")


@dataclass
class ActionResult:
    success: bool
    action_type: str
    rows_affected: int = 0
    output_file: str | None = None
    message: str = ""
    columns: list[str] = field(default_factory=list)
    error: str | None = None


# --- helpers -----------------------------------------------------------------


def get_target_rows_mask(frame: pd.DataFrame, filters: list[dict[str, Any]]) -> pd.Series:
    if not filters:
        return pd.Series(True, index=frame.index)
    return _build_mask(frame, filters)


def append_note(existing_value: Any, new_note: str) -> str:
    existing = "" if existing_value is None else str(existing_value).strip()
    if existing.lower() in {"", "nan", "none"}:
        return new_note
    return f"{existing}; {new_note}"


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _sanitize(text: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", str(text).lower()).strip("_")
    return cleaned[:40] or "output"


def _unique_path(output_dir: Path, prefix: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    base = f"{_sanitize(prefix)}_{_timestamp()}"
    candidate = output_dir / f"{base}.xlsx"
    counter = 1
    while candidate.exists():
        candidate = output_dir / f"{base}_{counter}.xlsx"
        counter += 1
    return candidate


def write_output_workbook(
    sheets: dict[str, pd.DataFrame],
    prefix: str,
    output_dir: Path | None = None,
) -> Path:
    """Write the given sheets to a NEW timestamped workbook. Never overwrites."""
    path = _unique_path(Path(output_dir or DEFAULT_OUTPUT_DIR), prefix)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=str(name)[:31], index=False)
    return path


def _write_output_workbook_with_formatting(
    sheets: dict[str, pd.DataFrame],
    prefix: str,
    output_dir: Path | None = None,
    chart_specs: list[dict[str, Any]] | None = None,
) -> Path:
    path = write_output_workbook(sheets, prefix, output_dir)
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        _format_worksheet(worksheet)
    for spec in chart_specs or []:
        _add_bar_chart(workbook, spec)
    workbook.save(path)
    return path


def _format_worksheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    if worksheet.max_row >= 1:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _add_bar_chart(workbook, spec: dict[str, Any]) -> None:
    sheet_name = spec.get("sheet")
    if sheet_name not in workbook.sheetnames:
        return
    worksheet = workbook[sheet_name]
    if worksheet.max_row < 2 or worksheet.max_column < 2:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = spec.get("title") or sheet_name
    chart.y_axis.title = spec.get("value_title") or worksheet.cell(row=1, column=2).value
    chart.x_axis.title = spec.get("category_title") or worksheet.cell(row=1, column=1).value
    data = Reference(worksheet, min_col=2, min_row=1, max_row=worksheet.max_row)
    cats = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 14
    worksheet.add_chart(chart, spec.get("anchor") or "E2")


def _sensitive_fields(columns: list[str]) -> list[str]:
    return [c for c in columns if classify_sensitivity(c)[0]]


def build_audit_record(
    *,
    action_type: str,
    target_sheet: str,
    filters: list[dict[str, Any]],
    rows_affected: int,
    columns: list[str],
    output_file: str | None,
    sensitive_fields: list[str],
    confirmation_status: str,
    request_summary: str,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    record = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "action_type": action_type,
        "target_sheet": target_sheet,
        "filters_applied": filters,
        "rows_affected": rows_affected,
        "columns_changed": columns,
        "output_file": output_file,
        "sensitive_fields_involved": sensitive_fields,
        "confirmation_status": confirmation_status,
        "user_request_summary": (request_summary or "")[:200],
    }
    if extra:
        record.update(extra)
    return record


_AUDIT_SUPPRESS = object()  # Sentinel: chain steps pass this to skip their own audit.


def record_audit(record: dict[str, Any], audit_path: "Path | None | object" = None) -> None:
    """Append one audit record. ``audit_path=_AUDIT_SUPPRESS`` disables logging
    for chained inner actions (the chain writes one combined entry instead).
    ``None`` falls back to ``DEFAULT_AUDIT_PATH``."""
    if audit_path is _AUDIT_SUPPRESS:
        return
    return _record_audit_impl(record, audit_path)


def _record_audit_impl(record: dict[str, Any], audit_path: Path | None = None) -> None:
    path = Path(audit_path or DEFAULT_AUDIT_PATH)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record) + "\n")


def _resolve_column(frame: pd.DataFrame, field_name: str) -> str | None:
    """Find an existing column matching a (possibly canonical) field name."""
    target = field_name.strip().casefold()
    for column in frame.columns:
        if str(column).strip().casefold() == target:
            return column
    canon = canonical_for(field_name)
    if canon:
        for column in frame.columns:
            if canonical_for(str(column)) == canon:
                return column
    return None


# --- actions -----------------------------------------------------------------


def execute_export_action(
    *,
    filters: list[dict[str, Any]],
    sheets: dict[str, pd.DataFrame],
    sheet: str,
    request_summary: str = "",
    output_dir: Path | None = None,
    audit_path: Path | None = None,
) -> ActionResult:
    frame = sheets[sheet]
    mask = get_target_rows_mask(frame, filters)
    subset = frame.loc[mask].copy().reset_index(drop=True)
    path = write_output_workbook({"Export": subset}, "student_export", output_dir)
    columns = list(subset.columns)
    result = ActionResult(
        success=True,
        action_type="export",
        rows_affected=int(len(subset)),
        output_file=str(path),
        columns=columns,
        message=f"Exported {len(subset)} record(s) to a new file: {path}",
    )
    record_audit(
        build_audit_record(
            action_type="export",
            target_sheet=sheet,
            filters=filters,
            rows_affected=result.rows_affected,
            columns=columns,
            output_file=str(path),
            sensitive_fields=_sensitive_fields(columns),
            confirmation_status="confirmed",
            request_summary=request_summary,
        ),
        audit_path,
    )
    return result


def execute_dashboard_report_action(
    *,
    sheets: dict[str, pd.DataFrame],
    sheet: str,
    request_summary: str = "",
    output_dir: Path | None = None,
    audit_path: Path | None = None,
) -> ActionResult:
    """Create a local advisor/intervention dashboard workbook.

    The dashboard is deterministic: pandas computes every table, and the
    original upload is never modified.
    """
    frame = sheets[sheet]

    intervention = _safe_query({
        "operation": "student_intervention_summary",
        "sheet": sheet,
        "filters": [],
        "limit": 100,
    }, sheets)
    advisors_best = _safe_query({
        "operation": "advisor_outcome_summary",
        "sheet": sheet,
        "filters": [],
        "group_by": _first_existing(frame, ["Advisor", "Teacher", "Professor"]) or "",
        "sort": {"column": "Outcome Score", "direction": "desc"},
        "limit": None,
    }, sheets)
    advisors_support = _safe_query({
        "operation": "advisor_outcome_summary",
        "sheet": sheet,
        "filters": [],
        "group_by": _first_existing(frame, ["Advisor", "Teacher", "Professor"]) or "",
        "sort": {"column": "Outcome Score", "direction": "asc"},
        "limit": None,
    }, sheets)
    standing_by_advisor = _safe_query({
        "operation": "pivot_table_summary",
        "sheet": sheet,
        "filters": [],
        "group_by": _first_existing(frame, ["Advisor", "Teacher", "Professor"]) or "",
        "pivot_rows": _first_existing(frame, ["Advisor", "Teacher", "Professor"]) or "",
        "pivot_columns": _first_existing(frame, ["Standing", "Academic Standing", "Academic Status"]) or "",
        "metric": "count",
    }, sheets)
    gpa_by_advisor_year = _safe_query({
        "operation": "pivot_table_summary",
        "sheet": sheet,
        "filters": [],
        "group_by": _first_existing(frame, ["Advisor", "Teacher", "Professor"]) or "",
        "pivot_rows": _first_existing(frame, ["Advisor", "Teacher", "Professor"]) or "",
        "pivot_columns": _first_existing(frame, ["Year", "Grade"]) or "",
        "value_column": _first_existing(frame, ["GPA"]) or "",
        "metric": "average",
    }, sheets)
    trends = _safe_query({
        "operation": "trend_summary",
        "sheet": sheet,
        "filters": [],
    }, sheets)

    dashboard_sheets: dict[str, pd.DataFrame] = {
        "Dashboard Summary": _dashboard_summary_frame(
            source_rows=len(frame.index),
            intervention=intervention,
            advisors_best=advisors_best,
            advisors_support=advisors_support,
            trends=trends,
        ),
        "Intervention Review": _result_frame(intervention),
        "Advisor Outcomes": _result_frame(advisors_best),
        "Advisors Need Support": _result_frame(advisors_support),
        "Standing by Advisor": _result_frame(standing_by_advisor),
        "GPA by Advisor Year": _result_frame(gpa_by_advisor_year),
        "Trend Findings": _result_frame(trends),
    }
    chart_data = _dashboard_chart_data(frame)
    dashboard_sheets.update(chart_data)

    path = _write_output_workbook_with_formatting(
        dashboard_sheets,
        "advisor_intervention_dashboard",
        output_dir,
        chart_specs=[
            {"sheet": "Chart Students by Standing", "title": "Students by Standing"},
            {"sheet": "Chart Avg GPA by Major", "title": "Average GPA by Major", "value_title": "Average GPA"},
            {"sheet": "Chart Students by Advisor", "title": "Students by Advisor"},
        ],
    )
    columns = list(frame.columns)
    rows_affected = int(len(frame.index))
    result = ActionResult(
        success=True,
        action_type="dashboard_report",
        rows_affected=rows_affected,
        output_file=str(path),
        columns=columns,
        message=f"Created advisor intervention dashboard workbook: {path}",
    )
    record_audit(
        build_audit_record(
            action_type="dashboard_report",
            target_sheet=sheet,
            filters=[],
            rows_affected=rows_affected,
            columns=columns,
            output_file=str(path),
            sensitive_fields=_sensitive_fields(columns),
            confirmation_status="confirmed",
            request_summary=request_summary,
            extra={
                "dashboard_sheets": list(dashboard_sheets),
                "intervention_rows": intervention.row_count if intervention else 0,
            },
        ),
        audit_path,
    )
    return result


def _safe_query(query: dict[str, Any], sheets: dict[str, pd.DataFrame]) -> QueryResult | None:
    try:
        return run_query(query, sheets)
    except Exception:
        return None


def _result_frame(result: QueryResult | None) -> pd.DataFrame:
    if result is None:
        return pd.DataFrame([{"Status": "Could not compute this table from the available workbook columns."}])
    if result.table:
        return pd.DataFrame(result.table)
    return pd.DataFrame([{"Description": result.description, "Value": result.value, "Rows": result.row_count}])


def _dashboard_summary_frame(
    *,
    source_rows: int,
    intervention: QueryResult | None,
    advisors_best: QueryResult | None,
    advisors_support: QueryResult | None,
    trends: QueryResult | None,
) -> pd.DataFrame:
    best = advisors_best.table[0] if advisors_best and advisors_best.table else {}
    support = advisors_support.table[0] if advisors_support and advisors_support.table else {}
    top_trend = trends.table[0] if trends and trends.table else {}
    advisor_key = next((key for key in ("Advisor", "Teacher", "Professor") if key in best or key in support), "Advisor")
    rows = [
        {"Metric": "Source rows", "Value": source_rows, "Notes": "Rows scanned from the selected sheet."},
        {
            "Metric": "Students for intervention review",
            "Value": intervention.row_count if intervention else 0,
            "Notes": "Metric-based review list; not a final intervention decision.",
        },
        {
            "Metric": "Highest advisor outcome score",
            "Value": best.get(advisor_key, ""),
            "Notes": f"Score: {best.get('Outcome Score', '')}",
        },
        {
            "Metric": "Lowest advisor outcome score",
            "Value": support.get(advisor_key, ""),
            "Notes": f"Score: {support.get('Outcome Score', '')}",
        },
        {
            "Metric": "Top trend",
            "Value": top_trend.get("Metric", ""),
            "Notes": (
                f"{top_trend.get('Group By', '')}: "
                f"{top_trend.get('Highest Group', '')} vs {top_trend.get('Lowest Group', '')}"
            ),
        },
    ]
    return pd.DataFrame(rows)


def _dashboard_chart_data(frame: pd.DataFrame) -> dict[str, pd.DataFrame]:
    output: dict[str, pd.DataFrame] = {}
    standing = _first_existing(frame, ["Standing", "Academic Standing", "Academic Status"])
    if standing:
        output["Chart Students by Standing"] = (
            frame[standing].fillna("(missing)").astype(str).value_counts()
            .rename_axis(standing)
            .reset_index(name="Count")
        )
    major = _first_existing(frame, ["Major", "Program", "Department", "Discipline"])
    gpa = _first_existing(frame, ["GPA"])
    if major and gpa:
        numeric = pd.to_numeric(frame[gpa], errors="coerce")
        output["Chart Avg GPA by Major"] = (
            frame.assign(**{gpa: numeric})
            .groupby(major, dropna=False)[gpa]
            .mean()
            .round(4)
            .reset_index(name="Average GPA")
            .sort_values("Average GPA", ascending=False)
        )
    advisor = _first_existing(frame, ["Advisor", "Teacher", "Professor"])
    if advisor:
        output["Chart Students by Advisor"] = (
            frame[advisor].fillna("(missing)").astype(str).value_counts()
            .rename_axis(advisor)
            .reset_index(name="Count")
        )
    return output


def _first_existing(frame: pd.DataFrame, candidates: list[str]) -> str | None:
    targets = {candidate.strip().casefold(): candidate for candidate in candidates}
    for column in frame.columns:
        if str(column).strip().casefold() in targets:
            return column
    return None


def execute_add_note_action(
    *,
    filters: list[dict[str, Any]],
    note: str,
    sheets: dict[str, pd.DataFrame],
    sheet: str,
    request_summary: str = "",
    output_dir: Path | None = None,
    audit_path: Path | None = None,
) -> ActionResult:
    if not note.strip():
        return ActionResult(success=False, action_type="note_edit", error="empty_note",
                            message="What note should I add? Try: add note: Advisor follow-up needed.")
    frame = sheets[sheet]
    notes_col = _resolve_column(frame, "Notes") or "Notes"
    modified = frame.copy()
    if notes_col not in modified.columns:
        modified[notes_col] = ""
    mask = get_target_rows_mask(modified, filters)
    modified.loc[mask, notes_col] = modified.loc[mask, notes_col].apply(lambda v: append_note(v, note))

    output_sheets = dict(sheets)
    output_sheets[sheet] = modified
    path = write_output_workbook(output_sheets, "student_records_modified", output_dir)
    rows = int(mask.sum())
    result = ActionResult(
        success=True,
        action_type="note_edit",
        rows_affected=rows,
        output_file=str(path),
        columns=[notes_col],
        message=f"Added the note to {rows} matching record(s). A new workbook was saved: {path}",
    )
    record_audit(
        build_audit_record(
            action_type="add_note",
            target_sheet=sheet,
            filters=filters,
            rows_affected=rows,
            columns=[notes_col],
            output_file=str(path),
            sensitive_fields=_sensitive_fields([notes_col]),
            confirmation_status="confirmed",
            request_summary=request_summary,
        ),
        audit_path,
    )
    return result


def execute_action_chain(
    *,
    actions: list[dict[str, Any]],
    filters: list[dict[str, Any]],
    sheets: dict[str, pd.DataFrame],
    sheet: str,
    request_summary: str = "",
    output_dir: Path | None = None,
    audit_path: Path | None = None,
) -> ActionResult:
    """Phase Q: run an edit-then-export chain in a single confirmed pass.

    Edits run first (academic_watch / note_edit / field_update). The output
    workbook is written ONCE — if the chain ends with an export step, the
    export simply points at the file the edit just produced. Audit gets one
    'action_chain' entry summarizing the whole sequence so reviewers see the
    workflow as a unit.
    """
    if not actions:
        return ActionResult(success=False, action_type="action_chain",
                            error="empty_chain",
                            message="No actions to run.")

    edit_steps = [a for a in actions if a.get("type") != "export"]
    if len(edit_steps) > 1:
        # The Phase Q scope is single-edit + optional export; chains with
        # multiple edits would need a dedicated review UX. Fail safely.
        return ActionResult(success=False, action_type="action_chain",
                            error="too_many_edits",
                            message="Only one edit + export per chain is supported right now.")
    has_export = any(a.get("type") == "export" for a in actions)
    if not edit_steps and not has_export:
        return ActionResult(success=False, action_type="action_chain",
                            error="no_actionable_steps",
                            message="The chain has no executable steps.")

    edit = edit_steps[0] if edit_steps else None
    rows_affected = 0
    columns_changed: list[str] = []
    output_file: str | None = None

    if edit is None:
        # Pure export. Mirror the existing export action.
        export_result = execute_export_action(
            filters=filters, sheets=sheets, sheet=sheet,
            request_summary=request_summary,
            output_dir=output_dir, audit_path=_AUDIT_SUPPRESS,
        )
        rows_affected = export_result.rows_affected
        columns_changed = export_result.columns
        output_file = export_result.output_file
    else:
        step_type = edit.get("type")
        if step_type in {"academic_watch", "attendance_watch"}:
            default_col = ("Attendance Watch"
                           if step_type == "attendance_watch"
                           else "Academic Watch")
            column = edit.get("column_hint") or default_col
            value = edit.get("value", "Yes")
            edit_result = execute_academic_watch_action(
                filters=filters, sheets=sheets, sheet=sheet,
                request_summary=request_summary,
                output_dir=output_dir, audit_path=_AUDIT_SUPPRESS,
                column_name=column, value=value,
            )
        elif step_type == "note_edit":
            edit_result = execute_add_note_action(
                filters=filters, note=edit.get("note", ""),
                sheets=sheets, sheet=sheet,
                request_summary=request_summary,
                output_dir=output_dir, audit_path=_AUDIT_SUPPRESS,
            )
        elif step_type == "field_update":
            edit_result = execute_update_field_action(
                filters=filters, field_name=edit.get("field", ""),
                value=edit.get("value"),
                sheets=sheets, sheet=sheet,
                request_summary=request_summary,
                output_dir=output_dir, audit_path=_AUDIT_SUPPRESS,
            )
        else:
            return ActionResult(success=False, action_type="action_chain",
                                error=f"unsupported_step:{step_type}",
                                message=f"I can't run a {step_type} step in a chain.")
        if not edit_result.success:
            return ActionResult(
                success=False, action_type="action_chain",
                error=edit_result.error or "edit_failed",
                message=edit_result.message,
            )
        rows_affected = edit_result.rows_affected
        columns_changed = edit_result.columns
        output_file = edit_result.output_file
        # The export step simply points at the workbook the edit just wrote.

    # Build a short user-facing message.
    if edit is None:
        message = f"Done. Exported {rows_affected} row(s). New workbook: {output_file}"
    else:
        column_phrase = ", ".join(columns_changed) or "the target column"
        message = (
            f"Done. I marked {rows_affected} student(s) in {column_phrase} "
            f"and created a new Excel workbook. The original workbook was not modified."
        )
        if has_export:
            message += f" Export ready: {output_file}"

    record_audit(
        build_audit_record(
            action_type="action_chain",
            target_sheet=sheet,
            filters=filters,
            rows_affected=rows_affected,
            columns=columns_changed,
            output_file=output_file,
            sensitive_fields=_sensitive_fields(columns_changed),
            confirmation_status="confirmed",
            request_summary=request_summary,
            extra={"actions": [a.get("type") for a in actions],
                   "original_modified": False},
        ),
        audit_path,
    )

    return ActionResult(
        success=True,
        action_type="action_chain",
        rows_affected=rows_affected,
        output_file=output_file,
        columns=columns_changed,
        message=message,
    )


def execute_academic_watch_action(
    *,
    filters: list[dict[str, Any]],
    sheets: dict[str, pd.DataFrame],
    sheet: str,
    request_summary: str = "",
    output_dir: Path | None = None,
    audit_path: Path | None = None,
    column_name: str = "Academic Watch",
    value: str = "Yes",
) -> ActionResult:
    """Mark the currently filtered students as Academic Watch / Follow Up.

    School-roster workflow action: sets the Academic Watch column (creating it
    if missing) to "Yes" for rows that match ``filters``. Always writes a brand-
    new workbook; the original sheets dict is never mutated. Audits the action.
    """
    frame = sheets[sheet]
    # Resolve the target column case-insensitively and accept "Follow Up Needed"
    # as a synonym so e.g. "mark as follow up needed" lands on the right column.
    column = _resolve_column(frame, column_name) or column_name
    text_low = (request_summary or "").lower()
    if "follow up" in text_low or "follow-up" in text_low or "followup" in text_low:
        alt = _resolve_column(frame, "Follow Up Needed")
        if alt:
            column = alt

    modified = frame.copy()
    if column not in modified.columns:
        modified[column] = ""
    # All-blank columns parse as float64 (NaN); coerce to object so writing
    # 'Yes' doesn't raise from pandas' strict dtype check.
    if not pd.api.types.is_object_dtype(modified[column]):
        modified[column] = modified[column].astype(object)
    modified[column] = modified[column].where(modified[column].notna(), "")

    mask = get_target_rows_mask(modified, filters)
    rows = int(mask.sum())
    if rows == 0:
        return ActionResult(
            success=False,
            action_type="academic_watch",
            error="no_rows",
            message="No students match the current selection, so nothing was marked.",
        )
    modified.loc[mask, column] = value

    # Provenance: add a Watch Reason column to the exported workbook so a
    # counsellor reading the file later understands WHY each student was
    # flagged. Reason text comes from the filters that produced the mask.
    reason_column = _watch_reason_column_name(column)
    reason_text = describe_filters_as_reason(filters)
    date_column = "Date Flagged"
    if reason_column not in modified.columns:
        modified[reason_column] = ""
    if not pd.api.types.is_object_dtype(modified[reason_column]):
        modified[reason_column] = modified[reason_column].astype(object)
    modified[reason_column] = modified[reason_column].where(
        modified[reason_column].notna(), "",
    )
    if date_column not in modified.columns:
        modified[date_column] = ""
    if not pd.api.types.is_object_dtype(modified[date_column]):
        modified[date_column] = modified[date_column].astype(object)
    modified[date_column] = modified[date_column].where(
        modified[date_column].notna(), "",
    )
    if reason_text:
        modified.loc[mask, reason_column] = reason_text
    modified.loc[mask, date_column] = datetime.now().date().isoformat()

    output_sheets = dict(sheets)
    output_sheets[sheet] = modified
    path = write_output_workbook(output_sheets, "academic_watch", output_dir)
    result = ActionResult(
        success=True,
        action_type="academic_watch",
        rows_affected=rows,
        output_file=str(path),
        columns=[column, reason_column, date_column] if reason_text else [column, date_column],
        message=(f"Marked {rows} student(s) as {column} = '{value}'. "
                 f"A new workbook was saved: {path}"),
    )
    record_audit(
        build_audit_record(
            action_type="academic_watch",
            target_sheet=sheet,
            filters=filters,
            rows_affected=rows,
            columns=[column, reason_column, date_column] if reason_text else [column, date_column],
            output_file=str(path),
            sensitive_fields=_sensitive_fields([column]),
            confirmation_status="confirmed",
            request_summary=request_summary,
        ),
        audit_path,
    )
    return result


def _watch_reason_column_name(watch_column: str) -> str:
    """Pair a Watch column with a parallel Reason column. 'Academic Watch'
    → 'Academic Watch Reason'; 'Attendance Watch' → 'Attendance Watch
    Reason'; anything else → '<Column> Reason'."""
    cleaned = watch_column.strip()
    if cleaned.lower().endswith(" reason"):
        return cleaned
    return f"{cleaned} Reason"


def describe_filters_as_reason(filters: list[dict[str, Any]]) -> str:
    """Turn a list of filter conditions into a human-readable reason.

    Example: [GPA < 2.0, Attendance Rate < 90] → "GPA below 2.0; Attendance
    Rate below 90". The intent is provenance the counsellor can read later
    without re-running the query; we keep it short and side-step any
    sensitive-value detail (only the operator + threshold appear).
    """
    if not filters:
        return ""
    pieces: list[str] = []
    for condition in filters:
        column = condition.get("column", "")
        operator = condition.get("operator", "")
        value = condition.get("value")
        if not column or not operator:
            continue
        text = _filter_phrase(column, operator, value)
        if text:
            pieces.append(text)
    return "; ".join(pieces)


def _filter_phrase(column: str, operator: str, value: Any) -> str:
    op = operator
    if op in {"equals", "in"}:
        if isinstance(value, list):
            return f"{column} in [{', '.join(map(str, value))}]"
        if isinstance(value, bool):
            return column if value else f"{column}=No"
        return f"{column} = {value}"
    if op in {"not_equals", "not_in"}:
        return f"{column} ≠ {value}"
    if op == "greater_than":
        return f"{column} above {value}"
    if op == "greater_or_equal":
        return f"{column} ≥ {value}"
    if op == "less_than":
        return f"{column} below {value}"
    if op == "less_or_equal":
        return f"{column} ≤ {value}"
    if op == "between" and isinstance(value, list) and len(value) == 2:
        return f"{column} between {value[0]} and {value[1]}"
    if op in {"is_missing", "is_blank"}:
        return f"{column} missing"
    if op in {"is_not_missing", "is_not_blank"}:
        return f"{column} present"
    if op in {"contains", "contains_text"}:
        return f"{column} contains '{value}'"
    return f"{column} {op.replace('_', ' ')} {value}"


def execute_update_field_action(
    *,
    filters: list[dict[str, Any]],
    field_name: str,
    value: Any,
    sheets: dict[str, pd.DataFrame],
    sheet: str,
    request_summary: str = "",
    output_dir: Path | None = None,
    audit_path: Path | None = None,
) -> ActionResult:
    status = field_status(field_name)
    if status == "protected":
        return ActionResult(
            success=False,
            action_type="field_update",
            error="protected_field",
            message=f"I can't update {field_name} because it is a protected field. You can export the list for manual review instead.",
        )
    if status != "safe":
        return ActionResult(
            success=False,
            action_type="field_update",
            error="not_editable",
            message=f"'{field_name}' is not an editable field for this assistant.",
        )

    frame = sheets[sheet]
    column = _resolve_column(frame, field_name) or field_name
    modified = frame.copy()
    if column not in modified.columns:
        modified[column] = ""
    mask = get_target_rows_mask(modified, filters)
    modified.loc[mask, column] = value

    output_sheets = dict(sheets)
    output_sheets[sheet] = modified
    path = write_output_workbook(output_sheets, "student_records_modified", output_dir)
    rows = int(mask.sum())
    result = ActionResult(
        success=True,
        action_type="field_update",
        rows_affected=rows,
        output_file=str(path),
        columns=[column],
        message=f"Set {column} to '{value}' for {rows} matching record(s). A new workbook was saved: {path}",
    )
    record_audit(
        build_audit_record(
            action_type="update_field",
            target_sheet=sheet,
            filters=filters,
            rows_affected=rows,
            columns=[column],
            output_file=str(path),
            sensitive_fields=_sensitive_fields([column]),
            confirmation_status="confirmed",
            request_summary=request_summary,
        ),
        audit_path,
    )
    return result
