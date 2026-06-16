from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.chart_engine import create_chart
from core.data_quality_engine import create_data_quality_report
from core.formula_engine import create_formula
from core.report_engine import create_report
from core.validator import validate_command, validate_plan


@dataclass(frozen=True)
class ActionResult:
    message: str
    preview: pd.DataFrame | None = None
    result_sheet: str | None = None


def execute_command(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
    original_file_name: str,
) -> ActionResult:
    validate_command(command, sheets, original_file_name)

    action = command["action"]
    if action == "create_data_quality_report":
        result = create_data_quality_report(command, workbook, sheets)
        return ActionResult(
            message=result.message,
            preview=result.preview,
            result_sheet=result.result_sheet,
        )
    if action == "create_chart":
        result = create_chart(command, workbook, sheets)
        return ActionResult(
            message=result.message,
            preview=result.preview,
            result_sheet=result.result_sheet,
        )
    if action == "create_formula":
        result = create_formula(command, workbook, sheets)
        return ActionResult(
            message=result.message,
            preview=result.preview,
            result_sheet=result.result_sheet,
        )
    if action == "create_report":
        result = create_report(command, workbook, sheets)
        return ActionResult(
            message=result.message,
            preview=result.preview,
            result_sheet=result.result_sheet,
        )
    if action == "filter_rows":
        return filter_rows(command, workbook, sheets)
    if action == "highlight_rows":
        return highlight_rows(command, workbook, sheets)
    if action == "sum_column":
        return sum_column(command, workbook, sheets)
    if action == "count_rows":
        return count_rows(command, workbook, sheets)
    if action == "count_by_group":
        return count_by_group(command, workbook, sheets)
    if action == "detect_missing_values":
        return detect_missing_values(command, workbook, sheets)
    if action == "remove_duplicates":
        return remove_duplicates(command, workbook, sheets)
    if action == "format_report":
        return format_report(command, workbook, sheets)

    raise ValueError(f"Unsupported action: {action}")


def filter_rows(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    dataframe = sheets[sheet_name]
    filtered = dataframe[_condition_mask(dataframe, command["conditions"])].copy()
    result_sheet = _unique_sheet_name(workbook, f"Filtered_{sheet_name}")
    _write_dataframe_to_new_sheet(workbook, result_sheet, filtered)

    return ActionResult(
        message=f"Filtered {len(filtered)} rows from {sheet_name}.",
        preview=filtered,
        result_sheet=result_sheet,
    )


def highlight_rows(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    worksheet = workbook[sheet_name]
    dataframe = sheets[sheet_name]
    mask = _condition_mask(dataframe, command["conditions"])

    fill_color = command.get("format", {}).get("fill_color", "FFFF00")
    fill = PatternFill(
        fill_type="solid",
        fgColor=_normalize_fill_color(fill_color),
    )

    highlighted_count = 0
    for dataframe_index in dataframe.index[mask]:
        excel_row = int(dataframe_index) + 2
        for cell in worksheet[excel_row]:
            cell.fill = fill
        highlighted_count += 1

    return ActionResult(
        message=f"Highlighted {highlighted_count} rows in {sheet_name}.",
        preview=dataframe[mask].copy(),
    )


def sum_column(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    column = command["column"]
    group_by = command.get("group_by")

    if group_by:
        result = (
            sheets[sheet_name]
            .groupby(group_by, dropna=False)[column]
            .sum()
            .reset_index(name=f"{column} sum")
        )
        message = f"Summed {column} by {group_by} in {sheet_name}."
    else:
        total = sheets[sheet_name][column].sum()
        result = pd.DataFrame(
            [{"sheet": sheet_name, "column": column, "sum": total}]
        )
        message = f"Summed {column} in {sheet_name}: {total}"

    result_sheet = _write_report_sheet(workbook, "Sum_Report", result)

    return ActionResult(
        message=message,
        preview=result,
        result_sheet=result_sheet,
    )


def count_rows(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    dataframe = sheets[sheet_name]

    if command.get("conditions"):
        count = int(_condition_mask(dataframe, command["conditions"]).sum())
    else:
        count = len(dataframe.index)

    result = pd.DataFrame([{"sheet": sheet_name, "row_count": count}])
    result_sheet = _write_report_sheet(workbook, "Count_Report", result)

    return ActionResult(
        message=f"Counted {count} rows in {sheet_name}.",
        preview=result,
        result_sheet=result_sheet,
    )


def count_by_group(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    group_by = command.get("group_by") or command["column"]
    dataframe = sheets[sheet_name]
    if command.get("conditions"):
        dataframe = dataframe[_condition_mask(dataframe, command["conditions"])].copy()
    result = (
        dataframe
        .groupby(group_by, dropna=False)
        .size()
        .reset_index(name="count")
    )
    result_sheet = _write_report_sheet(workbook, "Group_Count_Report", result)

    return ActionResult(
        message=f"Counted rows by {group_by} in {sheet_name}.",
        preview=result,
        result_sheet=result_sheet,
    )


def detect_missing_values(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    dataframe = sheets[sheet_name]
    column = command.get("column")

    if column:
        result = pd.DataFrame(
            [{"sheet": sheet_name, "column": column, "missing_values": int(dataframe[column].isna().sum())}]
        )
    else:
        result = (
            dataframe.isna()
            .sum()
            .rename_axis("column")
            .reset_index(name="missing_values")
        )
        result.insert(0, "sheet", sheet_name)

    result_sheet = _write_report_sheet(workbook, "Missing_Values_Report", result)

    return ActionResult(
        message=f"Detected missing values in {sheet_name}.",
        preview=result,
        result_sheet=result_sheet,
    )


def remove_duplicates(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    dataframe = sheets[sheet_name]
    columns = command.get("columns") or None
    deduplicated = dataframe.drop_duplicates(subset=columns).copy()
    removed_count = len(dataframe.index) - len(deduplicated.index)

    worksheet = workbook[sheet_name]
    _replace_worksheet_values(worksheet, deduplicated)

    return ActionResult(
        message=f"Removed {removed_count} duplicate rows from {sheet_name}.",
        preview=deduplicated,
    )


def format_report(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    sheet_name = command["sheet"]
    worksheet = workbook[sheet_name]

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)

    return ActionResult(
        message=f"Formatted report sheet {sheet_name}.",
        preview=sheets[sheet_name].head(100).copy(),
    )


# Multi-step plan execution ---------------------------------------------------


@dataclass(frozen=True)
class PlanExecutionResult:
    message: str
    step_messages: list[str]
    result_sheet: str | None = None
    preview: pd.DataFrame | None = None


def execute_plan(
    plan: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
    original_file_name: str,
) -> PlanExecutionResult:
    """Validate and run a (single- or multi-step) planner envelope.

    Steps run in order. Each step that creates an output_sheet registers its
    result so later steps can reference it. Because the requested output_sheet
    name may collide with an existing sheet, we map requested names to the
    actual worksheet titles and rewrite later references before each step.
    """
    validate_plan(plan, sheets, original_file_name)

    commands = [c for c in plan.get("commands", []) if c.get("action") != "column_mapping_request"]
    working: dict[str, pd.DataFrame] = dict(sheets)
    name_map: dict[str, str] = {}  # requested output name -> actual worksheet title

    messages: list[str] = []
    last_sheet: str | None = None
    last_preview: pd.DataFrame | None = None

    for command in commands:
        cmd = dict(command)
        if cmd.get("sheet") in name_map:
            cmd["sheet"] = name_map[cmd["sheet"]]
        if isinstance(cmd.get("sheets"), list):
            cmd["sheets"] = [name_map.get(s, s) for s in cmd["sheets"]]

        requested_out = command.get("output_sheet")
        result = _execute_planner_step(cmd, workbook, working)
        messages.append(result.message)
        if result.result_sheet:
            last_sheet = result.result_sheet
            if requested_out:
                name_map[requested_out] = result.result_sheet
        if result.preview is not None:
            last_preview = result.preview

    # Surface any new intermediate sheets to the caller's dict.
    for name, frame in working.items():
        sheets.setdefault(name, frame)

    return PlanExecutionResult(
        message="; ".join(messages) if messages else "Plan executed.",
        step_messages=messages,
        result_sheet=last_sheet,
        preview=last_preview,
    )


def _execute_planner_step(
    command: dict[str, Any],
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
) -> ActionResult:
    action = command["action"]
    if action == "filter_rows":
        return _planner_filter_rows(command, workbook, sheets)
    if action == "highlight_rows":
        return highlight_rows(command, workbook, sheets)
    if action == "sort_rows":
        return sort_rows(command, workbook, sheets)
    if action == "sum_column":
        return _planner_sum_column(command, workbook, sheets)
    if action == "average_column":
        return average_column(command, workbook, sheets)
    if action == "sum_by_group":
        return sum_by_group(command, workbook, sheets)
    if action == "count_rows":
        return count_rows(command, workbook, sheets)
    if action == "count_by_group":
        return _planner_count_by_group(command, workbook, sheets)
    if action == "detect_missing_values":
        return detect_missing_values(command, workbook, sheets)
    if action == "remove_duplicates":
        result = remove_duplicates(command, workbook, sheets)
        if result.preview is not None:
            sheets[command["sheet"]] = result.preview.reset_index(drop=True)
        return result
    if action == "move_rows_to_sheet":
        return move_rows_to_sheet(command, workbook, sheets)
    if action == "create_summary_sheet":
        return create_summary_sheet(command, workbook, sheets)
    if action == "create_chart":
        return _planner_create_chart(command, workbook, sheets)
    if action == "create_report":
        return create_report(command, workbook, sheets)
    if action == "create_formula":
        return create_formula(command, workbook, sheets)
    if action == "create_data_quality_report":
        return create_data_quality_report(command, workbook, sheets)
    if action == "format_report":
        return _planner_format_report(command, workbook, sheets)
    if action == "freeze_header":
        return freeze_header(command, workbook, sheets)
    if action == "autofit_columns":
        return autofit_columns(command, workbook, sheets)
    if action == "apply_conditional_formatting":
        return apply_conditional_formatting(command, workbook, sheets)
    raise ValueError(f"Unsupported plan action: {action}")


def _write_plan_sheet(
    workbook: Workbook,
    sheets: dict[str, pd.DataFrame],
    base_name: str,
    dataframe: pd.DataFrame,
) -> str:
    actual = _unique_sheet_name(workbook, base_name)
    _write_dataframe_to_new_sheet(workbook, actual, dataframe)
    sheets[actual] = dataframe.reset_index(drop=True)
    return actual


def _planner_filter_rows(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    dataframe = sheets[sheet]
    filtered = dataframe[_condition_mask(dataframe, command["conditions"])].copy().reset_index(drop=True)
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or f"Filtered_{sheet}", filtered)
    return ActionResult(message=f"Filtered {len(filtered)} rows from {sheet}.", preview=filtered, result_sheet=actual)


def _planner_count_by_group(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    group_by = command.get("group_by") or command["column"]
    dataframe = sheets[sheet]
    if command.get("conditions"):
        dataframe = dataframe[_condition_mask(dataframe, command["conditions"])]
    out = dataframe.groupby(group_by, dropna=False).size().reset_index(name="count")
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or f"{group_by} Counts", out)
    return ActionResult(message=f"Counted rows by {group_by} in {sheet}.", preview=out, result_sheet=actual)


def _planner_sum_column(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    column = command["column"]
    group_by = command.get("group_by")
    dataframe = sheets[sheet]
    if group_by:
        # Keep the value column's name so downstream steps (and the validator's
        # projected schema) can reference it as `column`.
        out = dataframe.groupby(group_by, dropna=False)[column].sum().reset_index()
        message = f"Summed {column} by {group_by} in {sheet}."
    else:
        out = pd.DataFrame([{"sheet": sheet, "column": column, "sum": dataframe[column].sum()}])
        message = f"Summed {column} in {sheet}."
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or "Sum_Report", out)
    return ActionResult(message=message, preview=out, result_sheet=actual)


def sort_rows(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    dataframe = sheets[sheet]
    sort_by = command["sort_by"]
    columns = [item["column"] for item in sort_by]
    ascending = [str(item.get("direction", "asc")).lower() != "desc" for item in sort_by]
    ordered = dataframe.sort_values(by=columns, ascending=ascending, kind="stable").reset_index(drop=True)
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or f"Sorted_{sheet}", ordered)
    return ActionResult(message=f"Sorted {sheet} by {', '.join(columns)}.", preview=ordered, result_sheet=actual)


def average_column(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    column = command["column"]
    group_by = command.get("group_by")
    dataframe = sheets[sheet]
    numeric = pd.to_numeric(dataframe[column], errors="coerce")
    if group_by:
        # Keep the value column's name to match the validator's projected schema.
        out = (
            dataframe.assign(**{column: numeric})
            .groupby(group_by, dropna=False)[column]
            .mean()
            .round(4)
            .reset_index()
        )
        message = f"Averaged {column} by {group_by} in {sheet}."
    else:
        average = round(float(numeric.mean()), 4)
        out = pd.DataFrame([{"sheet": sheet, "column": column, "average": average}])
        message = f"Average {column} in {sheet}: {average}."
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or "Average_Report", out)
    return ActionResult(message=message, preview=out, result_sheet=actual)


def sum_by_group(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    group_by = command["group_by"]
    sum_column = command.get("sum_column") or command.get("column")
    dataframe = sheets[sheet]
    out = dataframe.groupby(group_by, dropna=False)[sum_column].sum().round(4).reset_index()
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or f"{group_by} Totals", out)
    return ActionResult(message=f"Summed {sum_column} by {group_by} in {sheet}.", preview=out, result_sheet=actual)


def move_rows_to_sheet(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    dataframe = sheets[sheet]
    subset = dataframe[_condition_mask(dataframe, command["conditions"])].copy().reset_index(drop=True)
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or f"Moved_{sheet}", subset)
    return ActionResult(
        message=f"Copied {len(subset)} rows to {actual} (the original rows are kept).",
        preview=subset,
        result_sheet=actual,
    )


def create_summary_sheet(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    dataframe = sheets[sheet]
    rows = []
    for metric in command["metrics"]:
        kind = metric["kind"]
        label = metric.get("label") or kind
        if kind == "count":
            value = int(len(dataframe.index))
        elif kind == "count_missing":
            value = int(_blank_count(dataframe[metric["column"]]))
        elif kind == "sum":
            value = round(float(pd.to_numeric(dataframe[metric["column"]], errors="coerce").sum()), 4)
        elif kind == "average":
            value = round(float(pd.to_numeric(dataframe[metric["column"]], errors="coerce").mean()), 4)
        else:
            continue
        rows.append({"Metric": label, "Value": value})
    out = pd.DataFrame(rows)
    actual = _write_plan_sheet(workbook, sheets, command.get("output_sheet") or "Summary", out)
    return ActionResult(message=f"Created summary sheet {actual}.", preview=out, result_sheet=actual)


def freeze_header(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    workbook[sheet].freeze_panes = "A2"
    return ActionResult(message=f"Froze the header row on {sheet}.", result_sheet=sheet)


def autofit_columns(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    _autosize_columns(workbook[sheet])
    return ActionResult(message=f"Auto-fit column widths on {sheet}.", result_sheet=sheet)


def apply_conditional_formatting(command, workbook, sheets) -> ActionResult:
    sheet = command["sheet"]
    worksheet = workbook[sheet]
    dataframe = sheets[sheet]
    mask = _condition_mask(dataframe, command["conditions"])
    fill = PatternFill(
        fill_type="solid",
        fgColor=_normalize_fill_color(command.get("format", {}).get("fill_color", "green")),
    )
    target_column = command.get("column")
    column_index = list(dataframe.columns).index(target_column) + 1 if target_column in dataframe.columns else None

    count = 0
    for dataframe_index in dataframe.index[mask]:
        excel_row = int(dataframe_index) + 2
        if column_index:
            worksheet.cell(row=excel_row, column=column_index).fill = fill
        else:
            for cell in worksheet[excel_row]:
                cell.fill = fill
        count += 1

    where = f" on {target_column}" if target_column else ""
    return ActionResult(message=f"Applied conditional formatting to {count} rows{where} in {sheet}.", result_sheet=sheet)


def _planner_create_chart(command, workbook, sheets) -> ActionResult:
    cmd = dict(command)
    if not cmd.get("group_by"):
        cmd["group_by"] = cmd.get("category_column")
    if not cmd.get("metric"):
        cmd["metric"] = "sum" if cmd.get("value_column") else "count_rows"
    return create_chart(cmd, workbook, sheets)


def _planner_format_report(command, workbook, sheets) -> ActionResult:
    targets = command.get("sheets")
    if not (isinstance(targets, list) and targets):
        targets = [command["sheet"]]
    for sheet in targets:
        _format_one(workbook[sheet])
    return ActionResult(message=f"Formatted {', '.join(targets)}.", result_sheet=targets[-1])


def _format_one(worksheet: Worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)


def _blank_count(series: pd.Series) -> int:
    normalized = series.astype(str).str.strip().str.casefold()
    return int((series.isna() | normalized.isin({"", "nan", "none", "null", "n/a", "na"})).sum())


def _condition_mask(dataframe: pd.DataFrame, conditions: list[dict[str, Any]]) -> pd.Series:
    mask = pd.Series(True, index=dataframe.index)
    for condition in conditions:
        column = condition["column"]
        operator = condition["operator"]
        value = condition.get("value")
        series = dataframe[column]
        comparable_value = _coerce_comparison_value(series, value)

        if operator == "equals":
            if pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series):
                condition_mask = series.astype(str).str.casefold() == str(comparable_value).casefold()
            else:
                condition_mask = series == comparable_value
        elif operator == "not_equals":
            if pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series):
                condition_mask = series.astype(str).str.casefold() != str(comparable_value).casefold()
            else:
                condition_mask = series != comparable_value
        elif operator == "greater_than":
            condition_mask = pd.to_numeric(series, errors="coerce") > float(value)
        elif operator == "greater_or_equal":
            condition_mask = pd.to_numeric(series, errors="coerce") >= float(value)
        elif operator == "less_than":
            condition_mask = pd.to_numeric(series, errors="coerce") < float(value)
        elif operator == "less_or_equal":
            condition_mask = pd.to_numeric(series, errors="coerce") <= float(value)
        elif operator == "contains":
            condition_mask = series.astype(str).str.contains(str(value), case=False, na=False)
        elif operator == "contains_any":
            condition_mask = series.astype(str).apply(
                lambda item: any(str(term).casefold() in item.casefold() for term in value)
            )
        elif operator == "not_contains":
            condition_mask = ~series.astype(str).str.contains(str(value), case=False, na=False)
        elif operator == "starts_with":
            condition_mask = series.astype(str).str.casefold().str.startswith(str(value).casefold()) & series.notna()
        elif operator == "ends_with":
            condition_mask = series.astype(str).str.casefold().str.endswith(str(value).casefold()) & series.notna()
        elif operator == "between":
            if not isinstance(value, (list, tuple)) or len(value) != 2:
                raise ValueError("between requires a [low, high] value.")
            low, high = sorted((float(value[0]), float(value[1])))
            numeric = pd.to_numeric(series, errors="coerce")
            condition_mask = (numeric >= low) & (numeric <= high)
        elif operator == "not_in":
            values = value if isinstance(value, list) else [value]
            tokens = {str(v).strip().casefold() for v in values}
            condition_mask = ~series.astype(str).str.strip().str.casefold().isin(tokens)
        elif operator in ("is_missing", "is_blank"):
            normalized = series.astype(str).str.strip().str.casefold()
            condition_mask = series.isna() | normalized.isin({"", "nan", "none", "null", "n/a", "na"})
        elif operator in ("is_not_missing", "is_not_blank"):
            normalized = series.astype(str).str.strip().str.casefold()
            condition_mask = ~(series.isna() | normalized.isin({"", "nan", "none", "null", "n/a", "na"}))
        else:
            raise ValueError(f"Unsupported operator: {operator}")

        mask &= condition_mask.fillna(False)

    return mask


def _write_dataframe_to_new_sheet(
    workbook: Workbook,
    sheet_name: str,
    dataframe: pd.DataFrame,
) -> Worksheet:
    worksheet = workbook.create_sheet(sheet_name)
    _write_dataframe_values(worksheet, dataframe)
    _format_header(worksheet)
    _autosize_columns(worksheet)
    return worksheet


def _write_report_sheet(
    workbook: Workbook,
    base_sheet_name: str,
    dataframe: pd.DataFrame,
) -> str:
    sheet_name = _unique_sheet_name(workbook, base_sheet_name)
    _write_dataframe_to_new_sheet(workbook, sheet_name, dataframe)
    return sheet_name


def _write_dataframe_values(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    worksheet.append([str(column) for column in dataframe.columns])
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append([_excel_safe_value(value) for value in row])


def _replace_worksheet_values(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
    header_styles = [_copy_cell_style(cell) for cell in worksheet[1]]
    worksheet.delete_rows(1, worksheet.max_row)
    _write_dataframe_values(worksheet, dataframe)

    for cell, style in zip(worksheet[1], header_styles):
        _apply_cell_style(cell, style)

    _autosize_columns(worksheet)


def _format_header(worksheet: Worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def _copy_cell_style(cell) -> dict[str, Any]:
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def _apply_cell_style(cell, style: dict[str, Any]) -> None:
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.border = style["border"]
    cell.alignment = style["alignment"]
    cell.number_format = style["number_format"]
    cell.protection = style["protection"]


def _excel_safe_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _coerce_comparison_value(series: pd.Series, value: Any) -> Any:
    if pd.api.types.is_numeric_dtype(series):
        return float(value)
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(value)
    return value


def _normalize_fill_color(fill_color: str) -> str:
    color_map = {
        "yellow": "FFFF00",
        "green": "C6EFCE",
        "red": "FFC7CE",
        "blue": "BDD7EE",
        "orange": "F4B183",
    }
    normalized = color_map.get(str(fill_color).lower(), str(fill_color).replace("#", ""))
    if len(normalized) == 6:
        return normalized.upper()
    if len(normalized) == 8:
        return normalized.upper()
    return "FFFF00"


def _unique_sheet_name(workbook: Workbook, base_name: str) -> str:
    safe_base_name = base_name[:31]
    if safe_base_name not in workbook.sheetnames:
        return safe_base_name

    counter = 1
    while True:
        suffix = f"_{counter}"
        candidate = f"{safe_base_name[:31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        counter += 1
