import csv
import os
import re
import sqlite3
import subprocess
import sys
import tkinter as tk
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    DND_FILES = None
    TkinterDnD = None


APP_TITLE = "Offline Dean Student Assistant"
DB_PATH = Path(__file__).with_name("students.db")
TkBase = TkinterDnD.Tk if TkinterDnD else tk.Tk

COLLEGES = [
    "Arts and Sciences",
    "Business",
    "Education",
    "Engineering",
    "Health Sciences",
    "Nursing",
]

STATUSES = ["Enrolled", "Inactive", "Withdrawn", "Graduated"]

SAMPLE_STUDENTS = [
    ("1001", "Alicia Rivera", "Engineering", "Computer Science", "Enrolled", "Main Campus", "Dr. Kim"),
    ("1002", "Marcus Chen", "Business", "Accounting", "Enrolled", "Downtown Campus", "Prof. Lane"),
    ("1003", "Jasmine Patel", "Nursing", "Nursing", "Inactive", "Health Campus", "Dr. Brooks"),
    ("1004", "Elena Gomez", "Arts and Sciences", "Biology", "Enrolled", "Main Campus", "Dr. Stone"),
    ("1005", "Noah Williams", "Health Sciences", "Public Health", "Withdrawn", "Health Campus", "Dr. Ford"),
    ("1006", "Samira Hassan", "Education", "Elementary Education", "Graduated", "Main Campus", "Prof. Carter"),
]

UNIVERSITY_TERMS = {
    "student_id": ["student id", "student_id", "id", "banner id", "university id"],
    "name": ["name", "student name", "full name"],
    "program": ["program", "major", "department", "degree program", "discipline"],
    "balance": ["balance", "balance due", "amount due", "tuition balance", "student balance"],
    "fafsa": ["fafsa", "financial aid", "aid status", "missing aid"],
    "enrollment": ["enrollment", "enrollment status", "status", "active", "enrolled"],
    "registration": ["registered", "registration", "registered credits", "credits"],
    "standing": ["standing", "academic standing", "good standing", "bad standing"],
    "year": ["year", "class year", "student year"],
}


@dataclass
class SpreadsheetCommand:
    action: str
    description: str
    conditions: list
    target_column: str = ""
    group_by: str = ""
    new_sheet: str = ""


class WorkbookData:
    def __init__(self, path, headers, rows):
        self.path = Path(path)
        self.headers = headers
        self.rows = rows

    @property
    def row_count(self):
        return len(self.rows)

    @property
    def column_count(self):
        return len(self.headers)

    def summary(self):
        header_text = ", ".join(self.headers) if self.headers else "No headers found"
        sample_lines = []
        for row in self.rows[:5]:
            preview = [
                f"{header}: {value}"
                for header, value in zip(self.headers, row)
                if str(value).strip()
            ][:4]
            if preview:
                sample_lines.append("- " + " | ".join(preview))

        lines = [
            f"Loaded workbook: {self.path.name}",
            f"Headers: {header_text}",
            f"Rows: {self.row_count}",
            f"Columns: {self.column_count}",
        ]
        if sample_lines:
            lines.append("Sample rows:")
            lines.extend(sample_lines)
        return "\n".join(lines)

    def matching_rows(self, terms):
        clean_terms = [term.lower() for term in terms if term.strip()]
        if not clean_terms:
            return []

        matches = []
        for index, row in enumerate(self.rows):
            row_text = " ".join(str(value).lower() for value in row)
            if all(term in row_text for term in clean_terms):
                matches.append((index, row))
        return matches

    def matching_columns(self, requested_columns):
        if not requested_columns:
            return list(range(len(self.headers)))

        matched = []
        for requested in requested_columns:
            requested = requested.lower().strip()
            for index, header in enumerate(self.headers):
                header_text = header.lower()
                if requested == header_text or requested in header_text or header_text in requested:
                    if index not in matched:
                        matched.append(index)
        return matched or list(range(len(self.headers)))


class XlsxReader:
    XML_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    @classmethod
    def read_first_sheet(cls, path):
        path = Path(path)
        if path.suffix.lower() != ".xlsx":
            raise ValueError("Please choose an .xlsx Excel file.")

        with zipfile.ZipFile(path) as workbook:
            shared_strings = cls.read_shared_strings(workbook)
            sheet_name = cls.first_sheet_name(workbook)
            raw_rows = cls.read_sheet_rows(workbook, sheet_name, shared_strings)

        rows = [row for row in raw_rows if any(str(value).strip() for value in row)]
        if not rows:
            raise ValueError("The workbook sheet is empty.")

        headers = [str(value).strip() or f"Column {index + 1}" for index, value in enumerate(rows[0])]
        body_rows = rows[1:]
        width = len(headers)
        padded_rows = [(row + [""] * width)[:width] for row in body_rows]
        return WorkbookData(path, headers, padded_rows)

    @classmethod
    def read_shared_strings(cls, workbook):
        if "xl/sharedStrings.xml" not in workbook.namelist():
            return []

        root = ET.fromstring(workbook.read("xl/sharedStrings.xml"))
        values = []
        for item in root.findall("a:si", cls.XML_NS):
            text_parts = [node.text or "" for node in item.findall(".//a:t", cls.XML_NS)]
            values.append("".join(text_parts))
        return values

    @classmethod
    def first_sheet_name(cls, workbook):
        preferred = "xl/worksheets/sheet1.xml"
        if preferred in workbook.namelist():
            return preferred

        for name in workbook.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                return name
        raise ValueError("No worksheet XML file was found in this workbook.")

    @classmethod
    def read_sheet_rows(cls, workbook, sheet_name, shared_strings):
        root = ET.fromstring(workbook.read(sheet_name))
        rows = []
        for row_node in root.findall(".//a:sheetData/a:row", cls.XML_NS):
            cells = {}
            max_col = 0
            for cell in row_node.findall("a:c", cls.XML_NS):
                reference = cell.attrib.get("r", "")
                column_index = cls.column_index(reference)
                max_col = max(max_col, column_index)
                cells[column_index] = cls.cell_value(cell, shared_strings)
            rows.append([cells.get(index, "") for index in range(max_col + 1)])
        return rows

    @classmethod
    def column_index(cls, reference):
        letters = "".join(char for char in reference if char.isalpha())
        total = 0
        for char in letters:
            total = total * 26 + (ord(char.upper()) - ord("A") + 1)
        return max(total - 1, 0)

    @classmethod
    def cell_value(cls, cell, shared_strings):
        cell_type = cell.attrib.get("t")
        value_node = cell.find("a:v", cls.XML_NS)
        inline_node = cell.find("a:is/a:t", cls.XML_NS)

        if inline_node is not None:
            return inline_node.text or ""
        if value_node is None:
            return ""

        raw_value = value_node.text or ""
        if cell_type == "s":
            try:
                return shared_strings[int(raw_value)]
            except (ValueError, IndexError):
                return raw_value
        return raw_value


class SpreadsheetCommandParser:
    def parse(self, message, headers):
        lowered = message.lower().strip()
        if not headers:
            return None

        if "format" in lowered and ("report" in lowered or "sheet" in lowered):
            return SpreadsheetCommand(
                action="format_report",
                description="Format the report header, freeze the top row, and fit column widths.",
                conditions=[],
            )

        if "move" in lowered and ("sheet" in lowered or "tab" in lowered):
            conditions = self.conditions_from_text(lowered, headers)
            if conditions:
                return SpreadsheetCommand(
                    action="move_rows_to_sheet",
                    description="Move matching rows to a new sheet.",
                    conditions=conditions,
                    new_sheet=self.safe_sheet_name(lowered),
                )

        if "sum" in lowered or "total" in lowered:
            target_text = self.text_between_sum_and_group(lowered) or "balance"
            target = self.find_column(headers, target_text)
            group_by = self.find_column(headers, self.words_after(lowered, ["by", "per"]) or "program")
            if target:
                return SpreadsheetCommand(
                    action="sum_column",
                    description=f"Sum {target}" + (f" by {group_by}." if group_by else "."),
                    conditions=[],
                    target_column=target,
                    group_by=group_by,
                    new_sheet="Summary",
                )

        if "count" in lowered or "how many" in lowered:
            group_by = self.find_column(headers, self.words_after(lowered, ["by", "per"]) or "program")
            return SpreadsheetCommand(
                action="count_by_group",
                description=f"Count rows" + (f" by {group_by}." if group_by else "."),
                conditions=self.conditions_from_text(lowered, headers),
                group_by=group_by,
                new_sheet="Counts",
            )

        if "highlight" in lowered or "missing" in lowered or "blank" in lowered:
            conditions = self.conditions_from_text(lowered, headers)
            if conditions:
                return SpreadsheetCommand(
                    action="highlight_rows",
                    description="Highlight rows that match the request.",
                    conditions=conditions,
                )

        if "show" in lowered or "filter" in lowered or "find" in lowered or "who" in lowered:
            conditions = self.conditions_from_text(lowered, headers)
            if conditions:
                return SpreadsheetCommand(
                    action="filter_rows_to_sheet",
                    description="Create a filtered results sheet with matching rows.",
                    conditions=conditions,
                    new_sheet="Filtered Results",
                )

        return None

    def conditions_from_text(self, lowered, headers):
        conditions = []

        if any(word in lowered for word in ["missing", "blank", "not submitted"]):
            column = self.find_column(headers, lowered)
            if not column and "fafsa" in lowered:
                column = self.find_column(headers, "fafsa")
            if column:
                conditions.append({"column": column, "operator": "is_blank_or_missing", "value": ""})

        if "fafsa" in lowered and not any(condition["column"] == self.find_column(headers, "fafsa") for condition in conditions):
            column = self.find_column(headers, "fafsa")
            if column:
                if "missing" in lowered:
                    conditions.append({"column": column, "operator": "contains", "value": "missing"})
                else:
                    conditions.append({"column": column, "operator": "is_not_blank", "value": ""})

        if "balance" in lowered or "owe" in lowered or "tuition" in lowered or "paid" in lowered:
            column = self.find_column(headers, "balance")
            if column:
                conditions.append({"column": column, "operator": "greater_than", "value": 0})

        if "inactive" in lowered:
            column = self.find_column(headers, "status enrollment")
            if column:
                conditions.append({"column": column, "operator": "contains", "value": "inactive"})

        if "active" in lowered or "enrolled" in lowered:
            column = self.find_column(headers, "status enrollment")
            if column:
                conditions.append({"column": column, "operator": "contains_any", "value": ["active", "enrolled"]})

        if "not registered" in lowered or "didn't register" in lowered or "not register" in lowered:
            column = self.find_column(headers, "registered registration credits")
            if column:
                conditions.append({"column": column, "operator": "is_blank_or_zero_or_no", "value": ""})

        if "bad standing" in lowered:
            column = self.find_column(headers, "standing")
            if column:
                conditions.append({"column": column, "operator": "contains", "value": "bad standing"})

        if "good standing" in lowered:
            column = self.find_column(headers, "standing")
            if column:
                conditions.append({"column": column, "operator": "contains", "value": "good standing"})

        for header in headers:
            header_lower = header.lower()
            if header_lower in lowered and not any(condition["column"] == header for condition in conditions):
                value = self.value_after_header(lowered, header_lower)
                if value:
                    conditions.append({"column": header, "operator": "contains", "value": value})

        return conditions

    def find_column(self, headers, text):
        text = text.lower()
        normalized_headers = {header: header.lower().replace("_", " ") for header in headers}

        for header, normalized in normalized_headers.items():
            if normalized in text or text in normalized:
                return header

        for concept, aliases in UNIVERSITY_TERMS.items():
            if concept in text or any(alias in text for alias in aliases):
                for header, normalized in normalized_headers.items():
                    if any(alias in normalized or normalized in alias for alias in aliases):
                        return header
        return ""

    def words_after(self, text, markers):
        for marker in markers:
            match = re.search(rf"\b{marker}\b\s+(.+)", text)
            if match:
                return match.group(1)
        return ""

    def text_between_sum_and_group(self, text):
        match = re.search(r"\b(?:sum|total)\b\s+(.+?)(?:\s+\b(?:by|per)\b\s+|$)", text)
        return match.group(1).strip() if match else ""

    def value_after_header(self, text, header):
        match = re.search(rf"{re.escape(header)}\s+(?:is|=|equals|with|as)?\s*([a-z0-9 ._-]+)", text)
        if not match:
            return ""
        value = match.group(1)
        value = re.split(r"\b(and|or|by|to|in workbook|in sheet)\b", value)[0]
        return value.strip()

    def safe_sheet_name(self, lowered):
        if "inactive" in lowered:
            return "Inactive Students"
        if "balance" in lowered or "owe" in lowered:
            return "Outstanding Balance"
        if "fafsa" in lowered:
            return "Missing FAFSA"
        return "Extracted Rows"


class SpreadsheetActionEngine:
    def __init__(self, path):
        if load_workbook is None:
            raise RuntimeError("openpyxl is required for editing Excel files. Install it with: python3 -m pip install -r requirements.txt")
        self.path = Path(path)
        self.workbook = load_workbook(self.path)
        self.sheet = self.workbook.active
        self.headers = [str(cell.value or "").strip() for cell in self.sheet[1]]
        self.header_map = {header: index + 1 for index, header in enumerate(self.headers)}

    def preview(self, command):
        matches = self.matching_row_numbers(command.conditions)
        if command.action in {"sum_column", "count_by_group"} and not matches:
            matches = list(range(2, self.sheet.max_row + 1))
        return matches

    def apply(self, command):
        matches = self.preview(command)
        if command.action == "highlight_rows":
            self.highlight_rows(matches)
        elif command.action == "filter_rows_to_sheet":
            self.copy_rows_to_sheet(command.new_sheet, matches)
        elif command.action == "move_rows_to_sheet":
            self.copy_rows_to_sheet(command.new_sheet, matches)
            self.delete_rows(matches)
        elif command.action == "sum_column":
            self.create_summary_sheet(command, matches, mode="sum")
        elif command.action == "count_by_group":
            self.create_summary_sheet(command, matches, mode="count")
        elif command.action == "format_report":
            self.format_report()
        else:
            raise ValueError(f"Unsupported action: {command.action}")

        output_path = self.edited_path()
        self.workbook.save(output_path)
        return output_path, len(matches)

    def matching_row_numbers(self, conditions):
        matches = []
        for row_number in range(2, self.sheet.max_row + 1):
            if all(self.row_matches(row_number, condition) for condition in conditions):
                matches.append(row_number)
        return matches

    def row_matches(self, row_number, condition):
        column_index = self.header_map.get(condition["column"])
        if not column_index:
            return False
        value = self.sheet.cell(row_number, column_index).value
        text = str(value or "").strip().lower()
        operator = condition["operator"]
        expected = condition["value"]

        if operator == "is_blank_or_missing":
            return text in {"", "missing", "not submitted", "no", "n/a", "none"}
        if operator == "is_not_blank":
            return bool(text)
        if operator == "contains":
            return str(expected).lower() in text
        if operator == "contains_any":
            return any(item in text for item in expected)
        if operator == "greater_than":
            try:
                return float(str(value).replace("$", "").replace(",", "")) > float(expected)
            except (TypeError, ValueError):
                return False
        if operator == "is_blank_or_zero_or_no":
            if text in {"", "no", "not registered", "none", "n/a"}:
                return True
            try:
                return float(text) == 0
            except ValueError:
                return False
        return False

    def highlight_rows(self, row_numbers):
        fill = PatternFill(fill_type="solid", fgColor="FFF2A8")
        for row_number in row_numbers:
            for cell in self.sheet[row_number]:
                cell.fill = fill

    def copy_rows_to_sheet(self, name, row_numbers):
        target = self.replace_sheet(name)
        target.append(self.headers)
        for row_number in row_numbers:
            target.append([cell.value for cell in self.sheet[row_number]])
        self.format_sheet(target)

    def delete_rows(self, row_numbers):
        for row_number in sorted(row_numbers, reverse=True):
            self.sheet.delete_rows(row_number)

    def create_summary_sheet(self, command, row_numbers, mode):
        target = self.replace_sheet(command.new_sheet)
        target_column_index = self.header_map.get(command.target_column)
        group_column_index = self.header_map.get(command.group_by)
        summary = {}

        for row_number in row_numbers:
            group = self.sheet.cell(row_number, group_column_index).value if group_column_index else "All Rows"
            group = str(group or "Blank")
            if mode == "sum":
                value = self.sheet.cell(row_number, target_column_index).value
                try:
                    summary[group] = summary.get(group, 0) + float(str(value).replace("$", "").replace(",", ""))
                except (TypeError, ValueError):
                    summary[group] = summary.get(group, 0)
            else:
                summary[group] = summary.get(group, 0) + 1

        target.append([command.group_by or "Group", "Total" if mode == "sum" else "Count"])
        for group, value in sorted(summary.items()):
            target.append([group, value])
        self.format_sheet(target)

    def format_report(self):
        self.format_sheet(self.sheet)

    def format_sheet(self, sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 42)

    def replace_sheet(self, name):
        name = name[:31]
        if name in self.workbook.sheetnames:
            del self.workbook[name]
        return self.workbook.create_sheet(name)

    def edited_path(self):
        return self.path.with_name(f"{self.path.stem}_edited{self.path.suffix}")


class StudentDatabase:
    def __init__(self, path: Path):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.setup()

    def setup(self):
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                college TEXT NOT NULL,
                major TEXT NOT NULL,
                status TEXT NOT NULL,
                location TEXT NOT NULL,
                advisor TEXT NOT NULL,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        count = self.conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        if count == 0:
            self.conn.executemany(
                """
                INSERT INTO students
                    (student_id, name, college, major, status, location, advisor)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                SAMPLE_STUDENTS,
            )
        self.conn.commit()

    def add_student(self, student):
        self.conn.execute(
            """
            INSERT INTO students
                (student_id, name, college, major, status, location, advisor)
            VALUES (:student_id, :name, :college, :major, :status, :location, :advisor)
            """,
            student,
        )
        self.conn.commit()

    def save_student(self, student, original_student_id=None):
        if original_student_id:
            result = self.conn.execute(
                """
                UPDATE students
                SET student_id = :student_id,
                    name = :name,
                    college = :college,
                    major = :major,
                    status = :status,
                    location = :location,
                    advisor = :advisor,
                    updated_at = CURRENT_TIMESTAMP
                WHERE student_id = :original_student_id
                """,
                {**student, "original_student_id": original_student_id},
            )
            self.conn.commit()
            return result.rowcount

        self.add_student(student)
        return 1

    def update_status(self, student_id, status):
        result = self.conn.execute(
            """
            UPDATE students
            SET status = ?, updated_at = CURRENT_TIMESTAMP
            WHERE student_id = ?
            """,
            (status, student_id),
        )
        self.conn.commit()
        return result.rowcount

    def delete_student(self, student_id):
        result = self.conn.execute("DELETE FROM students WHERE student_id = ?", (student_id,))
        self.conn.commit()
        return result.rowcount

    def all_students(self):
        return self.conn.execute(
            """
            SELECT student_id, name, college, major, status, location, advisor
            FROM students
            ORDER BY college, name
            """
        ).fetchall()

    def search_students(self, college=None, status=None, location=None, keyword=None):
        clauses = []
        params = []
        if college:
            clauses.append("LOWER(college) = LOWER(?)")
            params.append(college)
        if status:
            clauses.append("LOWER(status) = LOWER(?)")
            params.append(status)
        if location:
            clauses.append("LOWER(location) = LOWER(?)")
            params.append(location)
        if keyword:
            clauses.append(
                """
                (
                    LOWER(student_id) LIKE LOWER(?)
                    OR LOWER(name) LIKE LOWER(?)
                    OR LOWER(college) LIKE LOWER(?)
                    OR LOWER(major) LIKE LOWER(?)
                    OR LOWER(status) LIKE LOWER(?)
                    OR LOWER(location) LIKE LOWER(?)
                    OR LOWER(advisor) LIKE LOWER(?)
                )
                """
            )
            term = f"%{keyword}%"
            params.extend([term] * 7)

        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        return self.conn.execute(
            f"""
            SELECT student_id, name, college, major, status, location, advisor
            FROM students
            {where}
            ORDER BY name
            """,
            params,
        ).fetchall()

    def count_students(self, college=None, status=None, location=None):
        return len(self.search_students(college=college, status=status, location=location))

    def find_student(self, student_id):
        return self.conn.execute(
            """
            SELECT student_id, name, college, major, status, location, advisor
            FROM students
            WHERE student_id = ?
            """,
            (student_id,),
        ).fetchone()

    def export_csv(self, path):
        rows = self.all_students()
        with open(path, "w", newline="", encoding="utf-8") as output:
            writer = csv.writer(output)
            writer.writerow(["student_id", "name", "college", "major", "status", "location", "advisor"])
            for row in rows:
                writer.writerow([row[key] for key in row.keys()])

    def import_csv(self, path):
        required = {"student_id", "name", "college", "major", "status", "location", "advisor"}
        imported = 0
        updated = 0

        with open(path, newline="", encoding="utf-8-sig") as source:
            reader = csv.DictReader(source)
            if not reader.fieldnames:
                raise ValueError("The CSV file has no header row.")

            normalized = {name.strip().lower(): name for name in reader.fieldnames}
            missing = sorted(required - set(normalized.keys()))
            if missing:
                raise ValueError(f"Missing columns: {', '.join(missing)}")

            for row in reader:
                student = {
                    key: row[normalized[key]].strip()
                    for key in ["student_id", "name", "college", "major", "status", "location", "advisor"]
                }
                if not all(student.values()):
                    continue

                exists = self.find_student(student["student_id"])
                self.save_student(student, original_student_id=student["student_id"] if exists else None)
                if exists:
                    updated += 1
                else:
                    imported += 1

        return imported, updated


class RuleBasedAssistant:
    def __init__(self, database: StudentDatabase):
        self.database = database
        self.last_college = None
        self.last_status = None

    def respond(self, message):
        text = message.strip()
        lowered = text.lower()

        if not text:
            return "Type a question or command about student enrollment."

        college = self.extract_college(lowered)
        status = self.extract_status(lowered)
        student_id = self.extract_student_id(lowered)

        if college:
            self.last_college = college
        if status:
            self.last_status = status

        if "help" in lowered:
            return self.help_text()

        if "where" in lowered and student_id:
            return self.describe_student(student_id)

        if ("student" in lowered or "id" in lowered) and student_id and not any(
            word in lowered for word in ["update", "change", "set"]
        ):
            return self.describe_student(student_id)

        update_match = re.search(r"(?:update|change|set)\s+(?:student\s+)?(\d+).*\b(enrolled|inactive|withdrawn|graduated)\b", lowered)
        if update_match:
            return self.update_status(update_match.group(1), update_match.group(2))

        if "report" in lowered or "summary" in lowered:
            return self.report(college or self.last_college)

        if "how many" in lowered or "count" in lowered or "number of" in lowered:
            return self.count(college=college or self.last_college, status=status, location=self.extract_location(lowered))

        if "list" in lowered or "show" in lowered or "find" in lowered:
            return self.list_students(college=college or self.last_college, status=status, location=self.extract_location(lowered))

        if college or status:
            return self.list_students(college=college or self.last_college, status=status)

        return (
            "I can help with enrollment counts, student lookup, college lists, status updates, and reports. "
            "Try: 'How many enrolled students are in Engineering?'"
        )

    def extract_college(self, lowered):
        for college in COLLEGES:
            if college.lower() in lowered:
                return college
        return None

    def extract_status(self, lowered):
        for status in STATUSES:
            if status.lower() in lowered:
                return status
        return None

    def extract_location(self, lowered):
        match = re.search(r"(main campus|downtown campus|health campus|online)", lowered)
        return match.group(1).title() if match else None

    def extract_student_id(self, lowered):
        match = re.search(r"\b\d{3,}\b", lowered)
        return match.group(0) if match else None

    def describe_student(self, student_id):
        student = self.database.find_student(student_id)
        if not student:
            return f"I could not find student ID {student_id}."
        return (
            f"{student['name']} ({student['student_id']}) is {student['status']} in "
            f"{student['college']}, majoring in {student['major']}. Location: "
            f"{student['location']}. Advisor: {student['advisor']}."
        )

    def update_status(self, student_id, status_text):
        status = status_text.title()
        changed = self.database.update_status(student_id, status)
        if not changed:
            return f"I could not find student ID {student_id}, so no status was changed."
        return f"Student {student_id} is now marked as {status}."

    def count(self, college=None, status=None, location=None):
        count = self.database.count_students(college=college, status=status, location=location)
        filters = self.describe_filters(college=college, status=status, location=location)
        return f"There are {count} student(s){filters}."

    def list_students(self, college=None, status=None, location=None):
        rows = self.database.search_students(college=college, status=status, location=location)
        filters = self.describe_filters(college=college, status=status, location=location)
        if not rows:
            return f"I found no student records{filters}."

        lines = [f"I found {len(rows)} student record(s){filters}:"]
        for row in rows[:20]:
            lines.append(
                f"- {row['student_id']} | {row['name']} | {row['college']} | "
                f"{row['major']} | {row['status']} | {row['location']}"
            )
        if len(rows) > 20:
            lines.append(f"...and {len(rows) - 20} more.")
        return "\n".join(lines)

    def report(self, college=None):
        rows = self.database.search_students(college=college)
        if not rows:
            return "I found no records for that report."

        status_counts = {status: 0 for status in STATUSES}
        for row in rows:
            status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1

        title = f"{college} report" if college else "All colleges report"
        lines = [title, f"Total students: {len(rows)}"]
        for status, count in status_counts.items():
            lines.append(f"{status}: {count}")
        return "\n".join(lines)

    def describe_filters(self, college=None, status=None, location=None):
        filters = []
        if status:
            filters.append(status.lower())
        if college:
            filters.append(f"in {college}")
        if location:
            filters.append(f"at {location}")
        return " " + " ".join(filters) if filters else ""

    def help_text(self):
        return (
            "Examples I understand:\n"
            "- How many enrolled students are in Engineering?\n"
            "- Show inactive students in Nursing\n"
            "- Where is student 1001?\n"
            "- Generate a report for Business\n"
            "- Update student 1003 to enrolled"
        )


class StudentAssistantApp(TkBase):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1500x760")
        self.minsize(1180, 620)

        self.database = StudentDatabase(DB_PATH)
        self.assistant = RuleBasedAssistant(self.database)

        self.workbook = None
        self.workbook_filter_var = tk.StringVar()
        self.command_parser = SpreadsheetCommandParser()
        self.pending_command = None
        self.action_log = []

        self.build_ui()
        self.add_bot_message(
            "Offline assistant ready. Type 'help' for examples, or ask about enrollment."
        )

    def build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        self.main_panes = tk.PanedWindow(
            self,
            orient=tk.HORIZONTAL,
            sashwidth=8,
            sashrelief=tk.RAISED,
            showhandle=True,
            borderwidth=0,
        )
        self.main_panes.grid(row=0, column=0, sticky="nsew")

        chat_frame = ttk.Frame(self.main_panes, padding=12)
        chat_frame.rowconfigure(1, weight=1)
        chat_frame.columnconfigure(0, weight=1)

        ttk.Label(chat_frame, text="Dean Assistant", font=("Helvetica", 18, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 8)
        )

        self.chat_log = tk.Text(chat_frame, wrap="word", state="disabled", height=18)
        self.chat_log.grid(row=1, column=0, sticky="nsew")
        self.chat_log.tag_configure("user", foreground="#124b8f", spacing3=8)
        self.chat_log.tag_configure("bot", foreground="#1f5c3b", spacing3=8)

        input_frame = ttk.Frame(chat_frame)
        input_frame.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        input_frame.columnconfigure(0, weight=1)

        self.message_entry = ttk.Entry(input_frame)
        self.message_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self.message_entry.bind("<Return>", self.send_message)

        ttk.Button(input_frame, text="Send", command=self.send_message).grid(row=0, column=1)

        workbook_frame = ttk.Frame(self.main_panes, padding=12)
        workbook_frame.rowconfigure(3, weight=1)
        workbook_frame.columnconfigure(0, weight=1)

        ttk.Label(workbook_frame, text="Excel Workbook", font=("Helvetica", 18, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 8)
        )
        self.build_workbook_view(workbook_frame)

        self.main_panes.add(chat_frame, minsize=260, width=330)
        self.main_panes.add(workbook_frame, minsize=620, width=1050)

        self.after(100, self.set_default_pane_sizes)

    def set_default_pane_sizes(self):
        chat_width = 320
        self.main_panes.sash_place(0, chat_width, 0)

    def expand_workbook_pane(self):
        chat_width = 260
        self.main_panes.sash_place(0, chat_width, 0)

    def build_workbook_view(self, parent):
        controls = ttk.Frame(parent)
        controls.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        controls.columnconfigure(3, weight=1)

        ttk.Button(controls, text="Open XLSX", command=self.open_xlsx).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(controls, text="Open Original", command=self.open_original_workbook).grid(row=0, column=1, padx=(0, 8))
        ttk.Label(controls, text="Find").grid(row=0, column=2, sticky="e", padx=(0, 6))
        find_entry = ttk.Entry(controls, textvariable=self.workbook_filter_var)
        find_entry.grid(row=0, column=3, sticky="ew", padx=(0, 8))
        find_entry.bind("<Return>", lambda event: self.highlight_workbook_terms([self.workbook_filter_var.get()]))
        ttk.Button(controls, text="Highlight", command=lambda: self.highlight_workbook_terms([self.workbook_filter_var.get()])).grid(
            row=0, column=4
        )
        ttk.Button(controls, text="Wider Sheet", command=self.expand_workbook_pane).grid(row=0, column=5, padx=(8, 0))
        ttk.Button(controls, text="Reset Layout", command=self.set_default_pane_sizes).grid(row=0, column=6, padx=(8, 0))

        self.drop_label = ttk.Label(
            parent,
            text="Open an .xlsx file here. Drag-and-drop works when tkinterdnd2 is installed.",
            relief="groove",
            padding=12,
            anchor="center",
        )
        self.drop_label.grid(row=2, column=0, sticky="ew", pady=(0, 8))

        if DND_FILES:
            self.drop_label.drop_target_register(DND_FILES)
            self.drop_label.dnd_bind("<<Drop>>", self.handle_file_drop)

        self.workbook_table = ttk.Treeview(parent, show="headings")
        workbook_scroll_y = ttk.Scrollbar(parent, orient="vertical", command=self.workbook_table.yview)
        workbook_scroll_x = ttk.Scrollbar(parent, orient="horizontal", command=self.workbook_table.xview)
        self.workbook_table.configure(yscrollcommand=workbook_scroll_y.set, xscrollcommand=workbook_scroll_x.set)
        self.workbook_table.grid(row=3, column=0, sticky="nsew")
        workbook_scroll_y.grid(row=3, column=1, sticky="ns")
        workbook_scroll_x.grid(row=4, column=0, sticky="ew")
        self.workbook_table.tag_configure("match", background="#fff2a8")

    def send_message(self, event=None):
        message = self.message_entry.get().strip()
        if not message:
            return
        self.message_entry.delete(0, tk.END)
        self.add_user_message(message)
        response = self.handle_pending_confirmation(message)
        if not response:
            response = self.handle_workbook_command(message) or self.assistant.respond(message)
        self.add_bot_message(response)

    def add_user_message(self, message):
        self.append_chat(f"You: {message}\n", "user")

    def add_bot_message(self, message):
        self.append_chat(f"Assistant: {message}\n", "bot")

    def append_chat(self, message, tag):
        self.chat_log.configure(state="normal")
        self.chat_log.insert(tk.END, message, tag)
        self.chat_log.configure(state="disabled")
        self.chat_log.see(tk.END)

    def open_xlsx(self):
        path = filedialog.askopenfilename(
            title="Open Excel Workbook",
            filetypes=[("Excel workbooks", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.load_xlsx(path)

    def handle_file_drop(self, event):
        paths = self.tk.splitlist(event.data)
        if paths:
            self.load_xlsx(paths[0])

    def load_xlsx(self, path):
        try:
            self.workbook = XlsxReader.read_first_sheet(path)
        except (OSError, ValueError, zipfile.BadZipFile, ET.ParseError) as error:
            messagebox.showerror("Workbook Error", str(error))
            return

        self.display_workbook()
        self.add_bot_message(self.workbook.summary())

    def open_original_workbook(self):
        if not self.workbook:
            messagebox.showinfo("No Workbook", "Open an .xlsx workbook first.")
            return

        path = str(self.workbook.path)
        try:
            if sys.platform == "darwin":
                subprocess.Popen(["open", path])
            elif os.name == "nt":
                os.startfile(path)
            else:
                subprocess.Popen(["xdg-open", path])
        except OSError as error:
            messagebox.showerror("Open Failed", str(error))

    def display_workbook(self):
        if not self.workbook:
            return

        self.workbook_table.delete(*self.workbook_table.get_children())
        columns = [f"c{index}" for index in range(len(self.workbook.headers))]
        self.workbook_table.configure(columns=columns)

        for column_id, header in zip(columns, self.workbook.headers):
            self.workbook_table.heading(column_id, text=header)
            self.workbook_table.column(column_id, width=max(110, min(len(header) * 10 + 35, 220)), anchor="w")

        for index, row in enumerate(self.workbook.rows):
            self.workbook_table.insert("", tk.END, iid=str(index), values=row)

    def handle_workbook_command(self, message):
        lowered = message.lower().strip()
        workbook_words = [
            "workbook",
            "xlsx",
            "excel",
            "sheet",
            "headers",
            "rows",
            "extract",
            "highlight",
            "sum",
            "total",
            "count",
            "move",
            "format",
            "missing",
            "balance",
            "fafsa",
            "registered",
        ]
        if not any(word in lowered for word in workbook_words):
            return None

        if not self.workbook:
            return "Open an .xlsx workbook first, then I can summarize, search, highlight, or extract from it."

        if "summary" in lowered or "summarize" in lowered or "what is shown" in lowered:
            return self.workbook.summary()

        if "header" in lowered or "column" in lowered:
            return "Workbook headers:\n- " + "\n- ".join(self.workbook.headers)

        if "extract" in lowered:
            return self.extract_from_workbook(message)

        command = self.command_parser.parse(message, self.workbook.headers)
        if command:
            return self.preview_command(command)

        if "highlight" in lowered or "find" in lowered or "show" in lowered:
            terms = self.terms_after_first_match(
                message,
                ["highlight", "find", "show", "search"],
            )
            if not terms:
                return "Tell me what to highlight in the workbook."
            match_count = self.highlight_workbook_terms(terms)
            return f"Highlighted {match_count} workbook row(s) matching: {', '.join(terms)}."

        return None

    def preview_command(self, command):
        try:
            engine = SpreadsheetActionEngine(self.workbook.path)
            matches = engine.preview(command)
        except RuntimeError as error:
            return str(error)
        except (OSError, ValueError) as error:
            return f"I could not validate that action: {error}"

        if command.conditions and not matches:
            return "I understood the action, but I found 0 matching rows. No change is queued."

        self.pending_command = command
        condition_text = self.describe_conditions(command.conditions)
        lines = [
            f"Planned action: {command.description}",
            f"Action type: {command.action}",
            f"Matching rows: {len(matches)}",
        ]
        if condition_text:
            lines.append(f"Conditions: {condition_text}")
        if command.new_sheet:
            lines.append(f"New sheet: {command.new_sheet}")
        lines.append("Type 'confirm' to apply this to a new edited workbook, or 'cancel' to stop.")
        return "\n".join(lines)

    def handle_pending_confirmation(self, message):
        if not self.pending_command:
            return None

        lowered = message.lower().strip()
        if lowered in {"cancel", "no", "stop"}:
            self.pending_command = None
            return "Canceled the pending spreadsheet action."

        if lowered not in {"confirm", "yes", "apply"}:
            return "A spreadsheet action is waiting. Type 'confirm' to apply it or 'cancel' to stop."

        command = self.pending_command
        self.pending_command = None
        try:
            engine = SpreadsheetActionEngine(self.workbook.path)
            output_path, changed_rows = engine.apply(command)
        except RuntimeError as error:
            return str(error)
        except (OSError, ValueError) as error:
            return f"I could not apply that action: {error}"

        self.action_log.append(
            {
                "action": command.action,
                "description": command.description,
                "rows": changed_rows,
                "output": str(output_path),
            }
        )
        self.load_xlsx(output_path)
        return f"Applied {command.action} to {changed_rows} row(s). Saved edited workbook: {output_path}"

    def describe_conditions(self, conditions):
        parts = []
        for condition in conditions:
            parts.append(f"{condition['column']} {condition['operator']} {condition['value']}")
        return "; ".join(parts)

    def extract_from_workbook(self, message):
        lowered = message.lower()
        after_extract = re.sub(r"^.*?\bextract\b", "", message, flags=re.IGNORECASE).strip()
        filter_text, requested_text = self.split_extract_request(after_extract)
        filter_terms = self.clean_filter_terms(filter_text)
        requested_columns = self.clean_requested_columns(requested_text)

        matches = self.workbook.matching_rows(filter_terms)
        column_indexes = self.workbook.matching_columns(requested_columns)
        self.highlight_workbook_rows([index for index, _row in matches])

        if not matches:
            return f"I found no workbook rows matching: {', '.join(filter_terms) or lowered}."

        lines = [
            f"Extracted {len(matches)} workbook row(s).",
            "Columns: " + ", ".join(self.workbook.headers[index] for index in column_indexes),
        ]
        for row_index, row in matches[:20]:
            values = []
            for column_index in column_indexes:
                values.append(f"{self.workbook.headers[column_index]}: {row[column_index]}")
            lines.append(f"- Row {row_index + 2}: " + " | ".join(values))
        if len(matches) > 20:
            lines.append(f"...and {len(matches) - 20} more matching row(s).")
        return "\n".join(lines)

    def split_extract_request(self, text):
        parts = re.split(r"\s+\bby\b\s+", text, maxsplit=1, flags=re.IGNORECASE)
        filter_text = parts[0] if parts else text
        requested_text = parts[1] if len(parts) > 1 else ""
        return filter_text, requested_text

    def clean_filter_terms(self, text):
        text = re.sub(r"\b(all|students?|records?|rows?|from|with|where|that|are|is|the)\b", " ", text, flags=re.IGNORECASE)
        terms = re.findall(r'"([^"]+)"|\'([^\']+)\'|([A-Za-z0-9_-]+)', text)
        return [next(part for part in group if part).strip() for group in terms]

    def clean_requested_columns(self, text):
        if not text:
            return []
        text = re.sub(r"\b(and|the|columns?|fields?)\b", " ", text, flags=re.IGNORECASE)
        return [part.strip() for part in re.split(r"[,/]|  +", text) if part.strip()]

    def terms_after_first_match(self, message, words):
        for word in words:
            match = re.search(rf"\b{word}\b\s*(.*)", message, flags=re.IGNORECASE)
            if match:
                text = match.group(1).strip()
                text = re.sub(r"\b(in|from|the|workbook|sheet|excel|xlsx)\b", " ", text, flags=re.IGNORECASE)
                return [term for term in re.split(r"\s+", text.strip()) if term]
        return []

    def highlight_workbook_terms(self, terms):
        if not self.workbook:
            return 0

        clean_terms = [term.strip() for term in terms if term.strip()]
        matches = self.workbook.matching_rows(clean_terms)
        self.highlight_workbook_rows([index for index, _row in matches])
        return len(matches)

    def highlight_workbook_rows(self, row_indexes):
        wanted = {str(index) for index in row_indexes}
        for item in self.workbook_table.get_children():
            self.workbook_table.item(item, tags=("match",) if item in wanted else ())

        if wanted:
            first = min(wanted, key=lambda value: int(value))
            self.workbook_table.selection_set(first)
            self.workbook_table.see(first)

def main():
    app = StudentAssistantApp()
    app.mainloop()


if __name__ == "__main__":
    main()
