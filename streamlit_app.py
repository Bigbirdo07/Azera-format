import json
import re
import sqlite3
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DB_PATH = Path(__file__).with_name("assistant_logs.db")
SUPPORTED_ACTIONS = {
    "filter_rows",
    "highlight_rows",
    "sum_column",
    "count_by_group",
    "move_rows_to_sheet",
    "detect_missing_values",
    "format_report",
}

SYNONYMS = {
    "Balance Due": ["amount due", "tuition balance", "unpaid balance", "owes money", "owe money", "still owe", "balance"],
    "FAFSA Status": ["fafsa", "financial aid status", "aid documents", "financial aid", "aid status"],
    "Enrollment Status": ["active", "inactive", "withdrawn", "registered", "not registered", "enrollment status", "status"],
    "Program": ["major", "department", "degree program", "program", "discipline"],
    "Student ID": ["student number", "banner id", "university id", "student id", "id"],
    "Standing": ["standing", "academic standing", "good standing", "bad standing"],
}


def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS request_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                filename TEXT NOT NULL,
                original_request TEXT NOT NULL,
                generated_json TEXT NOT NULL,
                succeeded INTEGER NOT NULL,
                message TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS corrections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                filename TEXT NOT NULL,
                original_request TEXT NOT NULL,
                generated_json TEXT NOT NULL,
                correct INTEGER NOT NULL,
                corrected_action TEXT NOT NULL
            )
            """
        )


def log_request(filename, original_request, command, succeeded, message):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO request_logs
                (timestamp, filename, original_request, generated_json, succeeded, message)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                filename,
                original_request,
                json.dumps(command, indent=2),
                1 if succeeded else 0,
                message,
            ),
        )


def log_correction(filename, original_request, command, correct, corrected_action):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO corrections
                (timestamp, filename, original_request, generated_json, correct, corrected_action)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                filename,
                original_request,
                json.dumps(command, indent=2),
                1 if correct else 0,
                corrected_action,
            ),
        )


def normalize(text):
    return re.sub(r"[^a-z0-9]+", " ", str(text).lower()).strip()


def resolve_column(phrase, columns):
    phrase_norm = normalize(phrase)
    normalized_columns = {column: normalize(column) for column in columns}

    for column, normalized in normalized_columns.items():
        if normalized and (normalized in phrase_norm or phrase_norm in normalized):
            return column

    for canonical, aliases in SYNONYMS.items():
        alias_hit = normalize(canonical) in phrase_norm or any(normalize(alias) in phrase_norm for alias in aliases)
        if not alias_hit:
            continue
        for column, normalized in normalized_columns.items():
            names = [canonical] + aliases
            if any(normalize(name) == normalized or normalize(name) in normalized or normalized in normalize(name) for name in names):
                return column
    return ""


def text_after_marker(text, markers):
    for marker in markers:
        match = re.search(rf"\b{re.escape(marker)}\b\s+(.+)", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return ""


def text_between(text, start_words, end_words):
    starts = "|".join(re.escape(word) for word in start_words)
    ends = "|".join(re.escape(word) for word in end_words)
    match = re.search(rf"\b(?:{starts})\b\s+(.+?)(?:\s+\b(?:{ends})\b\s+|$)", text, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def parse_conditions(request, columns):
    text = normalize(request)
    conditions = []

    balance_col = resolve_column("Balance Due", columns)
    fafsa_col = resolve_column("FAFSA Status", columns)
    enrollment_col = resolve_column("Enrollment Status", columns)
    standing_col = resolve_column("Standing", columns)

    if balance_col and any(term in text for term in ["owe", "owes", "money", "balance", "amount due", "unpaid", "tuition"]):
        conditions.append({"column": balance_col, "operator": "greater_than", "value": 0})

    if fafsa_col and "fafsa" in text:
        if any(term in text for term in ["missing", "blank", "not submitted", "incomplete", "aid documents"]):
            conditions.append({"column": fafsa_col, "operator": "is_blank_or_missing", "value": ""})
        else:
            conditions.append({"column": fafsa_col, "operator": "is_not_blank", "value": ""})

    if enrollment_col:
        if "not registered" in text or "didn t register" in text or "did not register" in text:
            conditions.append({"column": enrollment_col, "operator": "not_contains", "value": "registered"})
        elif "registered" in text:
            conditions.append({"column": enrollment_col, "operator": "contains", "value": "registered"})
        elif "inactive" in text:
            conditions.append({"column": enrollment_col, "operator": "contains", "value": "inactive"})
        elif "withdrawn" in text:
            conditions.append({"column": enrollment_col, "operator": "contains", "value": "withdrawn"})
        elif "active" in text or "enrolled" in text:
            conditions.append({"column": enrollment_col, "operator": "contains_any", "value": ["active", "enrolled"]})

    if "missing" in text or "blank" in text:
        target_col = resolve_column(request, columns)
        if target_col and not any(condition["column"] == target_col for condition in conditions):
            conditions.append({"column": target_col, "operator": "is_blank_or_missing", "value": ""})

    if standing_col and "bad standing" in text:
        conditions.append({"column": standing_col, "operator": "contains", "value": "bad standing"})
    elif standing_col and "good standing" in text:
        conditions.append({"column": standing_col, "operator": "contains", "value": "good standing"})

    return conditions


def parse_request_to_command(request, columns):
    text = normalize(request)

    if "format" in text and ("report" in text or "sheet" in text or "workbook" in text):
        return {"action": "format_report"}

    if "detect" in text and ("missing" in text or "blank" in text):
        target_col = resolve_column(request, columns)
        return {"action": "detect_missing_values", "column": target_col or None}

    if "move" in text and ("sheet" in text or "tab" in text):
        return {
            "action": "move_rows_to_sheet",
            "conditions": parse_conditions(request, columns),
            "new_sheet": suggested_sheet_name(request),
        }

    if "sum" in text or "total" in text:
        target_text = text_between(request, ["sum", "total"], ["by", "per"]) or "Balance Due"
        group_text = text_after_marker(request, ["by", "per"]) or "Program"
        return {
            "action": "sum_column",
            "target_column": resolve_column(target_text, columns),
            "group_by": resolve_column(group_text, columns),
            "new_sheet": "Summary",
        }

    if "count" in text or "how many" in text:
        group_text = text_after_marker(request, ["by", "per"]) or "Program"
        return {
            "action": "count_by_group",
            "conditions": parse_conditions(request, columns),
            "group_by": resolve_column(group_text, columns),
            "new_sheet": "Counts",
        }

    if "highlight" in text:
        return {
            "action": "highlight_rows",
            "conditions": parse_conditions(request, columns),
            "format": {"fill_color": "yellow"},
        }

    if any(word in text for word in ["show", "filter", "find", "list", "who"]):
        return {
            "action": "filter_rows",
            "conditions": parse_conditions(request, columns),
            "new_sheet": "Filtered Rows",
        }

    return {"action": "unknown", "raw_request": request}


def suggested_sheet_name(request):
    text = normalize(request)
    if "fafsa" in text:
        return "Missing FAFSA"
    if "balance" in text or "owe" in text:
        return "Outstanding Balance"
    if "inactive" in text:
        return "Inactive Students"
    return "Moved Rows"


def validate_command(command, columns):
    errors = []
    action = command.get("action")
    if action not in SUPPORTED_ACTIONS:
        errors.append(f"Unsupported action: {action}")

    referenced = []
    for condition in command.get("conditions", []):
        referenced.append(condition.get("column"))
    for key in ["target_column", "group_by", "column"]:
        if command.get(key):
            referenced.append(command[key])

    for column in referenced:
        if column and column not in columns:
            errors.append(f"Column does not exist: {column}")

    if action in {"filter_rows", "highlight_rows", "move_rows_to_sheet"} and not command.get("conditions"):
        errors.append("This action needs at least one condition.")
    if action == "sum_column" and not command.get("target_column"):
        errors.append("sum_column needs a target_column.")
    if action == "count_by_group" and not command.get("group_by"):
        errors.append("count_by_group needs a group_by column.")

    return errors


def cell_matches(value, condition):
    text = normalize(value)
    operator = condition["operator"]
    expected = condition.get("value")

    if operator == "greater_than":
        try:
            return float(str(value).replace("$", "").replace(",", "")) > float(expected)
        except (TypeError, ValueError):
            return False
    if operator == "contains":
        return normalize(expected) in text
    if operator == "not_contains":
        return normalize(expected) not in text
    if operator == "contains_any":
        return any(normalize(item) in text for item in expected)
    if operator == "is_blank_or_missing":
        return text in {"", "missing", "not submitted", "incomplete", "n a", "none", "no"}
    if operator == "is_not_blank":
        return bool(text)
    return False


def matching_row_numbers(sheet, header_map, conditions):
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        ok = True
        for condition in conditions:
            column_index = header_map[condition["column"]]
            if not cell_matches(sheet.cell(row_number, column_index).value, condition):
                ok = False
                break
        if ok:
            rows.append(row_number)
    return rows


def copy_row(source_row, target_sheet):
    values = [cell.value for cell in source_row]
    target_sheet.append(values)


def format_sheet(sheet):
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 45)


def replace_sheet(workbook, name):
    safe_name = str(name or "Results")[:31]
    if safe_name in workbook.sheetnames:
        del workbook[safe_name]
    return workbook.create_sheet(safe_name)


def preview_match_count(file_bytes, sheet_name, command):
    workbook = load_workbook(BytesIO(file_bytes))
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}

    action = command["action"]
    if action in {"filter_rows", "highlight_rows", "move_rows_to_sheet", "count_by_group"}:
        conditions = command.get("conditions", [])
        return len(matching_row_numbers(sheet, header_map, conditions)) if conditions else sheet.max_row - 1
    if action == "detect_missing_values":
        target = command.get("column")
        if target:
            return len(
                [
                    row
                    for row in range(2, sheet.max_row + 1)
                    if cell_matches(sheet.cell(row, header_map[target]).value, {"operator": "is_blank_or_missing", "value": ""})
                ]
            )
        return sum(
            1
            for row in range(2, sheet.max_row + 1)
            for column in range(1, sheet.max_column + 1)
            if cell_matches(sheet.cell(row, column).value, {"operator": "is_blank_or_missing", "value": ""})
        )
    return sheet.max_row - 1


def apply_command(file_bytes, sheet_name, command):
    workbook = load_workbook(BytesIO(file_bytes))
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    header_map = {header: index + 1 for index, header in enumerate(headers)}
    action = command["action"]
    affected = 0

    if action == "highlight_rows":
        rows = matching_row_numbers(sheet, header_map, command["conditions"])
        fill = PatternFill(fill_type="solid", fgColor="FFF2A8")
        for row_number in rows:
            for cell in sheet[row_number]:
                cell.fill = fill
        affected = len(rows)

    elif action == "filter_rows":
        rows = matching_row_numbers(sheet, header_map, command["conditions"])
        target = replace_sheet(workbook, command.get("new_sheet", "Filtered Rows"))
        target.append(headers)
        for row_number in rows:
            copy_row(sheet[row_number], target)
        format_sheet(target)
        affected = len(rows)

    elif action == "move_rows_to_sheet":
        rows = matching_row_numbers(sheet, header_map, command["conditions"])
        target = replace_sheet(workbook, command.get("new_sheet", "Moved Rows"))
        target.append(headers)
        for row_number in rows:
            copy_row(sheet[row_number], target)
        for row_number in sorted(rows, reverse=True):
            sheet.delete_rows(row_number)
        format_sheet(target)
        affected = len(rows)

    elif action == "sum_column":
        target_col = header_map[command["target_column"]]
        group_col = header_map.get(command.get("group_by"))
        summary = {}
        for row_number in range(2, sheet.max_row + 1):
            group = sheet.cell(row_number, group_col).value if group_col else "All Rows"
            group = str(group or "Blank")
            value = sheet.cell(row_number, target_col).value
            try:
                summary[group] = summary.get(group, 0) + float(str(value).replace("$", "").replace(",", ""))
            except (TypeError, ValueError):
                summary[group] = summary.get(group, 0)
        target = replace_sheet(workbook, command.get("new_sheet", "Summary"))
        target.append([command.get("group_by") or "Group", f"Sum of {command['target_column']}"])
        for group, value in sorted(summary.items()):
            target.append([group, value])
        format_sheet(target)
        affected = len(summary)

    elif action == "count_by_group":
        rows = matching_row_numbers(sheet, header_map, command.get("conditions", [])) if command.get("conditions") else range(2, sheet.max_row + 1)
        group_col = header_map[command["group_by"]]
        summary = {}
        for row_number in rows:
            group = str(sheet.cell(row_number, group_col).value or "Blank")
            summary[group] = summary.get(group, 0) + 1
        target = replace_sheet(workbook, command.get("new_sheet", "Counts"))
        target.append([command["group_by"], "Count"])
        for group, value in sorted(summary.items()):
            target.append([group, value])
        format_sheet(target)
        affected = len(summary)

    elif action == "detect_missing_values":
        target = replace_sheet(workbook, "Missing Values")
        target.append(["Row", "Column", "Student ID", "Issue"])
        student_col = header_map.get(resolve_column("Student ID", headers))
        columns_to_check = [command["column"]] if command.get("column") else headers
        for row_number in range(2, sheet.max_row + 1):
            for column_name in columns_to_check:
                column_index = header_map[column_name]
                if cell_matches(sheet.cell(row_number, column_index).value, {"operator": "is_blank_or_missing", "value": ""}):
                    student_id = sheet.cell(row_number, student_col).value if student_col else ""
                    target.append([row_number, column_name, student_id, "Missing value"])
                    affected += 1
        format_sheet(target)

    elif action == "format_report":
        format_sheet(sheet)
        affected = sheet.max_row - 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue(), affected


def edited_filename(filename):
    path = Path(filename)
    return f"{path.stem}_edited{path.suffix}"


def read_workbook(uploaded_file):
    bytes_data = uploaded_file.getvalue()
    sheets = pd.read_excel(BytesIO(bytes_data), sheet_name=None, dtype=object)
    return bytes_data, sheets


def show_logs():
    with sqlite3.connect(DB_PATH) as conn:
        logs = pd.read_sql_query("SELECT * FROM request_logs ORDER BY id DESC LIMIT 25", conn)
        corrections = pd.read_sql_query("SELECT * FROM corrections ORDER BY id DESC LIMIT 25", conn)
    st.subheader("Recent Requests")
    st.dataframe(logs, use_container_width=True, hide_index=True)
    st.subheader("Corrections")
    st.dataframe(corrections, use_container_width=True, hide_index=True)


def main():
    st.set_page_config(page_title="University Enrollment Excel Assistant", layout="wide")
    init_db()

    st.title("University Enrollment Excel Assistant")
    st.caption("Offline-style MVP: rule-based parser, validated Excel actions, edited-file export, SQLite logging.")

    tab_work, tab_corrections, tab_logs = st.tabs(["Assistant", "Correction Screen", "Logs"])

    with tab_work:
        uploaded = st.file_uploader("Upload an .xlsx workbook", type=["xlsx"])
        if not uploaded:
            st.info("Upload an Excel workbook to begin.")
            return

        file_bytes, sheets = read_workbook(uploaded)
        sheet_name = st.selectbox("Sheet", list(sheets.keys()))
        df = sheets[sheet_name]
        columns = [str(column).strip() for column in df.columns]

        col_a, col_b = st.columns([1, 2])
        with col_a:
            st.subheader("Detected Columns")
            st.write(columns)
        with col_b:
            st.subheader("Preview")
            st.dataframe(df.head(50), use_container_width=True)

        request = st.text_area(
            "Natural-language request",
            placeholder="Example: Highlight students who still owe money",
            key="request_text",
        )

        if st.button("Generate JSON Command", type="primary"):
            command = parse_request_to_command(request, columns)
            errors = validate_command(command, columns)
            match_count = None
            if not errors:
                match_count = preview_match_count(file_bytes, sheet_name, command)
            else:
                log_request(uploaded.name, request, command, False, "; ".join(errors))
            st.session_state["last_command"] = command
            st.session_state["last_errors"] = errors
            st.session_state["last_request"] = request
            st.session_state["last_filename"] = uploaded.name
            st.session_state["last_sheet"] = sheet_name
            st.session_state["last_file_bytes"] = file_bytes
            st.session_state["last_match_count"] = match_count

        command = st.session_state.get("last_command")
        if command:
            st.subheader("Generated JSON Command")
            st.json(command)

            errors = st.session_state.get("last_errors", [])
            if errors:
                st.error("Validation failed:\n" + "\n".join(f"- {error}" for error in errors))
            else:
                st.success(f"Validation passed. Matching/affected rows preview: {st.session_state.get('last_match_count')}")
                st.warning("This will not overwrite the original workbook. It will create a new edited .xlsx file.")
                if st.button("Confirm and Apply Action"):
                    try:
                        edited_bytes, affected = apply_command(
                            st.session_state["last_file_bytes"],
                            st.session_state["last_sheet"],
                            command,
                        )
                        out_name = edited_filename(st.session_state["last_filename"])
                        st.session_state["edited_bytes"] = edited_bytes
                        st.session_state["edited_name"] = out_name
                        message = f"Action succeeded. Affected rows/items: {affected}"
                        log_request(uploaded.name, st.session_state.get("last_request", ""), command, True, message)
                        st.success(message)
                    except Exception as exc:
                        log_request(uploaded.name, st.session_state.get("last_request", ""), command, False, str(exc))
                        st.error(f"Action failed: {exc}")

        if st.session_state.get("edited_bytes"):
            st.download_button(
                "Download Edited Workbook",
                data=st.session_state["edited_bytes"],
                file_name=st.session_state["edited_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    with tab_corrections:
        st.subheader("Correction Screen")
        command = st.session_state.get("last_command")
        if not command:
            st.info("Generate a command first, then come here to mark whether it understood correctly.")
        else:
            st.write("Original request:")
            st.code(st.session_state.get("last_request", ""))
            st.write("Generated command:")
            st.json(command)
            correct = st.radio("Did the bot understand correctly?", ["Yes", "No"], horizontal=True)
            corrected = st.text_area(
                "If incorrect, provide the correct action or JSON command",
                placeholder='{"action": "highlight_rows", "conditions": [...]}',
            )
            if st.button("Save Correction"):
                log_correction(
                    st.session_state.get("last_filename", ""),
                    st.session_state.get("last_request", ""),
                    command,
                    correct == "Yes",
                    corrected,
                )
                st.success("Correction saved to SQLite.")

    with tab_logs:
        show_logs()


if __name__ == "__main__":
    main()
